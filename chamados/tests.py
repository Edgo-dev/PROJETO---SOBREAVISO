"""Testes do app chamados."""

from io import BytesIO
from pathlib import Path

from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, transaction
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook

from .models import (
    Ativo,
    AtualizacaoChamado,
    Chamado,
    Fornecedor,
    Obra,
    StatusChamado,
    TipoEventoAtualizacao,
)


class _LoginClienteMixin:
    """Autentica um usuário operador antes de cada teste de view interna.

    O ``LoginRequiredMiddleware`` exige autenticação em todas as telas
    internas para identificar quem está operando o sistema. Testes que
    exercitam essas views via ``self.client`` precisam logar — caso
    contrário recebem redirect 302 para ``/login/`` e qualquer
    ``assertContains``/``response.context[...]`` falha.

    Como cada ``TestCase`` roda em sua própria transação, o usuário
    criado aqui é descartado entre testes; podemos reutilizar o mesmo
    username sem risco de colisão entre classes.
    """

    LOGIN_USERNAME = "operador_teste"
    LOGIN_PASSWORD = "senhaforte123"

    def setUp(self):
        super().setUp()
        from django.contrib.auth import get_user_model

        User = get_user_model()
        self.user = User.objects.create_user(
            username=self.LOGIN_USERNAME,
            password=self.LOGIN_PASSWORD,
        )
        self.client.force_login(self.user)


class HomeViewTests(_LoginClienteMixin, TestCase):
    def test_home_retorna_status_200(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertEqual(response.status_code, 200)

    def test_home_contem_titulo(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, "Sistema de Sobreaviso")


class FornecedorModelTests(TestCase):
    def test_criar_fornecedor(self):
        # Contagem RELATIVA: a migration seed `0004_seed_fornecedores_padrao`
        # já popula a tabela, então a contagem absoluta não é estável.
        antes_count = Fornecedor.objects.count()
        fornecedor = Fornecedor.objects.create(
            nome="Fornecedor Alpha",
            contato="Joao",
            telefone="11999999999",
            email="contato@alpha.com",
        )
        self.assertEqual(Fornecedor.objects.count(), antes_count + 1)
        criado = Fornecedor.objects.get(nome="Fornecedor Alpha")
        self.assertEqual(criado.pk, fornecedor.pk)
        self.assertEqual(criado.contato, "Joao")
        self.assertEqual(criado.telefone, "11999999999")
        self.assertEqual(criado.email, "contato@alpha.com")
        self.assertTrue(fornecedor.ativo)
        self.assertIsNotNone(fornecedor.criado_em)
        self.assertIsNotNone(fornecedor.atualizado_em)

    def test_fornecedor_str(self):
        fornecedor = Fornecedor.objects.create(nome="Fornecedor Beta")
        self.assertEqual(str(fornecedor), "Fornecedor Beta")


class FornecedoresViewTests(_LoginClienteMixin, TestCase):
    def _dados_fornecedor(self, **overrides):
        dados = {
            "nome": "Fornecedor Operacional",
            "contato": "Maria",
            "telefone": "11999990000",
            "email": "contato@fornecedor.com",
            "ativo": "on",
        }
        dados.update(overrides)
        return dados

    def test_fornecedores_list_status_200(self):
        response = self.client.get(reverse("chamados:fornecedores_list"))
        self.assertEqual(response.status_code, 200)

    def test_fornecedores_list_lista_fornecedor_existente(self):
        Fornecedor.objects.create(nome="Fornecedor Lista")
        response = self.client.get(reverse("chamados:fornecedores_list"))
        self.assertContains(response, "Fornecedor Lista")

    def test_busca_q_encontra_fornecedor_por_nome(self):
        Fornecedor.objects.create(nome="Alpha Manutencao")
        Fornecedor.objects.create(nome="Beta Servicos")
        response = self.client.get(
            reverse("chamados:fornecedores_list"), {"q": "Alpha"}
        )
        self.assertContains(response, "Alpha Manutencao")
        self.assertNotContains(response, "Beta Servicos")

    def test_busca_q_encontra_fornecedor_por_empresa(self):
        Fornecedor.objects.create(nome="Fornecedor A", empresa="TechAtende Eng.")
        Fornecedor.objects.create(nome="Fornecedor B", empresa="OutraCorp")
        response = self.client.get(
            reverse("chamados:fornecedores_list"), {"q": "TechAtende"}
        )
        self.assertContains(response, "Fornecedor A")
        self.assertNotContains(response, "Fornecedor B")

    def test_busca_q_encontra_fornecedor_por_telefone(self):
        Fornecedor.objects.create(nome="Fornecedor Tel A", telefone="1133334444")
        Fornecedor.objects.create(nome="Fornecedor Tel B", telefone="1155556666")
        response = self.client.get(
            reverse("chamados:fornecedores_list"), {"q": "3333"}
        )
        self.assertContains(response, "Fornecedor Tel A")
        self.assertNotContains(response, "Fornecedor Tel B")

    def test_busca_q_encontra_fornecedor_por_email(self):
        Fornecedor.objects.create(nome="Fornecedor Email A", email="alpha@test.com")
        Fornecedor.objects.create(nome="Fornecedor Email B", email="beta@test.com")
        response = self.client.get(
            reverse("chamados:fornecedores_list"), {"q": "alpha@test.com"}
        )
        self.assertContains(response, "Fornecedor Email A")
        self.assertNotContains(response, "Fornecedor Email B")

    def test_filtro_ativo_sim_retorna_apenas_fornecedores_ativos(self):
        Fornecedor.objects.create(nome="Fornecedor Ativo", ativo=True)
        Fornecedor.objects.create(nome="Fornecedor Inativo", ativo=False)
        response = self.client.get(
            reverse("chamados:fornecedores_list"), {"ativo": "sim"}
        )
        self.assertContains(response, "Fornecedor Ativo")
        self.assertNotContains(response, "Fornecedor Inativo")

    def test_filtro_ativo_nao_retorna_apenas_fornecedores_inativos(self):
        Fornecedor.objects.create(nome="Fornecedor Ativo", ativo=True)
        Fornecedor.objects.create(nome="Fornecedor Inativo", ativo=False)
        response = self.client.get(
            reverse("chamados:fornecedores_list"), {"ativo": "nao"}
        )
        self.assertContains(response, "Fornecedor Inativo")
        self.assertNotContains(response, "Fornecedor Ativo")

    def test_fornecedor_create_status_200(self):
        response = self.client.get(reverse("chamados:fornecedor_create"))
        self.assertEqual(response.status_code, 200)

    def test_post_fornecedor_create_cria_fornecedor(self):
        response = self.client.post(
            reverse("chamados:fornecedor_create"), data=self._dados_fornecedor()
        )
        self.assertRedirects(response, reverse("chamados:fornecedores_list"))
        self.assertTrue(
            Fornecedor.objects.filter(nome="Fornecedor Operacional").exists()
        )

    def test_post_fornecedor_create_normaliza_nome(self):
        self.client.post(
            reverse("chamados:fornecedor_create"),
            data=self._dados_fornecedor(nome="  Fornecedor Strip  "),
        )
        self.assertTrue(Fornecedor.objects.filter(nome="Fornecedor Strip").exists())

    def test_post_fornecedor_create_normaliza_email_minusculo(self):
        self.client.post(
            reverse("chamados:fornecedor_create"),
            data=self._dados_fornecedor(email="CONTATO@EXEMPLO.COM"),
        )
        self.assertTrue(
            Fornecedor.objects.filter(email="contato@exemplo.com").exists()
        )

    def test_fornecedor_detail_status_200(self):
        fornecedor = Fornecedor.objects.create(nome="Fornecedor Detalhe")
        response = self.client.get(
            reverse("chamados:fornecedor_detail", args=[fornecedor.pk])
        )
        self.assertEqual(response.status_code, 200)

    def test_fornecedor_detail_exibe_dados(self):
        fornecedor = Fornecedor.objects.create(
            nome="Fornecedor Dados",
            telefone="1100000000",
            email="dados@test.com",
            empresa="ABC Manutenção",
            estados_atendidos="SP, RJ, MG",
        )
        response = self.client.get(
            reverse("chamados:fornecedor_detail", args=[fornecedor.pk])
        )
        self.assertContains(response, "Fornecedor Dados")
        self.assertContains(response, "1100000000")
        self.assertContains(response, "dados@test.com")
        self.assertContains(response, "ABC Manutenção")
        self.assertContains(response, "SP, RJ, MG")

    def test_fornecedor_detail_exibe_quantidade_chamados_vinculados(self):
        fornecedor = Fornecedor.objects.create(nome="Fornecedor Chamados")
        ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-FOR-01",
            nome_site="Site Fornecedor",
            endereco="Rua F, 1",
            cidade="Santos",
            uf="SP",
            regional="Sudeste",
        )
        Chamado.objects.create(
            ativo=ativo,
            fornecedor=fornecedor,
            numero_os="OS-FOR-1",
            data_abertura=timezone.now(),
        )
        response = self.client.get(
            reverse("chamados:fornecedor_detail", args=[fornecedor.pk])
        )
        self.assertContains(response, "Chamados vinculados")
        self.assertContains(response, "OS-FOR-1")

    def test_fornecedor_update_status_200(self):
        fornecedor = Fornecedor.objects.create(nome="Fornecedor Editar")
        response = self.client.get(
            reverse("chamados:fornecedor_update", args=[fornecedor.pk])
        )
        self.assertEqual(response.status_code, 200)

    def test_post_fornecedor_update_atualiza_fornecedor(self):
        fornecedor = Fornecedor.objects.create(nome="Fornecedor Antigo")
        response = self.client.post(
            reverse("chamados:fornecedor_update", args=[fornecedor.pk]),
            data=self._dados_fornecedor(nome="Fornecedor Novo"),
        )
        self.assertRedirects(response, reverse("chamados:fornecedores_list"))
        fornecedor.refresh_from_db()
        self.assertEqual(fornecedor.nome, "Fornecedor Novo")

    def test_menu_lateral_contem_contato_fornecedor(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, "Contato Fornecedor")

    def test_menu_lateral_aponta_cont_fornecedor_para_rota(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, reverse("chamados:fornecedores_list"))

    def test_models_existentes_no_app_chamados(self):
        # Invariante atual: o app contém os 4 models originais + Obra (gestão
        # do parque imobiliário em obra). Qualquer model adicional deve ser
        # avaliado.
        from django.apps import apps

        modelos = {model.__name__ for model in apps.get_app_config("chamados").get_models()}
        self.assertEqual(
            modelos,
            {"Fornecedor", "Ativo", "Chamado", "AtualizacaoChamado", "Obra"},
        )

    def test_funcionalidade_fornecedor_nao_cria_chamado(self):
        self.client.post(
            reverse("chamados:fornecedor_create"), data=self._dados_fornecedor()
        )
        self.assertEqual(Chamado.objects.count(), 0)

    def test_funcionalidade_fornecedor_nao_cria_atualizacao_chamado(self):
        self.client.post(
            reverse("chamados:fornecedor_create"), data=self._dados_fornecedor()
        )
        self.assertEqual(AtualizacaoChamado.objects.count(), 0)


class AtivoModelTests(TestCase):
    def _dados_ativo_validos(self, **overrides):
        dados = {
            "ativo_prisma": "PRISMA-001",
            "nome_site": "Site Central",
            "endereco": "Rua A, 100",
            "cidade": "Sao Paulo",
            "uf": "SP",
            "regional": "Sudeste",
        }
        dados.update(overrides)
        return dados

    def test_criar_ativo(self):
        ativo = Ativo.objects.create(**self._dados_ativo_validos())
        self.assertEqual(Ativo.objects.count(), 1)
        self.assertTrue(ativo.ativo)

    def test_ativo_prisma_unique(self):
        Ativo.objects.create(**self._dados_ativo_validos())
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Ativo.objects.create(
                    **self._dados_ativo_validos(nome_site="Outro Site")
                )

    def test_ativo_str(self):
        ativo = Ativo.objects.create(**self._dados_ativo_validos())
        self.assertEqual(str(ativo), "PRISMA-001 - Site Central")


class ChamadoModelTests(TestCase):
    def setUp(self):
        self.ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-002",
            nome_site="Site Norte",
            endereco="Av. B, 200",
            cidade="Manaus",
            uf="AM",
            regional="Norte",
        )
        self.fornecedor = Fornecedor.objects.create(nome="Fornecedor Gamma")

    def _criar_chamado(self, **overrides):
        dados = {
            "ativo": self.ativo,
            "fornecedor": self.fornecedor,
            "numero_os": "OS-0001",
            "data_abertura": timezone.now(),
        }
        dados.update(overrides)
        return Chamado.objects.create(**dados)

    def test_criar_chamado_status_default_aberto(self):
        chamado = self._criar_chamado()
        self.assertEqual(chamado.status, StatusChamado.ABERTO)
        self.assertEqual(chamado.status, "aberto")

    def test_numero_os_unique(self):
        self._criar_chamado()
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                self._criar_chamado(numero_os="OS-0001")

    def test_chamado_str(self):
        chamado = self._criar_chamado(numero_os="OS-9999")
        self.assertEqual(str(chamado), "OS OS-9999 - Site Norte")

    def test_emergencial_aberta_nao_eh_status(self):
        valores_status = {value for value, _ in StatusChamado.choices}
        self.assertNotIn("emergencial_aberta", valores_status)
        self.assertNotIn("emergencial", valores_status)


class AtualizacaoChamadoModelTests(TestCase):
    def setUp(self):
        self.ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-003",
            nome_site="Site Sul",
            endereco="Rua C, 300",
            cidade="Porto Alegre",
            uf="RS",
            regional="Sul",
        )
        self.chamado = Chamado.objects.create(
            ativo=self.ativo,
            numero_os="OS-0010",
            data_abertura=timezone.now(),
        )

    def test_criar_atualizacao(self):
        atualizacao = AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Equipe deslocada ate o site.",
        )
        self.assertEqual(atualizacao.tipo_evento, TipoEventoAtualizacao.COMENTARIO)
        self.assertEqual(atualizacao.status_resultante, StatusChamado.PENDENTE)

    def test_related_name_chamado_atualizacoes(self):
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Primeira atualizacao.",
        )
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Segunda atualizacao.",
            tipo_evento=TipoEventoAtualizacao.ATUALIZACAO_STATUS,
            status_resultante=StatusChamado.CONCLUIDO,
        )
        self.assertEqual(self.chamado.atualizacoes.count(), 2)

    def test_atualizacao_str(self):
        atualizacao = AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Abertura.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
        )
        self.assertEqual(str(atualizacao), "OS-0010 - Abertura emergencial")

    def test_abertura_emergencial_existe_como_tipo_evento(self):
        valores_tipo = {value for value, _ in TipoEventoAtualizacao.choices}
        self.assertIn("abertura_emergencial", valores_tipo)


def _criar_ativo(**overrides):
    """Helper para criar ativos validos com dados padrao."""
    dados = {
        "ativo_prisma": "PRISMA-100",
        "nome_site": "Site Centro",
        "endereco": "Av. Paulista, 1000",
        "cidade": "Sao Paulo",
        "uf": "SP",
        "regional": "Sudeste",
    }
    dados.update(overrides)
    return Ativo.objects.create(**dados)


class AtivosListViewTests(_LoginClienteMixin, TestCase):
    def test_pagina_ativos_status_200(self):
        response = self.client.get(reverse("chamados:ativos_list"))
        self.assertEqual(response.status_code, 200)

    def test_listagem_exibe_ativo_cadastrado(self):
        _criar_ativo(ativo_prisma="PRISMA-LIST-01", nome_site="Site Listagem")
        response = self.client.get(reverse("chamados:ativos_list"))
        self.assertContains(response, "PRISMA-LIST-01")
        self.assertContains(response, "Site Listagem")

    def test_filtro_q_por_ativo_prisma(self):
        _criar_ativo(ativo_prisma="PRISMA-Q-01", nome_site="Site A")
        _criar_ativo(ativo_prisma="PRISMA-Q-02", nome_site="Site B")
        response = self.client.get(reverse("chamados:ativos_list"), {"q": "PRISMA-Q-01"})
        self.assertContains(response, "PRISMA-Q-01")
        self.assertNotContains(response, "PRISMA-Q-02")

    def test_filtro_q_por_nome_site(self):
        _criar_ativo(ativo_prisma="PRISMA-N-01", nome_site="Estacao Verde")
        _criar_ativo(ativo_prisma="PRISMA-N-02", nome_site="Estacao Azul")
        response = self.client.get(reverse("chamados:ativos_list"), {"q": "Verde"})
        self.assertContains(response, "Estacao Verde")
        self.assertNotContains(response, "Estacao Azul")

    def test_filtro_cidade(self):
        _criar_ativo(ativo_prisma="PRISMA-C-01", cidade="Curitiba")
        _criar_ativo(ativo_prisma="PRISMA-C-02", cidade="Salvador")
        response = self.client.get(reverse("chamados:ativos_list"), {"cidade": "Curitiba"})
        self.assertContains(response, "PRISMA-C-01")
        self.assertNotContains(response, "PRISMA-C-02")

    def test_filtro_uf(self):
        _criar_ativo(ativo_prisma="PRISMA-U-01", uf="SP")
        _criar_ativo(ativo_prisma="PRISMA-U-02", uf="RJ")
        response = self.client.get(reverse("chamados:ativos_list"), {"uf": "RJ"})
        self.assertContains(response, "PRISMA-U-02")
        self.assertNotContains(response, "PRISMA-U-01")

    def test_filtro_regional(self):
        _criar_ativo(ativo_prisma="PRISMA-R-01", regional="Sul")
        _criar_ativo(ativo_prisma="PRISMA-R-02", regional="Nordeste")
        response = self.client.get(reverse("chamados:ativos_list"), {"regional": "Sul"})
        self.assertContains(response, "PRISMA-R-01")
        self.assertNotContains(response, "PRISMA-R-02")

    def test_filtro_ativo_sim_retorna_apenas_ativos(self):
        _criar_ativo(ativo_prisma="PRISMA-AT-01", ativo=True)
        _criar_ativo(ativo_prisma="PRISMA-AT-02", ativo=False)
        response = self.client.get(reverse("chamados:ativos_list"), {"ativo": "sim"})
        self.assertContains(response, "PRISMA-AT-01")
        self.assertNotContains(response, "PRISMA-AT-02")

    def test_filtro_ativo_nao_retorna_apenas_inativos(self):
        _criar_ativo(ativo_prisma="PRISMA-IN-01", ativo=True)
        _criar_ativo(ativo_prisma="PRISMA-IN-02", ativo=False)
        response = self.client.get(reverse("chamados:ativos_list"), {"ativo": "nao"})
        self.assertContains(response, "PRISMA-IN-02")
        self.assertNotContains(response, "PRISMA-IN-01")

    def test_filtro_sem_cidade_lista_apenas_ativos_com_cidade_vazia(self):
        _criar_ativo(ativo_prisma="QUAL-OK", cidade="São Paulo")
        _criar_ativo(ativo_prisma="QUAL-SEM-CIDADE", cidade="")
        response = self.client.get(reverse("chamados:ativos_list"), {"sem": "cidade"})
        self.assertContains(response, "QUAL-SEM-CIDADE")
        self.assertNotContains(response, "QUAL-OK")

    def test_filtro_sem_uf_lista_apenas_ativos_com_uf_vazia(self):
        _criar_ativo(ativo_prisma="QUAL-UF-OK", uf="SP")
        _criar_ativo(ativo_prisma="QUAL-SEM-UF", uf="")
        response = self.client.get(reverse("chamados:ativos_list"), {"sem": "uf"})
        self.assertContains(response, "QUAL-SEM-UF")
        self.assertNotContains(response, "QUAL-UF-OK")

    def test_filtro_sem_regional_lista_apenas_ativos_com_regional_vazia(self):
        _criar_ativo(ativo_prisma="QUAL-REG-OK", regional="Sul")
        _criar_ativo(ativo_prisma="QUAL-SEM-REG", regional="")
        response = self.client.get(reverse("chamados:ativos_list"), {"sem": "regional"})
        self.assertContains(response, "QUAL-SEM-REG")
        self.assertNotContains(response, "QUAL-REG-OK")

    def test_filtro_sem_invalido_e_ignorado(self):
        """Valor não-permitido em ?sem= é silenciosamente ignorado."""
        _criar_ativo(ativo_prisma="QUAL-IGN", cidade="X")
        response = self.client.get(
            reverse("chamados:ativos_list"), {"sem": "ativo_prisma"}
        )
        # Não deve filtrar nada; ativo aparece normalmente.
        self.assertContains(response, "QUAL-IGN")

    def test_contadores_de_qualidade_no_contexto(self):
        _criar_ativo(ativo_prisma="QC-1", cidade="X", uf="SP", regional="Sul")
        _criar_ativo(ativo_prisma="QC-2", cidade="", uf="SP", regional="Sul")
        _criar_ativo(ativo_prisma="QC-3", cidade="X", uf="", regional="")
        response = self.client.get(reverse("chamados:ativos_list"))
        contadores = response.context["contadores_qualidade"]
        # Pelo menos 1 ativo sem cidade e pelo menos 1 sem uf/regional
        self.assertGreaterEqual(contadores["cidade"], 1)
        self.assertGreaterEqual(contadores["uf"], 1)
        self.assertGreaterEqual(contadores["regional"], 1)

    def test_atalhos_de_qualidade_aparecem_quando_ha_problemas(self):
        _criar_ativo(ativo_prisma="QS-SEM-CIDADE", cidade="")
        response = self.client.get(reverse("chamados:ativos_list"))
        self.assertContains(response, "quality-chip")
        self.assertContains(response, "Sem Cidade")

    def test_celulas_vazias_renderizam_como_placeholder(self):
        _criar_ativo(ativo_prisma="EMPTY-1", cidade="")
        response = self.client.get(reverse("chamados:ativos_list"))
        self.assertContains(response, "empty-cell")


class AtivoCreateViewTests(_LoginClienteMixin, TestCase):
    def test_pagina_novo_ativo_status_200(self):
        response = self.client.get(reverse("chamados:ativo_create"))
        self.assertEqual(response.status_code, 200)

    def test_post_cria_ativo(self):
        dados = {
            "ativo_prisma": "PRISMA-NEW-01",
            "nome_site": "Site Novo",
            "endereco": "Rua Z, 99",
            "cidade": "  Recife  ",
            "uf": "pe",
            "regional": "  Nordeste  ",
            "tipo_imovel": "",
            "lider_coordenacao": "",
            "tipo_site_sla": "",
            "ativo": "on",
        }
        response = self.client.post(reverse("chamados:ativo_create"), data=dados)
        self.assertRedirects(response, reverse("chamados:ativos_list"))
        ativo = Ativo.objects.get(ativo_prisma="PRISMA-NEW-01")
        self.assertEqual(ativo.uf, "PE")
        self.assertEqual(ativo.cidade, "Recife")
        self.assertEqual(ativo.regional, "Nordeste")


def _arquivo_xlsx(linhas, nome="parque.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    for linha in linhas:
        sheet.append(linha)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)
    return SimpleUploadedFile(
        nome,
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class AtivosImportTests(TestCase):
    CABECALHO = [
        "Ativo Prisma",
        "Nome do Site",
        "Endereço",
        "Cidade",
        "UF",
        "Regional",
        "Tipo de imóvel",
        "Líder da coordenação",
        "Tipo site/SLA",
        "Ativo?",
    ]

    def setUp(self):
        # As telas internas exigem usuário autenticado para identificar o
        # operador. Os testes que exercitam essas telas precisam logar antes.
        from django.contrib.auth import get_user_model

        User = get_user_model()
        self.user = User.objects.create_user(
            username="operador_ativos_import",
            password="senhaforte123",
        )
        self.client.force_login(self.user)

    def _importar(self, linhas):
        from .importadores import importar_ativos_excel

        return importar_ativos_excel(_arquivo_xlsx(linhas))

    def test_pagina_importar_ativos_status_200(self):
        response = self.client.get(reverse("chamados:ativos_import"))
        self.assertEqual(response.status_code, 200)

    def test_tela_importar_contem_campo_upload(self):
        response = self.client.get(reverse("chamados:ativos_import"))
        self.assertContains(response, 'type="file"')
        self.assertContains(response, 'name="arquivo"')

    def test_tela_importar_contem_botao_importar(self):
        response = self.client.get(reverse("chamados:ativos_import"))
        self.assertContains(response, "Importar")

    def test_usuario_anonimo_recebe_302_em_ativos_import(self):
        """Confirma a regra atual: usuário anônimo é redirecionado para
        login ao tentar acessar a tela de importação."""
        self.client.logout()
        response = self.client.get(reverse("chamados:ativos_import"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response["Location"])

    def test_usuario_autenticado_recebe_200_em_ativos_import(self):
        """Confirma a regra atual: usuário logado acessa normalmente."""
        # setUp já faz force_login; aqui apenas explicitamos o contrato.
        response = self.client.get(reverse("chamados:ativos_import"))
        self.assertEqual(response.status_code, 200)

    def test_form_rejeita_arquivo_sem_extensao_xlsx(self):
        from .forms import AtivosImportForm

        arquivo = SimpleUploadedFile("parque.csv", b"conteudo", content_type="text/csv")
        form = AtivosImportForm(files={"arquivo": arquivo})
        self.assertFalse(form.is_valid())
        self.assertIn("formato .xlsx", str(form.errors["arquivo"]))

    def test_importacao_cria_ativo_novo(self):
        self._importar(
            [
                self.CABECALHO,
                ["PR-001", "Site Um", "Rua Um", "Santos", "sp", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertTrue(Ativo.objects.filter(ativo_prisma="PR-001").exists())

    def test_importacao_atualiza_ativo_existente_por_ativo_prisma(self):
        Ativo.objects.create(
            ativo_prisma="PR-002",
            nome_site="Antigo",
            endereco="Rua Antiga",
            cidade="Santos",
            uf="SP",
            regional="Sudeste",
        )
        self._importar(
            [
                self.CABECALHO,
                ["PR-002", "Site Atualizado", "Rua Nova", "Santos", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        ativo = Ativo.objects.get(ativo_prisma="PR-002")
        self.assertEqual(ativo.nome_site, "Site Atualizado")
        self.assertEqual(ativo.endereco, "Rua Nova")

    def test_importacao_nao_duplica_ativo_prisma(self):
        Ativo.objects.create(
            ativo_prisma="PR-003",
            nome_site="Antigo",
            endereco="Rua Antiga",
            cidade="Santos",
            uf="SP",
            regional="Sudeste",
        )
        self._importar(
            [
                self.CABECALHO,
                ["PR-003", "Site Atualizado", "Rua Nova", "Santos", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(Ativo.objects.filter(ativo_prisma="PR-003").count(), 1)

    def test_importacao_normaliza_uf_para_maiusculo(self):
        self._importar(
            [
                self.CABECALHO,
                ["PR-004", "Site UF", "Rua UF", "Santos", "sp", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(Ativo.objects.get(ativo_prisma="PR-004").uf, "SP")

    def test_importacao_ignora_linha_sem_ativo_prisma(self):
        resultado = self._importar(
            [
                self.CABECALHO,
                ["", "Site Sem Prisma", "Rua", "Santos", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(resultado.ignorados, 1)
        self.assertEqual(Ativo.objects.count(), 0)

    def test_importacao_aceita_linha_sem_nome_site_e_conta_como_vazio(self):
        """Nome do site não é mais obrigatório — linha é salva mas contada
        em campos_vazios."""
        resultado = self._importar(
            [
                self.CABECALHO,
                ["PR-005", "", "Rua", "Santos", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(resultado.ignorados, 0)
        self.assertTrue(Ativo.objects.filter(ativo_prisma="PR-005").exists())
        self.assertEqual(resultado.campos_vazios.get("nome_site", 0), 1)

    def test_importacao_aceita_linha_sem_cidade_e_conta_como_vazio(self):
        """Cidade não é mais obrigatória — linha é salva, mas se o endereço
        não tiver padrão extraível, cidade fica em campos_vazios."""
        resultado = self._importar(
            [
                self.CABECALHO,
                # Endereço sem " - " separator, então cidade não pode ser inferida.
                ["PR-006", "Site Sem Cidade", "Rua única", "", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(resultado.ignorados, 0)
        self.assertTrue(Ativo.objects.filter(ativo_prisma="PR-006").exists())
        self.assertEqual(resultado.campos_vazios.get("cidade", 0), 1)

    def test_importacao_rejeita_apenas_quando_ativo_prisma_em_branco(self):
        """A única linha realmente rejeitada é aquela sem o identificador único."""
        resultado = self._importar(
            [
                self.CABECALHO,
                ["", "Site Sem Prisma", "Rua", "Santos", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(resultado.ignorados, 1)
        self.assertEqual(resultado.criados, 0)

    def test_importacao_retorna_contagem_de_criados(self):
        resultado = self._importar(
            [
                self.CABECALHO,
                ["PR-007", "Site Criado", "Rua", "Santos", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(resultado.criados, 1)

    def test_importacao_retorna_contagem_de_atualizados(self):
        Ativo.objects.create(
            ativo_prisma="PR-008",
            nome_site="Antigo",
            endereco="Rua Antiga",
            cidade="Santos",
            uf="SP",
            regional="Sudeste",
        )
        resultado = self._importar(
            [
                self.CABECALHO,
                ["PR-008", "Atualizado", "Rua", "Santos", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(resultado.atualizados, 1)

    def test_importacao_retorna_contagem_de_ignorados(self):
        resultado = self._importar(
            [
                self.CABECALHO,
                ["", "Sem Prisma", "Rua", "Santos", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(resultado.ignorados, 1)

    def test_importacao_aceita_cabecalho_nome_do_site(self):
        self._importar(
            [
                self.CABECALHO,
                ["PR-009", "Nome Aceito", "Rua", "Santos", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(Ativo.objects.get(ativo_prisma="PR-009").nome_site, "Nome Aceito")

    def test_importacao_aceita_cabecalho_endereco(self):
        self._importar(
            [
                self.CABECALHO,
                ["PR-010", "Site", "Rua Endereço", "Santos", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(Ativo.objects.get(ativo_prisma="PR-010").endereco, "Rua Endereço")

    def test_importacao_aceita_cabecalho_municipio(self):
        cabecalho = [
            "Prisma",
            "Site",
            "Logradouro",
            "Município",
            "Estado",
            "Região",
        ]
        self._importar(
            [
                cabecalho,
                ["PR-011", "Site", "Rua", "Campinas", "SP", "Interior"],
            ]
        )
        self.assertEqual(Ativo.objects.get(ativo_prisma="PR-011").cidade, "Campinas")

    def test_importacao_interpreta_ativo_sim_como_true(self):
        self._importar(
            [
                self.CABECALHO,
                ["PR-012", "Site", "Rua", "Santos", "SP", "Sudeste", "", "", "", "sim"],
            ]
        )
        self.assertTrue(Ativo.objects.get(ativo_prisma="PR-012").ativo)

    def test_importacao_interpreta_ativo_nao_como_false(self):
        self._importar(
            [
                self.CABECALHO,
                ["PR-013", "Site", "Rua", "Santos", "SP", "Sudeste", "", "", "", "não"],
            ]
        )
        self.assertFalse(Ativo.objects.get(ativo_prisma="PR-013").ativo)

    def test_botao_importar_parque_imobiliario_aparece_em_ativos(self):
        response = self.client.get(reverse("chamados:ativos_list"))
        self.assertContains(response, "Importar parque imobiliário")

    def test_botao_importar_parque_imobiliario_aponta_para_rota(self):
        response = self.client.get(reverse("chamados:ativos_list"))
        self.assertContains(response, reverse("chamados:ativos_import"))

    def test_matriz_ativos_exibe_orientacao_quando_vazia(self):
        response = self.client.get(reverse("chamados:ativos_list"))
        self.assertContains(
            response,
            "Nenhum ativo encontrado. Importe o parque imobiliário ou cadastre um novo ativo.",
        )

    def test_importacao_nao_cria_chamado(self):
        self._importar(
            [
                self.CABECALHO,
                ["PR-014", "Site", "Rua", "Santos", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(Chamado.objects.count(), 0)

    def test_importacao_nao_cria_atualizacao_chamado(self):
        self._importar(
            [
                self.CABECALHO,
                ["PR-015", "Site", "Rua", "Santos", "SP", "Sudeste", "", "", "", ""],
            ]
        )
        self.assertEqual(AtualizacaoChamado.objects.count(), 0)


class AtivoUpdateViewTests(_LoginClienteMixin, TestCase):
    def test_post_atualiza_ativo(self):
        ativo = _criar_ativo(ativo_prisma="PRISMA-UP-01", nome_site="Antigo")
        url = reverse("chamados:ativo_update", args=[ativo.pk])
        dados = {
            "ativo_prisma": "PRISMA-UP-01",
            "nome_site": "Atualizado",
            "endereco": ativo.endereco,
            "cidade": ativo.cidade,
            "uf": ativo.uf,
            "regional": ativo.regional,
            "tipo_imovel": "",
            "lider_coordenacao": "",
            "tipo_site_sla": "",
            "ativo": "on",
        }
        response = self.client.post(url, data=dados)
        self.assertRedirects(response, reverse("chamados:ativos_list"))
        ativo.refresh_from_db()
        self.assertEqual(ativo.nome_site, "Atualizado")


class AtivoDetailViewTests(_LoginClienteMixin, TestCase):
    def test_detalhe_exibe_dados_do_ativo(self):
        ativo = _criar_ativo(ativo_prisma="PRISMA-DET-01", nome_site="Site Detalhe")
        response = self.client.get(reverse("chamados:ativo_detail", args=[ativo.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "PRISMA-DET-01")
        self.assertContains(response, "Site Detalhe")

    def test_detalhe_possui_botao_abrir_chamado(self):
        ativo = _criar_ativo(ativo_prisma="PRISMA-BTN-01")
        response = self.client.get(reverse("chamados:ativo_detail", args=[ativo.pk]))
        self.assertContains(response, "Abrir chamado para este ativo")
        self.assertContains(response, 'data-feature="abrir-chamado"')


class HomeLinkTests(_LoginClienteMixin, TestCase):
    def test_home_possui_link_para_matriz_de_ativos(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Matriz de Ativos")
        self.assertContains(response, reverse("chamados:ativos_list"))

    def test_home_possui_card_para_importar_parque(self):
        response = self.client.get(reverse("chamados:home"))
        # O card "Ações principais" renderiza com o título "Importar Parque".
        self.assertContains(response, "Importar Parque")
        self.assertContains(response, reverse("chamados:ativos_import"))

    def test_home_possui_card_para_contato_fornecedor(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, "Contato Fornecedor")
        self.assertContains(response, reverse("chamados:fornecedores_list"))

    def test_home_possui_card_para_consolidado_obras(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, "Consolidado Obras")
        # Card aponta direto para a tela de Obras (a URL antiga continua redirecionando).
        self.assertContains(response, reverse("chamados:obras_list"))

    def test_home_contem_acoes_rapidas(self):
        response = self.client.get(reverse("chamados:home"))
        # Títulos exibidos no novo layout (Title-Case).
        self.assertContains(response, "Novo Chamado")
        self.assertContains(response, "Atualizar Report")
        self.assertContains(response, "Importar Parque")

    def test_home_contem_wrapper_visual_principal(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, 'class="home-page"')

    def test_home_contem_hero_visual(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, 'class="home-hero"')

    def test_home_contem_quick_actions(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, 'class="quick-actions"')

    def test_home_contem_module_grid(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, 'class="module-grid"')

    def test_home_contem_tres_cards_de_modulo(self):
        # O layout aprovado lista exatamente 3 cards em "Visão operacional".
        response = self.client.get(reverse("chamados:home"))
        conteudo = response.content.decode("utf-8")
        self.assertGreaterEqual(conteudo.count('class="module-card'), 3)

    def test_home_contem_metricas_predios_e_lojas(self):
        # O hero expõe split Prédios | Lojas em Total/Pendentes/Concluídos.
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, "hero-metric-split")
        self.assertContains(response, "Prédios")
        self.assertContains(response, "Lojas")

    def test_css_home_contem_regras_visuais_obrigatorias(self):
        css_path = Path(__file__).resolve().parent.parent / "static" / "css" / "app.css"
        css = css_path.read_text(encoding="utf-8")
        self.assertIn(".home-hero", css)
        self.assertIn("border-radius", css)
        self.assertIn(".module-grid", css)
        self.assertIn("display: grid", css)
        self.assertIn(".module-card", css)
        self.assertIn("min-height", css)


class ChamadoCreateViewTests(_LoginClienteMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-CH-01",
            nome_site="Site Chamado",
            endereco="Rua C, 50",
            cidade="Belo Horizonte",
            uf="MG",
            regional="Sudeste",
        )

    def _post_minimo(self, **overrides):
        dados = {
            "ativo": self.ativo.pk,
            "fornecedor": "",
            "numero_os": "  OS-100  ",
            "data_abertura": "2026-05-16T10:30",
            "solicitante": "",
            "contato_solicitante": "",
            "dados_portaria_retirada_chave": "",
            "denominacao": "",
            "acao_tomada": "",
            "command_center": "",
            "detalhamento_situacao": "",
            "status": StatusChamado.ABERTO,
        }
        dados.update(overrides)
        return dados

    def test_pagina_novo_chamado_status_200(self):
        response = self.client.get(reverse("chamados:chamado_create"))
        self.assertEqual(response.status_code, 200)

    def test_novo_chamado_sem_ativo_exibe_orientacao(self):
        response = self.client.get(reverse("chamados:chamado_create"))
        self.assertContains(
            response,
            "Selecione um ativo ou abra o chamado a partir da Matriz de Ativos.",
        )

    def test_form_exibe_campos_obrigatorios(self):
        response = self.client.get(reverse("chamados:chamado_create"))
        self.assertContains(response, 'name="ativo"')
        self.assertContains(response, 'name="numero_os"')
        self.assertContains(response, 'name="data_abertura"')

    def test_post_cria_chamado(self):
        response = self.client.post(
            reverse("chamados:chamado_create"), data=self._post_minimo()
        )
        chamado = Chamado.objects.get(numero_os="OS-100")
        self.assertRedirects(
            response, reverse("chamados:chamado_detail", args=[chamado.pk])
        )
        self.assertEqual(chamado.ativo, self.ativo)

    def test_chamado_assume_status_aberto_quando_nao_informado(self):
        dados = self._post_minimo()
        dados.pop("status")
        self.client.post(reverse("chamados:chamado_create"), data=dados)
        chamado = Chamado.objects.get(numero_os="OS-100")
        self.assertEqual(chamado.status, StatusChamado.ABERTO)

    def test_numero_os_normalizado_sem_espacos(self):
        self.client.post(
            reverse("chamados:chamado_create"),
            data=self._post_minimo(numero_os="   OS-200   "),
        )
        self.assertTrue(Chamado.objects.filter(numero_os="OS-200").exists())

    def test_preselecao_ativo_via_querystring(self):
        url = reverse("chamados:chamado_create") + f"?ativo={self.ativo.pk}"
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        form = response.context["form"]
        self.assertEqual(form.initial.get("ativo"), self.ativo)

    def test_preselecao_exibe_resumo_do_ativo(self):
        url = reverse("chamados:chamado_create") + f"?ativo={self.ativo.pk}"
        response = self.client.get(url)
        self.assertContains(response, "Ativo selecionado")
        self.assertContains(response, self.ativo.ativo_prisma)
        self.assertContains(response, self.ativo.nome_site)
        self.assertContains(response, self.ativo.endereco)
        self.assertContains(response, self.ativo.cidade)
        self.assertContains(response, self.ativo.uf)
        self.assertContains(response, self.ativo.regional)

    def test_status_emergencial_aberta_nao_aparece_no_form(self):
        response = self.client.get(reverse("chamados:chamado_create"))
        self.assertNotContains(response, 'value="emergencial_aberta"')

    def test_nenhuma_atualizacao_chamado_criada_automaticamente(self):
        self.client.post(reverse("chamados:chamado_create"), data=self._post_minimo())
        self.assertEqual(AtualizacaoChamado.objects.count(), 0)

    def test_form_renderiza_combobox_de_ativo(self):
        response = self.client.get(reverse("chamados:chamado_create"))
        self.assertContains(response, "data-combobox")
        self.assertContains(response, "data-combobox-input")
        self.assertContains(response, 'id="id_ativo"')
        self.assertContains(
            response, reverse("chamados:ativos_autocomplete")
        )

    def test_combobox_prefilled_quando_ativo_preselecionado(self):
        url = reverse("chamados:chamado_create") + f"?ativo={self.ativo.pk}"
        response = self.client.get(url)
        # Hidden input do form vem com o PK; input de busca vem com o Prisma.
        self.assertContains(response, f'value="{self.ativo.pk}"')
        self.assertContains(response, f'value="{self.ativo.ativo_prisma}"')


class AtivosAutocompleteViewTests(_LoginClienteMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.ativo_sp = Ativo.objects.create(
            ativo_prisma="PRISMA-AC-01",
            nome_site="Site Paulista",
            endereco="Av Paulista, 100",
            cidade="São Paulo",
            uf="SP",
            regional="Sudeste",
        )
        self.ativo_rj = Ativo.objects.create(
            ativo_prisma="PRISMA-AC-02",
            nome_site="Site Copacabana",
            endereco="Av Atlântica, 200",
            cidade="Rio de Janeiro",
            uf="RJ",
            regional="Sudeste",
        )
        self.ativo_inativo = Ativo.objects.create(
            ativo_prisma="PRISMA-AC-99",
            nome_site="Site Antigo",
            cidade="Recife",
            uf="PE",
            regional="Nordeste",
            ativo=False,
        )

    def test_autocomplete_retorna_json(self):
        response = self.client.get(reverse("chamados:ativos_autocomplete"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/json")

    def test_autocomplete_retorna_todos_os_ativos_quando_sem_termo(self):
        response = self.client.get(reverse("chamados:ativos_autocomplete"))
        ids = [item["id"] for item in response.json()["results"]]
        self.assertIn(self.ativo_sp.id, ids)
        self.assertIn(self.ativo_rj.id, ids)

    def test_autocomplete_filtra_por_prisma(self):
        response = self.client.get(
            reverse("chamados:ativos_autocomplete") + "?q=AC-01"
        )
        ids = [item["id"] for item in response.json()["results"]]
        self.assertEqual(ids, [self.ativo_sp.id])

    def test_autocomplete_filtra_por_nome_site(self):
        response = self.client.get(
            reverse("chamados:ativos_autocomplete") + "?q=Copacabana"
        )
        ids = [item["id"] for item in response.json()["results"]]
        self.assertEqual(ids, [self.ativo_rj.id])

    def test_autocomplete_filtra_por_cidade(self):
        response = self.client.get(
            reverse("chamados:ativos_autocomplete") + "?q=São Paulo"
        )
        ids = [item["id"] for item in response.json()["results"]]
        self.assertIn(self.ativo_sp.id, ids)
        self.assertNotIn(self.ativo_rj.id, ids)

    def test_autocomplete_ignora_inativos(self):
        response = self.client.get(
            reverse("chamados:ativos_autocomplete") + "?q=Antigo"
        )
        self.assertEqual(response.json()["results"], [])

    def test_autocomplete_retorna_label_e_sub(self):
        response = self.client.get(
            reverse("chamados:ativos_autocomplete") + "?q=AC-01"
        )
        item = response.json()["results"][0]
        self.assertEqual(item["label"], "PRISMA-AC-01")
        self.assertIn("Site Paulista", item["sub"])
        self.assertIn("São Paulo/SP", item["sub"])

    def test_autocomplete_pagina_resultados(self):
        for i in range(50):
            Ativo.objects.create(
                ativo_prisma=f"PRISMA-AC-BULK-{i:03d}",
                cidade="Curitiba",
                uf="PR",
                regional="Sul",
            )
        # Primeira pagina: 30 itens, com has_more=True e total=50.
        response = self.client.get(
            reverse("chamados:ativos_autocomplete") + "?q=BULK"
        )
        payload = response.json()
        self.assertEqual(len(payload["results"]), 30)
        self.assertEqual(payload["total"], 50)
        self.assertTrue(payload["has_more"])

        # Segunda pagina: restantes 20, sem mais conteúdo.
        response = self.client.get(
            reverse("chamados:ativos_autocomplete") + "?q=BULK&offset=30"
        )
        payload = response.json()
        self.assertEqual(len(payload["results"]), 20)
        self.assertFalse(payload["has_more"])


class AtivoDetailLinkTests(_LoginClienteMixin, TestCase):
    def test_botao_abrir_chamado_aponta_para_rota_real(self):
        ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-LINK-01",
            nome_site="Site Link",
            endereco="Rua L, 10",
            cidade="Brasilia",
            uf="DF",
            regional="Centro-Oeste",
        )
        response = self.client.get(reverse("chamados:ativo_detail", args=[ativo.pk]))
        esperado = reverse("chamados:chamado_create") + f"?ativo={ativo.pk}"
        self.assertContains(response, esperado)


class ChamadosListViewTests(_LoginClienteMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-LST-01",
            nome_site="Site Listagem",
            endereco="Rua A, 1",
            cidade="Florianopolis",
            uf="SC",
            regional="Sul",
        )

    def test_pagina_chamados_status_200(self):
        response = self.client.get(reverse("chamados:chamados_list"))
        self.assertEqual(response.status_code, 200)

    def test_listagem_exibe_chamado_cadastrado(self):
        Chamado.objects.create(
            ativo=self.ativo,
            numero_os="OS-LST-1",
            data_abertura=timezone.now(),
        )
        response = self.client.get(reverse("chamados:chamados_list"))
        self.assertContains(response, "OS-LST-1")
        self.assertContains(response, self.ativo.nome_site)


class ChamadoDetailViewTests(_LoginClienteMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-DET-CH",
            nome_site="Site Detalhe Chamado",
            endereco="Av. Detalhe, 200",
            cidade="Recife",
            uf="PE",
            regional="Nordeste",
        )
        self.chamado = Chamado.objects.create(
            ativo=self.ativo,
            numero_os="OS-DET-1",
            data_abertura=timezone.now(),
        )

    def test_pagina_detalhe_chamado_status_200(self):
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertEqual(response.status_code, 200)

    def test_detalhe_exibe_numero_os(self):
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, "OS-DET-1")

    def test_detalhe_exibe_dados_do_ativo_vinculado(self):
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, self.ativo.ativo_prisma)
        self.assertContains(response, self.ativo.nome_site)
        self.assertContains(response, self.ativo.endereco)
        self.assertContains(response, self.ativo.cidade)
        self.assertContains(response, self.ativo.regional)

    def test_detalhe_exibe_botao_atualizar_report(self):
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, "Atualizar report")

    def test_botao_atualizar_report_aponta_para_rota_correta(self):
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(
            response, reverse("chamados:atualizar_report_form", args=[self.chamado.pk])
        )

    def test_detalhe_sem_atualizacoes_exibe_mensagem(self):
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(
            response, "Nenhuma atualização registrada até o momento."
        )

    def test_detalhe_lista_atualizacao_existente(self):
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Equipe chegou ao local.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
            status_resultante=StatusChamado.PENDENTE,
        )
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, "Equipe chegou ao local.")

    def test_detalhe_exibe_tipo_evento_display(self):
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Aberto.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
        )
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, "Abertura emergencial")

    def test_detalhe_exibe_status_resultante_display(self):
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Pendente.",
            status_resultante=StatusChamado.PENDENTE,
        )
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, "Pendente")

    def test_detalhe_exibe_texto_atualizacao(self):
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Texto da atualização no detalhe.",
        )
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, "Texto da atualização no detalhe.")

    def test_detalhe_exibe_data_hora_atualizacao(self):
        atualizacao = AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Com data.",
        )
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, atualizacao.criado_em.strftime("%d/%m/%Y %H:%M"))

    def test_detalhe_exibe_usuario_da_atualizacao_quando_existir(self):
        from django.contrib.auth import get_user_model

        usuario = get_user_model().objects.create_user(
            username="operador_detalhe", password="senhaforte123"
        )
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Com usuario.",
            criado_por=usuario,
        )
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, "operador_detalhe")

    def test_detalhe_exibe_bloco_texto_whatsapp_atual(self):
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, 'id="texto-whatsapp-detalhe"')

    def test_bloco_whatsapp_contem_emergencial_aberta_no_primeiro_report(self):
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, "EMERGENCIAL ABERTA")

    def test_bloco_whatsapp_contem_status_atualizado_com_atualizacao(self):
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Equipe aguardando peça.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
            status_resultante=StatusChamado.PENDENTE,
        )
        self.chamado.status = StatusChamado.PENDENTE
        self.chamado.save(update_fields=["status"])
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, "EMERGENCIAL PENDENTE")
        self.assertContains(response, "Equipe aguardando peça.")

    def test_botao_copiar_texto_whatsapp_aparece_no_detalhe(self):
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, "Copiar texto WhatsApp")

    def test_botao_copiar_usa_data_copy_target_correto(self):
        response = self.client.get(
            reverse("chamados:chamado_detail", args=[self.chamado.pk])
        )
        self.assertContains(response, 'data-copy-target="texto-whatsapp-detalhe"')

    def test_view_chamado_detail_nao_cria_atualizacao(self):
        count_antes = AtualizacaoChamado.objects.count()
        self.client.get(reverse("chamados:chamado_detail", args=[self.chamado.pk]))
        self.assertEqual(AtualizacaoChamado.objects.count(), count_antes)

    def test_view_chamado_detail_nao_altera_status(self):
        status_antes = self.chamado.status
        self.client.get(reverse("chamados:chamado_detail", args=[self.chamado.pk]))
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, status_antes)

    def test_template_detalhe_nao_contem_url_wa_me(self):
        template_path = (
            Path(__file__).resolve().parent.parent
            / "templates"
            / "chamados"
            / "chamado_detail.html"
        )
        self.assertNotIn("wa.me", template_path.read_text(encoding="utf-8"))

    def test_template_detalhe_nao_contem_exportar_excel(self):
        template_path = (
            Path(__file__).resolve().parent.parent
            / "templates"
            / "chamados"
            / "chamado_detail.html"
        )
        self.assertNotIn("Exportar Excel", template_path.read_text(encoding="utf-8"))


class ConsolidadoObrasViewTests(_LoginClienteMixin, TestCase):
    """A URL antiga /consolidado-obras/ agora redireciona para a tela de Obras."""

    def test_consolidado_obras_redireciona_para_obras_list(self):
        response = self.client.get(reverse("chamados:consolidado_obras"))
        self.assertRedirects(response, reverse("chamados:obras_list"))

    def test_consolidado_nao_cria_chamado(self):
        antes = Chamado.objects.count()
        self.client.get(reverse("chamados:consolidado_obras"))
        self.assertEqual(Chamado.objects.count(), antes)

    def test_consolidado_nao_cria_atualizacao_chamado(self):
        antes = AtualizacaoChamado.objects.count()
        self.client.get(reverse("chamados:consolidado_obras"))
        self.assertEqual(AtualizacaoChamado.objects.count(), antes)

class HomeNovoChamadoLinkTests(_LoginClienteMixin, TestCase):
    def test_home_possui_link_para_novo_chamado(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, "Novo Chamado")
        self.assertContains(response, reverse("chamados:chamado_create"))

    def test_home_card_atualizar_report_aponta_para_rota_real(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, "Atualizar Report")
        self.assertContains(response, reverse("chamados:atualizar_report_list"))


class ServiceIsPrimeiroReportTests(TestCase):
    def setUp(self):
        self.ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-SRV-01",
            nome_site="Site Service",
            endereco="Rua S, 1",
            cidade="Vitoria",
            uf="ES",
            regional="Sudeste",
        )
        self.chamado = Chamado.objects.create(
            ativo=self.ativo,
            numero_os="OS-SRV-1",
            data_abertura=timezone.now(),
        )

    def test_chamado_sem_atualizacoes_eh_primeiro_report(self):
        from .services import is_primeiro_report

        self.assertTrue(is_primeiro_report(self.chamado))

    def test_chamado_com_atualizacao_nao_eh_primeiro_report(self):
        from .services import is_primeiro_report

        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Primeira",
        )
        self.assertFalse(is_primeiro_report(self.chamado))


class ServiceObterTipoReportTests(TestCase):
    def setUp(self):
        self.ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-SRV-02",
            nome_site="Site Tipo",
            endereco="Rua T, 2",
            cidade="Joao Pessoa",
            uf="PB",
            regional="Nordeste",
        )
        self.chamado = Chamado.objects.create(
            ativo=self.ativo,
            numero_os="OS-SRV-2",
            data_abertura=timezone.now(),
        )

    def test_retorna_abertura_emergencial_para_primeiro_report(self):
        from .services import obter_tipo_report

        self.assertEqual(obter_tipo_report(self.chamado), "abertura_emergencial")

    def test_retorna_atualizacao_status_apos_primeiro_report(self):
        from .services import obter_tipo_report

        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Primeira",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
        )
        self.assertEqual(obter_tipo_report(self.chamado), "atualizacao_status")


class ServiceValidarStatusReportTests(TestCase):
    def test_aceita_aberto(self):
        from .services import validar_status_report

        self.assertEqual(validar_status_report("aberto"), "aberto")

    def test_aceita_pendente(self):
        from .services import validar_status_report

        self.assertEqual(validar_status_report("pendente"), "pendente")

    def test_aceita_concluido(self):
        from .services import validar_status_report

        self.assertEqual(validar_status_report("concluido"), "concluido")

    def test_aceita_cancelado(self):
        from .services import validar_status_report

        self.assertEqual(validar_status_report("cancelado"), "cancelado")

    def test_aceita_nao_emergencial(self):
        from .services import validar_status_report

        self.assertEqual(validar_status_report("nao_emergencial"), "nao_emergencial")

    def test_rejeita_emergencial_aberta(self):
        from .services import validar_status_report

        with self.assertRaises(ValueError):
            validar_status_report("emergencial_aberta")

    def test_rejeita_abertura_emergencial(self):
        from .services import validar_status_report

        with self.assertRaises(ValueError):
            validar_status_report("abertura_emergencial")

    def test_rejeita_string_invalida(self):
        from .services import validar_status_report

        with self.assertRaises(ValueError):
            validar_status_report("qualquer-coisa")


class ServiceRegistrarReportTests(TestCase):
    def setUp(self):
        from django.contrib.auth import get_user_model

        self.User = get_user_model()
        self.ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-SRV-03",
            nome_site="Site Report",
            endereco="Rua R, 3",
            cidade="Goiania",
            uf="GO",
            regional="Centro-Oeste",
        )
        self.chamado = Chamado.objects.create(
            ativo=self.ativo,
            numero_os="OS-SRV-3",
            data_abertura=timezone.now(),
        )

    def test_primeiro_report_cria_atualizacao_com_abertura_emergencial(self):
        from .services import registrar_report

        atualizacao = registrar_report(self.chamado, texto_atualizacao="Abertura.")
        self.assertEqual(atualizacao.tipo_evento, "abertura_emergencial")

    def test_primeiro_report_sem_status_assume_aberto(self):
        from .services import registrar_report

        registrar_report(self.chamado, texto_atualizacao="Abertura.")
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, "aberto")

    def test_primeiro_report_com_status_pendente_atualiza_chamado(self):
        from .services import registrar_report

        registrar_report(
            self.chamado,
            texto_atualizacao="Em deslocamento.",
            status_resultante="pendente",
        )
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, "pendente")

    def test_segunda_atualizacao_cria_tipo_atualizacao_status(self):
        from .services import registrar_report

        registrar_report(self.chamado, texto_atualizacao="Abertura.")
        atualizacao = registrar_report(
            self.chamado,
            texto_atualizacao="Equipe chegou.",
            status_resultante="pendente",
        )
        self.assertEqual(atualizacao.tipo_evento, "atualizacao_status")

    def test_segunda_atualizacao_exige_status_resultante(self):
        from .services import registrar_report

        registrar_report(self.chamado, texto_atualizacao="Abertura.")
        with self.assertRaises(ValueError):
            registrar_report(self.chamado, texto_atualizacao="Sem status.")

    def test_registrar_report_atualiza_status_do_chamado(self):
        from .services import registrar_report

        registrar_report(self.chamado, texto_atualizacao="Abertura.")
        registrar_report(
            self.chamado,
            texto_atualizacao="Resolvido no local.",
            status_resultante="concluido",
        )
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, "concluido")

    def test_registrar_report_rejeita_texto_vazio(self):
        from .services import registrar_report

        with self.assertRaises(ValueError):
            registrar_report(self.chamado, texto_atualizacao="")
        with self.assertRaises(ValueError):
            registrar_report(self.chamado, texto_atualizacao="   ")

    def test_registrar_report_nao_cria_status_emergencial_aberta(self):
        from .services import registrar_report

        registrar_report(self.chamado, texto_atualizacao="Abertura.")
        self.chamado.refresh_from_db()
        self.assertNotEqual(self.chamado.status, "emergencial_aberta")
        valores_status = {v for v, _ in StatusChamado.choices}
        self.assertNotIn("emergencial_aberta", valores_status)
        with self.assertRaises(ValueError):
            registrar_report(
                self.chamado,
                texto_atualizacao="Tentativa invalida.",
                status_resultante="emergencial_aberta",
            )

    def test_registrar_report_associa_usuario_autenticado(self):
        from .services import registrar_report

        usuario = self.User.objects.create_user(
            username="operador", password="senhaforte123"
        )
        atualizacao = registrar_report(
            self.chamado,
            texto_atualizacao="Abertura por operador.",
            usuario=usuario,
        )
        self.chamado.refresh_from_db()
        self.assertEqual(atualizacao.criado_por, usuario)
        self.assertEqual(self.chamado.atualizado_por, usuario)


class ChamadoCreatePersistenciaTests(_LoginClienteMixin, TestCase):
    def test_tela_novo_chamado_nao_cria_atualizacao_automaticamente(self):
        ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-NC-1",
            nome_site="Site Nao Auto",
            endereco="Av. NC, 1",
            cidade="Salvador",
            uf="BA",
            regional="Nordeste",
        )
        dados = {
            "ativo": ativo.pk,
            "fornecedor": "",
            "numero_os": "OS-NC-1",
            "data_abertura": "2026-05-16T12:00",
            "solicitante": "",
            "contato_solicitante": "",
            "dados_portaria_retirada_chave": "",
            "denominacao": "",
            "acao_tomada": "",
            "command_center": "",
            "detalhamento_situacao": "",
            "status": "aberto",
        }
        self.client.post(reverse("chamados:chamado_create"), data=dados)
        chamado = Chamado.objects.get(numero_os="OS-NC-1")
        self.assertEqual(chamado.atualizacoes.count(), 0)
        self.assertEqual(AtualizacaoChamado.objects.count(), 0)


class FormatadoresReportTests(TestCase):
    def test_formatar_valor_report_none(self):
        from .services import formatar_valor_report

        self.assertEqual(formatar_valor_report(None), "-")

    def test_formatar_valor_report_string_vazia(self):
        from .services import formatar_valor_report

        self.assertEqual(formatar_valor_report(""), "-")
        self.assertEqual(formatar_valor_report("   "), "-")

    def test_formatar_valor_report_strip(self):
        from .services import formatar_valor_report

        self.assertEqual(formatar_valor_report("  texto  "), "texto")

    def test_formatar_data_report_datetime_brasileiro(self):
        from datetime import datetime

        from .services import formatar_data_report

        dt = datetime(2026, 5, 16, 14, 30)
        self.assertEqual(formatar_data_report(dt), "16/05/2026 14:30")

    def test_formatar_data_report_none(self):
        from .services import formatar_data_report

        self.assertEqual(formatar_data_report(None), "-")


def _criar_chamado_completo(**overrides):
    ativo = Ativo.objects.create(
        ativo_prisma="PRISMA-WA-01",
        nome_site="Site WhatsApp",
        endereco="Av. WA, 100",
        cidade="Campinas",
        uf="SP",
        regional="Sudeste",
        tipo_imovel="Loja",
        lider_coordenacao="Ana Lider",
        tipo_site_sla="SLA-A",
    )
    fornecedor = Fornecedor.objects.create(nome="Fornecedor WA")
    dados = {
        "ativo": ativo,
        "fornecedor": fornecedor,
        "numero_os": "OS-WA-001",
        "data_abertura": timezone.make_aware(timezone.datetime(2026, 5, 16, 9, 0)),
        "solicitante": "Joao Silva",
        "contato_solicitante": "11988887777",
        "dados_portaria_retirada_chave": "Portaria 24h",
        "denominacao": "Falha eletrica",
        "acao_tomada": "Acionamento equipe",
        "command_center": "CC-Sul",
        "detalhamento_situacao": "Queda total de energia.",
    }
    dados.update(overrides)
    return Chamado.objects.create(**dados)


class ObterTituloWhatsappTests(TestCase):
    def setUp(self):
        self.chamado = _criar_chamado_completo()

    def test_titulo_pendente(self):
        from .services import obter_titulo_whatsapp

        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Aberto.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
        )
        self.assertEqual(
            obter_titulo_whatsapp(self.chamado, status_preview="pendente"),
            "EMERGENCIAL PENDENTE",
        )

    def test_titulo_concluido(self):
        from .services import obter_titulo_whatsapp

        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Aberto.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
        )
        self.assertEqual(
            obter_titulo_whatsapp(self.chamado, status_preview="concluido"),
            "EMERGENCIAL CONCLUÍDA",
        )

    def test_titulo_cancelado(self):
        from .services import obter_titulo_whatsapp

        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Aberto.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
        )
        self.assertEqual(
            obter_titulo_whatsapp(self.chamado, status_preview="cancelado"),
            "EMERGENCIAL CANCELADA",
        )

    def test_titulo_nao_emergencial(self):
        from .services import obter_titulo_whatsapp

        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Aberto.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
        )
        self.assertEqual(
            obter_titulo_whatsapp(self.chamado, status_preview="nao_emergencial"),
            "NÃO EMERGENCIAL",
        )

    def test_rejeita_emergencial_aberta(self):
        from .services import obter_titulo_whatsapp

        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Aberto.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
        )
        with self.assertRaises(ValueError):
            obter_titulo_whatsapp(self.chamado, status_preview="emergencial_aberta")

    def test_rejeita_abertura_emergencial(self):
        from .services import obter_titulo_whatsapp

        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Aberto.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
        )
        with self.assertRaises(ValueError):
            obter_titulo_whatsapp(self.chamado, status_preview="abertura_emergencial")


class GerarTextoWhatsappTests(TestCase):
    def setUp(self):
        self.chamado = _criar_chamado_completo()

    def test_primeiro_report_contem_emergencial_aberta(self):
        from .services import gerar_texto_whatsapp

        texto = gerar_texto_whatsapp(self.chamado)
        self.assertIn("EMERGENCIAL ABERTA", texto)

    def test_inclui_numero_os(self):
        from .services import gerar_texto_whatsapp

        texto = gerar_texto_whatsapp(self.chamado)
        self.assertIn("OS-WA-001", texto)
        self.assertIn("Número da OS:", texto)

    def test_inclui_ativo_prisma(self):
        from .services import gerar_texto_whatsapp

        texto = gerar_texto_whatsapp(self.chamado)
        self.assertIn("Ativo Prisma:", texto)
        self.assertIn("PRISMA-WA-01", texto)

    def test_inclui_nome_do_site(self):
        from .services import gerar_texto_whatsapp

        texto = gerar_texto_whatsapp(self.chamado)
        self.assertIn("Nome do Site:", texto)
        self.assertIn("Site WhatsApp", texto)

    def test_inclui_endereco(self):
        from .services import gerar_texto_whatsapp

        texto = gerar_texto_whatsapp(self.chamado)
        self.assertIn("Endereço:", texto)
        self.assertIn("Av. WA, 100", texto)

    def test_inclui_cidade_uf_regional(self):
        from .services import gerar_texto_whatsapp

        texto = gerar_texto_whatsapp(self.chamado)
        self.assertIn("Cidade: Campinas", texto)
        self.assertIn("UF: SP", texto)
        self.assertIn("Regional: Sudeste", texto)

    def test_inclui_fornecedor_quando_existir(self):
        from .services import gerar_texto_whatsapp

        texto = gerar_texto_whatsapp(self.chamado)
        self.assertIn("Fornecedor: Fornecedor WA", texto)

    def test_usa_hifen_para_campos_vazios(self):
        from .services import gerar_texto_whatsapp

        ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-VAZIO",
            nome_site="Site Vazio",
            endereco="Rua V, 1",
            cidade="Curitiba",
            uf="PR",
            regional="Sul",
        )
        chamado = Chamado.objects.create(
            ativo=ativo,
            numero_os="OS-VAZIO",
            data_abertura=timezone.now(),
        )
        texto = gerar_texto_whatsapp(chamado)
        self.assertIn("Fornecedor: -", texto)
        self.assertIn("Solicitante: -", texto)
        self.assertIn("Contato: -", texto)
        self.assertIn("Denominação: -", texto)
        self.assertIn("Tipo de Prédio: -", texto)
        self.assertIn("Líder Regional: -", texto)

    def test_nao_exibe_none(self):
        from .services import gerar_texto_whatsapp

        ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-NN",
            nome_site="Site Sem Dados",
            endereco="Rua N, 1",
            cidade="Natal",
            uf="RN",
            regional="Nordeste",
        )
        chamado = Chamado.objects.create(
            ativo=ativo,
            numero_os="OS-NN",
            data_abertura=timezone.now(),
        )
        texto = gerar_texto_whatsapp(chamado)
        self.assertNotIn("None", texto)

    def test_sem_atualizacoes_inclui_mensagem(self):
        from .services import gerar_texto_whatsapp

        texto = gerar_texto_whatsapp(self.chamado)
        self.assertIn("Atualizações:", texto)
        self.assertIn("Sem atualizações registradas.", texto)

    def test_com_atualizacoes_inclui_historico(self):
        from .services import gerar_texto_whatsapp

        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Equipe deslocada.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
            status_resultante=StatusChamado.PENDENTE,
        )
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Resolvido no local.",
            tipo_evento=TipoEventoAtualizacao.ATUALIZACAO_STATUS,
            status_resultante=StatusChamado.CONCLUIDO,
        )
        texto = gerar_texto_whatsapp(self.chamado)
        self.assertIn("Equipe deslocada.", texto)
        self.assertIn("Resolvido no local.", texto)
        self.assertNotIn("Sem atualizações registradas.", texto)

    def test_historico_ordenado_por_criado_em(self):
        from .services import gerar_linhas_historico

        primeira = AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Primeira atualizacao.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
        )
        segunda = AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Segunda atualizacao.",
            tipo_evento=TipoEventoAtualizacao.ATUALIZACAO_STATUS,
            status_resultante=StatusChamado.PENDENTE,
        )
        linhas = gerar_linhas_historico(self.chamado)
        self.assertEqual(len(linhas), 2)
        self.assertIn("Primeira atualizacao.", linhas[0])
        self.assertIn("Segunda atualizacao.", linhas[1])
        self.assertGreater(segunda.criado_em, primeira.criado_em)

    def test_status_preview_nao_altera_chamado_no_banco(self):
        from .services import gerar_texto_whatsapp

        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Aberta.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
        )
        status_antes = self.chamado.status
        gerar_texto_whatsapp(self.chamado, status_preview="pendente")
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, status_antes)

    def test_texto_preview_inclui_secao_previa(self):
        from .services import gerar_texto_whatsapp

        texto = gerar_texto_whatsapp(
            self.chamado, texto_preview="Equipe a caminho do site."
        )
        self.assertIn("Prévia da atualização:", texto)
        self.assertIn("Equipe a caminho do site.", texto)

    def test_texto_preview_nao_cria_atualizacao(self):
        from .services import gerar_texto_whatsapp

        count_antes = AtualizacaoChamado.objects.count()
        gerar_texto_whatsapp(self.chamado, texto_preview="Nao deve gravar.")
        self.assertEqual(AtualizacaoChamado.objects.count(), count_antes)

    def test_nao_altera_atualizado_por(self):
        from django.contrib.auth import get_user_model

        from .services import gerar_texto_whatsapp

        User = get_user_model()
        usuario = User.objects.create_user(
            username="op_wa", password="senhaforte123"
        )
        self.chamado.atualizado_por = usuario
        self.chamado.save(update_fields=["atualizado_por"])
        gerar_texto_whatsapp(
            self.chamado,
            status_preview=None,
            texto_preview="Algum preview.",
        )
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.atualizado_por, usuario)


class AtualizarReportViewTests(_LoginClienteMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-AR-01",
            nome_site="Site Atualizar Report",
            endereco="Rua Report, 10",
            cidade="Santos",
            uf="SP",
            regional="Sudeste",
        )
        self.fornecedor = Fornecedor.objects.create(nome="Fornecedor Report")
        self.chamado = Chamado.objects.create(
            ativo=self.ativo,
            fornecedor=self.fornecedor,
            numero_os="OS-AR-001",
            data_abertura=timezone.now(),
            detalhamento_situacao="Falha de energia no site.",
        )

    def test_listagem_status_200(self):
        response = self.client.get(reverse("chamados:atualizar_report_list"))
        self.assertEqual(response.status_code, 200)

    def test_atualizar_report_exibe_cards_de_metricas(self):
        response = self.client.get(reverse("chamados:atualizar_report_list"))
        self.assertContains(response, "report-metric-card")
        self.assertContains(response, "Total filtrado")

    def test_atualizar_report_nao_exibe_texto_colado_total_filtrado(self):
        response = self.client.get(reverse("chamados:atualizar_report_list"))
        self.assertNotContains(response, "Total filtrado0")

    def test_listagem_exibe_chamado_existente(self):
        response = self.client.get(reverse("chamados:atualizar_report_list"))
        self.assertContains(response, "OS-AR-001")
        self.assertContains(response, "PRISMA-AR-01")
        self.assertContains(response, "Site Atualizar Report")

    def test_busca_q_encontra_por_numero_os(self):
        outro_ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-AR-02",
            nome_site="Site Outro",
            endereco="Rua Outra, 20",
            cidade="Rio de Janeiro",
            uf="RJ",
            regional="Sudeste",
        )
        Chamado.objects.create(
            ativo=outro_ativo,
            numero_os="OS-AR-999",
            data_abertura=timezone.now(),
        )
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"q": "OS-AR-001"}
        )
        self.assertContains(response, "OS-AR-001")
        self.assertNotContains(response, "OS-AR-999")

    def test_busca_q_encontra_por_ativo_prisma(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"q": "PRISMA-AR-01"}
        )
        self.assertContains(response, "OS-AR-001")
        self.assertContains(response, "PRISMA-AR-01")

    def _data_abertura(self, ano, mes, dia):
        return timezone.make_aware(timezone.datetime(ano, mes, dia, 10, 0))

    def _criar_chamado_filtravel(
        self,
        numero_os,
        *,
        ativo_prisma=None,
        cidade="Santos",
        regional="Sudeste",
        fornecedor=None,
        status=StatusChamado.ABERTO,
        data_abertura=None,
    ):
        ativo = Ativo.objects.create(
            ativo_prisma=ativo_prisma or f"PRISMA-{numero_os}",
            nome_site=f"Site {numero_os}",
            endereco=f"Rua {numero_os}, 1",
            cidade=cidade,
            uf="SP",
            regional=regional,
        )
        return Chamado.objects.create(
            ativo=ativo,
            fornecedor=fornecedor,
            numero_os=numero_os,
            data_abertura=data_abertura or timezone.now(),
            status=status,
        )

    def test_filtro_data_inicio_retorna_apenas_chamados_a_partir_da_data(self):
        self._criar_chamado_filtravel(
            "OS-DI-ANTES", data_abertura=self._data_abertura(2026, 1, 10)
        )
        self._criar_chamado_filtravel(
            "OS-DI-DEPOIS", data_abertura=self._data_abertura(2026, 2, 10)
        )
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"data_inicio": "2026-02-01"}
        )
        self.assertContains(response, "OS-DI-DEPOIS")
        self.assertNotContains(response, "OS-DI-ANTES")

    def test_filtro_data_fim_retorna_apenas_chamados_ate_a_data(self):
        self._criar_chamado_filtravel(
            "OS-DF-ANTES", data_abertura=self._data_abertura(2026, 1, 10)
        )
        self._criar_chamado_filtravel(
            "OS-DF-DEPOIS", data_abertura=self._data_abertura(2026, 2, 10)
        )
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"data_fim": "2026-01-31"}
        )
        self.assertContains(response, "OS-DF-ANTES")
        self.assertNotContains(response, "OS-DF-DEPOIS")

    def test_filtro_periodo_retorna_apenas_chamados_no_intervalo(self):
        self._criar_chamado_filtravel(
            "OS-PER-ANTES", data_abertura=self._data_abertura(2026, 1, 10)
        )
        self._criar_chamado_filtravel(
            "OS-PER-MEIO", data_abertura=self._data_abertura(2026, 2, 10)
        )
        self._criar_chamado_filtravel(
            "OS-PER-DEPOIS", data_abertura=self._data_abertura(2026, 3, 10)
        )
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {"data_inicio": "2026-02-01", "data_fim": "2026-02-28"},
        )
        self.assertContains(response, "OS-PER-MEIO")
        self.assertNotContains(response, "OS-PER-ANTES")
        self.assertNotContains(response, "OS-PER-DEPOIS")

    def test_filtro_status_aberto_retorna_apenas_abertos(self):
        self._criar_chamado_filtravel("OS-ST-ABERTO", status=StatusChamado.ABERTO)
        self._criar_chamado_filtravel("OS-ST-PENDENTE", status=StatusChamado.PENDENTE)
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"status": "aberto"}
        )
        self.assertContains(response, "OS-ST-ABERTO")
        self.assertNotContains(response, "OS-ST-PENDENTE")

    def test_filtro_status_pendente_retorna_apenas_pendentes(self):
        self._criar_chamado_filtravel("OS-ST-PEND-1", status=StatusChamado.PENDENTE)
        self._criar_chamado_filtravel("OS-ST-CONC-1", status=StatusChamado.CONCLUIDO)
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"status": "pendente"}
        )
        self.assertContains(response, "OS-ST-PEND-1")
        self.assertNotContains(response, "OS-ST-CONC-1")

    def test_filtro_status_concluido_retorna_apenas_concluidos(self):
        self._criar_chamado_filtravel("OS-ST-CONC-2", status=StatusChamado.CONCLUIDO)
        self._criar_chamado_filtravel("OS-ST-CANC-2", status=StatusChamado.CANCELADO)
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"status": "concluido"}
        )
        self.assertContains(response, "OS-ST-CONC-2")
        self.assertNotContains(response, "OS-ST-CANC-2")

    def test_filtro_status_cancelado_retorna_apenas_cancelados(self):
        self._criar_chamado_filtravel("OS-ST-CANC-3", status=StatusChamado.CANCELADO)
        self._criar_chamado_filtravel("OS-ST-ABER-3", status=StatusChamado.ABERTO)
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"status": "cancelado"}
        )
        self.assertContains(response, "OS-ST-CANC-3")
        self.assertNotContains(response, "OS-ST-ABER-3")

    def test_filtro_status_nao_emergencial_retorna_apenas_nao_emergenciais(self):
        self._criar_chamado_filtravel(
            "OS-ST-NE-4", status=StatusChamado.NAO_EMERGENCIAL
        )
        self._criar_chamado_filtravel("OS-ST-ABER-4", status=StatusChamado.ABERTO)
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {"status": "nao_emergencial"},
        )
        self.assertContains(response, "OS-ST-NE-4")
        self.assertNotContains(response, "OS-ST-ABER-4")

    def test_status_invalido_nao_gera_erro_500(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {"status": "abertura_emergencial"},
        )
        self.assertEqual(response.status_code, 200)

    def test_filtro_regional_retorna_apenas_regional_filtrada(self):
        self._criar_chamado_filtravel("OS-REG-SUL", regional="Sul")
        self._criar_chamado_filtravel("OS-REG-NORTE", regional="Norte")
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"regional": "Sul"}
        )
        self.assertContains(response, "OS-REG-SUL")
        self.assertNotContains(response, "OS-REG-NORTE")

    def test_filtro_cidade_retorna_apenas_cidade_filtrada(self):
        self._criar_chamado_filtravel("OS-CID-SANTOS", cidade="Santos")
        self._criar_chamado_filtravel("OS-CID-CAMPINAS", cidade="Campinas")
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"cidade": "Campinas"}
        )
        self.assertContains(response, "OS-CID-CAMPINAS")
        self.assertNotContains(response, "OS-CID-SANTOS")

    def test_filtro_fornecedor_retorna_apenas_fornecedor_filtrado(self):
        fornecedor_a = Fornecedor.objects.create(nome="Fornecedor Filtro A")
        fornecedor_b = Fornecedor.objects.create(nome="Fornecedor Filtro B")
        self._criar_chamado_filtravel("OS-FOR-A", fornecedor=fornecedor_a)
        self._criar_chamado_filtravel("OS-FOR-B", fornecedor=fornecedor_b)
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {"fornecedor": str(fornecedor_a.pk)},
        )
        self.assertContains(response, "OS-FOR-A")
        self.assertNotContains(response, "OS-FOR-B")

    def test_fornecedor_invalido_nao_gera_erro_500(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"fornecedor": "abc"}
        )
        self.assertEqual(response.status_code, 200)

    def test_campos_de_filtro_preservam_valores_na_resposta(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {
                "q": "OS",
                "data_inicio": "2026-01-01",
                "data_fim": "2026-12-31",
                "status": "aberto",
                "regional": "Sudeste",
                "cidade": "Santos",
                "fornecedor": str(self.fornecedor.pk),
            },
        )
        self.assertContains(response, 'value="OS"')
        self.assertContains(response, 'value="2026-01-01"')
        self.assertContains(response, 'value="2026-12-31"')
        self.assertContains(response, 'value="aberto" selected')
        self.assertContains(response, 'value="Sudeste"')
        self.assertContains(response, 'value="Santos"')
        self.assertContains(response, f'value="{self.fornecedor.pk}" selected')

    def _criar_chamados_para_paginacao(self):
        for indice in range(21):
            self._criar_chamado_filtravel(
                f"OS-PAG-{indice:02d}",
                cidade="Cidade Pag",
                regional="Regional Pag",
                fornecedor=self.fornecedor,
                status=StatusChamado.PENDENTE,
                data_abertura=self._data_abertura(2026, 4, 1),
            )

    def test_paginacao_preserva_q(self):
        self._criar_chamados_para_paginacao()
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"q": "OS-PAG"}
        )
        self.assertIn("q=OS-PAG", response.context["querystring_filtros"])
        self.assertTrue(response.context["page_obj"].has_next())

    def test_paginacao_preserva_data_inicio_e_data_fim(self):
        self._criar_chamados_para_paginacao()
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {"data_inicio": "2026-04-01", "data_fim": "2026-04-30"},
        )
        self.assertIn("data_inicio=2026-04-01", response.context["querystring_filtros"])
        self.assertIn("data_fim=2026-04-30", response.context["querystring_filtros"])
        self.assertTrue(response.context["page_obj"].has_next())

    def test_paginacao_preserva_status(self):
        self._criar_chamados_para_paginacao()
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"status": "pendente"}
        )
        self.assertIn("status=pendente", response.context["querystring_filtros"])
        self.assertTrue(response.context["page_obj"].has_next())

    def test_paginacao_preserva_regional_cidade_e_fornecedor(self):
        self._criar_chamados_para_paginacao()
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {
                "regional": "Regional Pag",
                "cidade": "Cidade Pag",
                "fornecedor": str(self.fornecedor.pk),
            },
        )
        self.assertContains(response, "regional=Regional+Pag")
        self.assertContains(response, "cidade=Cidade+Pag")
        self.assertContains(response, f"fornecedor={self.fornecedor.pk}")
        self.assertContains(response, "page=2")

    def test_contador_total_filtrado_respeita_filtros(self):
        self._criar_chamado_filtravel("OS-COUNT-SUL", regional="Sul")
        self._criar_chamado_filtravel("OS-COUNT-NORTE", regional="Norte")
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"regional": "Sul"}
        )
        self.assertEqual(response.context["contadores"]["total"], 1)

    def test_contador_abertos_respeita_filtros(self):
        self._criar_chamado_filtravel(
            "OS-COUNT-ABERTO", regional="Contador Aberto", status=StatusChamado.ABERTO
        )
        self._criar_chamado_filtravel(
            "OS-COUNT-PENDENTE", regional="Contador Aberto", status=StatusChamado.PENDENTE
        )
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {"regional": "Contador Aberto"},
        )
        self.assertEqual(response.context["contadores"]["abertos"], 1)

    def test_contador_pendentes_respeita_filtros(self):
        self._criar_chamado_filtravel(
            "OS-COUNT-PEND", regional="Contador Pendente", status=StatusChamado.PENDENTE
        )
        self._criar_chamado_filtravel(
            "OS-COUNT-CONC", regional="Contador Pendente", status=StatusChamado.CONCLUIDO
        )
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {"regional": "Contador Pendente"},
        )
        self.assertEqual(response.context["contadores"]["pendentes"], 1)

    def test_contador_concluidos_respeita_filtros(self):
        self._criar_chamado_filtravel(
            "OS-COUNT-CONC-1",
            regional="Contador Concluido",
            status=StatusChamado.CONCLUIDO,
        )
        self._criar_chamado_filtravel(
            "OS-COUNT-CANC-1",
            regional="Contador Concluido",
            status=StatusChamado.CANCELADO,
        )
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {"regional": "Contador Concluido"},
        )
        self.assertEqual(response.context["contadores"]["concluidos"], 1)

    def test_contador_cancelados_respeita_filtros(self):
        self._criar_chamado_filtravel(
            "OS-COUNT-CANC-2",
            regional="Contador Cancelado",
            status=StatusChamado.CANCELADO,
        )
        self._criar_chamado_filtravel(
            "OS-COUNT-ABER-2",
            regional="Contador Cancelado",
            status=StatusChamado.ABERTO,
        )
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {"regional": "Contador Cancelado"},
        )
        self.assertEqual(response.context["contadores"]["cancelados"], 1)

    def test_contador_nao_emergenciais_respeita_filtros(self):
        self._criar_chamado_filtravel(
            "OS-COUNT-NE",
            regional="Contador NE",
            status=StatusChamado.NAO_EMERGENCIAL,
        )
        self._criar_chamado_filtravel(
            "OS-COUNT-ABER-NE", regional="Contador NE", status=StatusChamado.ABERTO
        )
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"regional": "Contador NE"}
        )
        self.assertEqual(response.context["contadores"]["nao_emergenciais"], 1)

    def test_contador_primeiro_report_pendente_conta_chamados_sem_atualizacoes(self):
        chamado_sem_report = self._criar_chamado_filtravel(
            "OS-COUNT-SEM-REPORT", regional="Contador Primeiro"
        )
        chamado_com_report = self._criar_chamado_filtravel(
            "OS-COUNT-COM-REPORT", regional="Contador Primeiro"
        )
        AtualizacaoChamado.objects.create(
            chamado=chamado_com_report,
            texto_atualizacao="Report registrado.",
        )
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {"regional": "Contador Primeiro"},
        )
        self.assertEqual(response.context["contadores"]["primeiro_report_pendente"], 1)
        self.assertContains(response, chamado_sem_report.numero_os)

    def test_link_limpar_filtros_existe(self):
        response = self.client.get(reverse("chamados:atualizar_report_list"))
        # Botao agora exibe apenas "Limpar" (texto reduzido para caber na linha unica de filtros)
        self.assertContains(response, "btn-clear")
        self.assertContains(response, "Limpar")

    def test_botao_exportar_excel_aparece_na_listagem(self):
        response = self.client.get(reverse("chamados:atualizar_report_list"))
        self.assertContains(response, "Exportar Excel")
        self.assertContains(response, reverse("chamados:atualizar_report_exportar_excel"))

    def test_form_get_status_200(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_form", args=[self.chamado.pk])
        )
        self.assertEqual(response.status_code, 200)

    def test_form_exibe_dados_principais_do_chamado(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_form", args=[self.chamado.pk])
        )
        self.assertContains(response, "OS-AR-001")
        self.assertContains(response, "Aberto")
        self.assertContains(response, "PRISMA-AR-01")
        self.assertContains(response, "Site Atualizar Report")
        self.assertContains(response, "Rua Report, 10")
        self.assertContains(response, "Santos")
        self.assertContains(response, "SP")
        self.assertContains(response, "Sudeste")
        self.assertContains(response, "Fornecedor Report")
        self.assertContains(response, "Falha de energia no site.")

    def test_sem_atualizacoes_exibe_aviso_primeiro_report(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_form", args=[self.chamado.pk])
        )
        self.assertContains(response, "Primeiro report desta OS")
        self.assertContains(response, "EMERGENCIAL ABERTA")

    def test_sem_atualizacoes_exibe_emergencial_aberta_na_previa(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_form", args=[self.chamado.pk])
        )
        self.assertContains(response, 'id="texto-whatsapp-preview"')
        self.assertContains(response, "EMERGENCIAL ABERTA")

    def test_com_atualizacao_anterior_exibe_aviso_atualizacao_status(self):
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Primeiro report.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
        )
        response = self.client.get(
            reverse("chamados:atualizar_report_form", args=[self.chamado.pk])
        )
        self.assertContains(response, "Atualização de status")
        self.assertContains(response, "já possui histórico")

    def test_formulario_nao_exibe_emergencial_aberta(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_form", args=[self.chamado.pk])
        )
        self.assertNotContains(response, 'value="emergencial_aberta"')

    def test_formulario_nao_exibe_abertura_emergencial(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_form", args=[self.chamado.pk])
        )
        self.assertNotContains(response, 'value="abertura_emergencial"')

    def test_botao_copiar_texto_whatsapp_aparece(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_form", args=[self.chamado.pk])
        )
        # Botao com texto reduzido ("Copiar") mas mesmo hook de copia.
        self.assertContains(response, 'data-copy-target="texto-whatsapp-preview"')
        self.assertContains(response, "btn-copy")

    def test_home_possui_link_funcional_para_atualizar_report(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertContains(response, reverse("chamados:atualizar_report_list"))

    def test_menu_lateral_possui_link_funcional_para_atualizar_report(self):
        response = self.client.get(reverse("chamados:home"))
        # O sidebar agora envolve o texto num <span> e tem ícone SVG; verificamos
        # apenas que a URL aparece como href e que o rótulo está presente.
        self.assertContains(
            response,
            f'href="{reverse("chamados:atualizar_report_list")}"',
        )
        self.assertContains(response, ">Atualizar Report</span>")

    def test_nenhuma_url_wa_me_aparece_nos_templates(self):
        templates_dir = Path(__file__).resolve().parent.parent / "templates"
        conteudo = "\n".join(
            path.read_text(encoding="utf-8") for path in templates_dir.rglob("*.html")
        )
        self.assertNotIn("wa.me", conteudo)

    def test_rota_exportacao_excel_foi_criada(self):
        self.assertEqual(
            reverse("chamados:atualizar_report_exportar_excel"),
            "/atualizar-report/exportar/",
        )

    def _exportar_excel(self, params=None):
        from io import BytesIO

        from openpyxl import load_workbook

        response = self.client.get(
            reverse("chamados:atualizar_report_exportar_excel"), params or {}
        )
        workbook = load_workbook(BytesIO(response.content))
        return response, workbook.active

    def _linhas_excel(self, params=None):
        _response, sheet = self._exportar_excel(params)
        return list(sheet.iter_rows(values_only=True))

    def _numeros_os_excel(self, params=None):
        linhas = self._linhas_excel(params)
        return [linha[0] for linha in linhas[1:]]

    def _indice_coluna_excel(self, cabecalho):
        linhas = self._linhas_excel()
        return linhas[0].index(cabecalho)

    def test_exportacao_excel_status_200(self):
        response = self.client.get(reverse("chamados:atualizar_report_exportar_excel"))
        self.assertEqual(response.status_code, 200)

    def test_exportacao_excel_content_type(self):
        response = self.client.get(reverse("chamados:atualizar_report_exportar_excel"))
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_exportacao_excel_content_disposition_xlsx(self):
        response = self.client.get(reverse("chamados:atualizar_report_exportar_excel"))
        self.assertIn(".xlsx", response["Content-Disposition"])

    def test_excel_contem_cabecalhos_obrigatorios(self):
        linhas = self._linhas_excel()
        self.assertEqual(
            list(linhas[0]),
            [
                "Número da OS",
                "Status atual",
                "Data de abertura",
                "Ativo Prisma",
                "Nome do site",
                "Endereço",
                "Cidade",
                "UF",
                "Regional",
                "Tipo de imóvel",
                "Tipo site/SLA",
                "Líder coordenação",
                "Fornecedor",
                "Solicitante",
                "Contato solicitante",
                "Command Center",
                "Denominação",
                "Ação tomada",
                "Detalhamento da situação",
                "Primeiro report pendente",
                "Quantidade de atualizações",
                "Última atualização",
                "Data da última atualização",
                "Status da última atualização",
                "Criado em",
                "Atualizado em",
            ],
        )

    def test_exportacao_sem_filtros_traz_chamados_existentes(self):
        self.assertIn("OS-AR-001", self._numeros_os_excel())

    def test_exportacao_respeita_filtro_q_por_numero_os(self):
        self._criar_chamado_filtravel("OS-EXP-Q-1")
        self._criar_chamado_filtravel("OS-EXP-Q-2")
        numeros = self._numeros_os_excel({"q": "OS-EXP-Q-1"})
        self.assertIn("OS-EXP-Q-1", numeros)
        self.assertNotIn("OS-EXP-Q-2", numeros)

    def test_exportacao_respeita_filtro_q_por_ativo_prisma(self):
        self._criar_chamado_filtravel("OS-EXP-PRISMA-1", ativo_prisma="PRISMA-EXP-1")
        self._criar_chamado_filtravel("OS-EXP-PRISMA-2", ativo_prisma="PRISMA-EXP-2")
        numeros = self._numeros_os_excel({"q": "PRISMA-EXP-1"})
        self.assertIn("OS-EXP-PRISMA-1", numeros)
        self.assertNotIn("OS-EXP-PRISMA-2", numeros)

    def test_exportacao_respeita_data_inicio(self):
        self._criar_chamado_filtravel(
            "OS-EXP-DI-ANTES", data_abertura=self._data_abertura(2026, 1, 10)
        )
        self._criar_chamado_filtravel(
            "OS-EXP-DI-DEPOIS", data_abertura=self._data_abertura(2026, 2, 10)
        )
        numeros = self._numeros_os_excel({"data_inicio": "2026-02-01"})
        self.assertIn("OS-EXP-DI-DEPOIS", numeros)
        self.assertNotIn("OS-EXP-DI-ANTES", numeros)

    def test_exportacao_respeita_data_fim(self):
        self._criar_chamado_filtravel(
            "OS-EXP-DF-ANTES", data_abertura=self._data_abertura(2026, 1, 10)
        )
        self._criar_chamado_filtravel(
            "OS-EXP-DF-DEPOIS", data_abertura=self._data_abertura(2026, 2, 10)
        )
        numeros = self._numeros_os_excel({"data_fim": "2026-01-31"})
        self.assertIn("OS-EXP-DF-ANTES", numeros)
        self.assertNotIn("OS-EXP-DF-DEPOIS", numeros)

    def test_exportacao_respeita_intervalo_datas(self):
        self._criar_chamado_filtravel(
            "OS-EXP-PER-ANTES", data_abertura=self._data_abertura(2026, 1, 10)
        )
        self._criar_chamado_filtravel(
            "OS-EXP-PER-MEIO", data_abertura=self._data_abertura(2026, 2, 10)
        )
        self._criar_chamado_filtravel(
            "OS-EXP-PER-DEPOIS", data_abertura=self._data_abertura(2026, 3, 10)
        )
        numeros = self._numeros_os_excel(
            {"data_inicio": "2026-02-01", "data_fim": "2026-02-28"}
        )
        self.assertIn("OS-EXP-PER-MEIO", numeros)
        self.assertNotIn("OS-EXP-PER-ANTES", numeros)
        self.assertNotIn("OS-EXP-PER-DEPOIS", numeros)

    def test_exportacao_respeita_status_aberto(self):
        self._criar_chamado_filtravel("OS-EXP-ST-ABERTO", status=StatusChamado.ABERTO)
        self._criar_chamado_filtravel("OS-EXP-ST-PEND", status=StatusChamado.PENDENTE)
        numeros = self._numeros_os_excel({"status": "aberto"})
        self.assertIn("OS-EXP-ST-ABERTO", numeros)
        self.assertNotIn("OS-EXP-ST-PEND", numeros)

    def test_exportacao_respeita_status_pendente(self):
        self._criar_chamado_filtravel("OS-EXP-ST-PEND-2", status=StatusChamado.PENDENTE)
        self._criar_chamado_filtravel("OS-EXP-ST-CONC-2", status=StatusChamado.CONCLUIDO)
        numeros = self._numeros_os_excel({"status": "pendente"})
        self.assertIn("OS-EXP-ST-PEND-2", numeros)
        self.assertNotIn("OS-EXP-ST-CONC-2", numeros)

    def test_exportacao_respeita_status_concluido(self):
        self._criar_chamado_filtravel("OS-EXP-ST-CONC-3", status=StatusChamado.CONCLUIDO)
        self._criar_chamado_filtravel("OS-EXP-ST-CANC-3", status=StatusChamado.CANCELADO)
        numeros = self._numeros_os_excel({"status": "concluido"})
        self.assertIn("OS-EXP-ST-CONC-3", numeros)
        self.assertNotIn("OS-EXP-ST-CANC-3", numeros)

    def test_exportacao_respeita_status_cancelado(self):
        self._criar_chamado_filtravel("OS-EXP-ST-CANC-4", status=StatusChamado.CANCELADO)
        self._criar_chamado_filtravel("OS-EXP-ST-ABER-4", status=StatusChamado.ABERTO)
        numeros = self._numeros_os_excel({"status": "cancelado"})
        self.assertIn("OS-EXP-ST-CANC-4", numeros)
        self.assertNotIn("OS-EXP-ST-ABER-4", numeros)

    def test_exportacao_respeita_status_nao_emergencial(self):
        self._criar_chamado_filtravel(
            "OS-EXP-ST-NE-5", status=StatusChamado.NAO_EMERGENCIAL
        )
        self._criar_chamado_filtravel("OS-EXP-ST-ABER-5", status=StatusChamado.ABERTO)
        numeros = self._numeros_os_excel({"status": "nao_emergencial"})
        self.assertIn("OS-EXP-ST-NE-5", numeros)
        self.assertNotIn("OS-EXP-ST-ABER-5", numeros)

    def test_exportacao_respeita_regional(self):
        self._criar_chamado_filtravel("OS-EXP-REG-SUL", regional="Sul")
        self._criar_chamado_filtravel("OS-EXP-REG-NORTE", regional="Norte")
        numeros = self._numeros_os_excel({"regional": "Sul"})
        self.assertIn("OS-EXP-REG-SUL", numeros)
        self.assertNotIn("OS-EXP-REG-NORTE", numeros)

    def test_exportacao_respeita_cidade(self):
        self._criar_chamado_filtravel("OS-EXP-CID-SANTOS", cidade="Santos")
        self._criar_chamado_filtravel("OS-EXP-CID-CAMPINAS", cidade="Campinas")
        numeros = self._numeros_os_excel({"cidade": "Campinas"})
        self.assertIn("OS-EXP-CID-CAMPINAS", numeros)
        self.assertNotIn("OS-EXP-CID-SANTOS", numeros)

    def test_exportacao_respeita_fornecedor(self):
        fornecedor_a = Fornecedor.objects.create(nome="Fornecedor Export A")
        fornecedor_b = Fornecedor.objects.create(nome="Fornecedor Export B")
        self._criar_chamado_filtravel("OS-EXP-FOR-A", fornecedor=fornecedor_a)
        self._criar_chamado_filtravel("OS-EXP-FOR-B", fornecedor=fornecedor_b)
        numeros = self._numeros_os_excel({"fornecedor": str(fornecedor_a.pk)})
        self.assertIn("OS-EXP-FOR-A", numeros)
        self.assertNotIn("OS-EXP-FOR-B", numeros)

    def test_exportacao_status_invalido_nao_gera_erro_500(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_exportar_excel"),
            {"status": "abertura_emergencial"},
        )
        self.assertEqual(response.status_code, 200)

    def test_exportacao_fornecedor_invalido_nao_gera_erro_500(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_exportar_excel"), {"fornecedor": "abc"}
        )
        self.assertEqual(response.status_code, 200)

    def test_excel_contem_coluna_primeiro_report_pendente(self):
        self._indice_coluna_excel("Primeiro report pendente")

    def test_excel_marca_sim_para_chamado_sem_atualizacoes(self):
        linhas = self._linhas_excel({"q": "OS-AR-001"})
        indice = linhas[0].index("Primeiro report pendente")
        self.assertEqual(linhas[1][indice], "Sim")

    def test_excel_marca_nao_para_chamado_com_atualizacoes(self):
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Report existente.",
        )
        linhas = self._linhas_excel({"q": "OS-AR-001"})
        indice = linhas[0].index("Primeiro report pendente")
        self.assertEqual(linhas[1][indice], "Não")

    def test_excel_contem_quantidade_atualizacoes_correta(self):
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Primeira.",
        )
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Segunda.",
        )
        linhas = self._linhas_excel({"q": "OS-AR-001"})
        indice = linhas[0].index("Quantidade de atualizações")
        self.assertEqual(linhas[1][indice], 2)

    def test_excel_contem_ultima_atualizacao_quando_existir(self):
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Primeira atualização.",
        )
        AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Última atualização exportada.",
        )
        linhas = self._linhas_excel({"q": "OS-AR-001"})
        indice = linhas[0].index("Última atualização")
        self.assertEqual(linhas[1][indice], "Última atualização exportada.")

    def test_excel_usa_hifen_quando_nao_houver_ultima_atualizacao(self):
        linhas = self._linhas_excel({"q": "OS-AR-001"})
        indice = linhas[0].index("Última atualização")
        self.assertEqual(linhas[1][indice], "-")

    def test_link_exportar_preserva_q(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"q": "OS-LINK"}
        )
        self.assertContains(response, "/atualizar-report/exportar/?q=OS-LINK")

    def test_link_exportar_preserva_data_inicio_e_data_fim(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {"data_inicio": "2026-05-01", "data_fim": "2026-05-31"},
        )
        self.assertContains(response, "data_inicio=2026-05-01")
        self.assertContains(response, "data_fim=2026-05-31")

    def test_link_exportar_preserva_status(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_list"), {"status": "pendente"}
        )
        self.assertContains(response, "/atualizar-report/exportar/?status=pendente")

    def test_link_exportar_preserva_regional_cidade_e_fornecedor(self):
        response = self.client.get(
            reverse("chamados:atualizar_report_list"),
            {
                "regional": "Litoral",
                "cidade": "Santos",
                "fornecedor": str(self.fornecedor.pk),
            },
        )
        self.assertContains(response, "regional=Litoral")
        self.assertContains(response, "cidade=Santos")
        self.assertContains(response, f"fornecedor={self.fornecedor.pk}")

    def _post_report(self, **data):
        follow = data.pop("follow", False)
        payload = {"texto_atualizacao": "Equipe em deslocamento."}
        payload.update(data)
        return self.client.post(
            reverse("chamados:atualizar_report_form", args=[self.chamado.pk]),
            data=payload,
            follow=follow,
        )

    def _criar_primeiro_report(self):
        return AtualizacaoChamado.objects.create(
            chamado=self.chamado,
            texto_atualizacao="Primeiro report.",
            tipo_evento=TipoEventoAtualizacao.ABERTURA_EMERGENCIAL,
            status_resultante=StatusChamado.ABERTO,
        )

    def test_post_primeiro_report_sem_status_cria_atualizacao(self):
        response = self._post_report()
        self.assertRedirects(
            response,
            reverse("chamados:atualizar_report_form", args=[self.chamado.pk]),
        )
        self.assertEqual(AtualizacaoChamado.objects.count(), 1)

    def test_post_primeiro_report_sem_status_cria_abertura_emergencial(self):
        self._post_report()
        atualizacao = AtualizacaoChamado.objects.get()
        self.assertEqual(
            atualizacao.tipo_evento, TipoEventoAtualizacao.ABERTURA_EMERGENCIAL
        )

    def test_post_primeiro_report_sem_status_assume_aberto(self):
        self._post_report()
        atualizacao = AtualizacaoChamado.objects.get()
        self.chamado.refresh_from_db()
        self.assertEqual(atualizacao.status_resultante, StatusChamado.ABERTO)
        self.assertEqual(self.chamado.status, StatusChamado.ABERTO)

    def test_post_primeiro_report_pendente_atualiza_status_chamado(self):
        self._post_report(status_resultante=StatusChamado.PENDENTE)
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, StatusChamado.PENDENTE)

    def test_post_primeiro_report_pendente_grava_status_resultante(self):
        self._post_report(status_resultante=StatusChamado.PENDENTE)
        atualizacao = AtualizacaoChamado.objects.get()
        self.assertEqual(atualizacao.status_resultante, StatusChamado.PENDENTE)

    def test_post_segunda_atualizacao_cria_tipo_atualizacao_status(self):
        self._criar_primeiro_report()
        self._post_report(status_resultante=StatusChamado.PENDENTE)
        atualizacao = AtualizacaoChamado.objects.order_by("criado_em").last()
        self.assertEqual(
            atualizacao.tipo_evento, TipoEventoAtualizacao.ATUALIZACAO_STATUS
        )

    def test_post_segunda_atualizacao_exige_status_resultante(self):
        self._criar_primeiro_report()
        response = self._post_report(status_resultante="")
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "Status resultante é obrigatório para atualizações posteriores.",
        )

    def test_post_segunda_atualizacao_sem_status_nao_cria_nova_atualizacao(self):
        self._criar_primeiro_report()
        self._post_report(status_resultante="")
        self.assertEqual(AtualizacaoChamado.objects.count(), 1)

    def test_post_segunda_atualizacao_concluido_altera_status(self):
        self._criar_primeiro_report()
        self._post_report(status_resultante=StatusChamado.CONCLUIDO)
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, StatusChamado.CONCLUIDO)

    def test_post_segunda_atualizacao_cancelado_altera_status(self):
        self._criar_primeiro_report()
        self._post_report(status_resultante=StatusChamado.CANCELADO)
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, StatusChamado.CANCELADO)

    def test_post_segunda_atualizacao_nao_emergencial_altera_status(self):
        self._criar_primeiro_report()
        self._post_report(status_resultante=StatusChamado.NAO_EMERGENCIAL)
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, StatusChamado.NAO_EMERGENCIAL)

    def test_post_texto_vazio_nao_cria_atualizacao(self):
        response = self._post_report(texto_atualizacao="")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(AtualizacaoChamado.objects.count(), 0)

    def test_post_texto_vazio_nao_altera_status(self):
        status_antes = self.chamado.status
        self._post_report(texto_atualizacao="", status_resultante=StatusChamado.PENDENTE)
        self.chamado.refresh_from_db()
        self.assertEqual(self.chamado.status, status_antes)

    def test_post_status_invalido_nao_cria_atualizacao(self):
        response = self._post_report(status_resultante="status_invalido")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(AtualizacaoChamado.objects.count(), 0)

    def test_post_status_invalido_exibe_erro_no_form(self):
        response = self._post_report(status_resultante="status_invalido")
        form = response.context["form"]
        self.assertIn("status_resultante", form.errors)

    def test_value_error_de_registrar_report_exibe_erro_no_form(self):
        from unittest.mock import patch

        with patch(
            "chamados.views.services.registrar_report",
            side_effect=ValueError("Erro controlado do service."),
        ):
            response = self._post_report(status_resultante=StatusChamado.PENDENTE)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Erro controlado do service.")

    def test_post_sucesso_redireciona_corretamente(self):
        response = self._post_report(status_resultante=StatusChamado.PENDENTE)
        self.assertRedirects(
            response,
            reverse("chamados:atualizar_report_form", args=[self.chamado.pk]),
        )

    def test_post_sucesso_recarrega_previa_com_status_atualizado(self):
        response = self._post_report(
            texto_atualizacao="Primeiro report em andamento.",
            status_resultante=StatusChamado.PENDENTE,
            follow=True,
        )
        self.assertContains(response, "EMERGENCIAL PENDENTE")
        self.assertContains(response, "Primeiro report em andamento.")

    def test_view_usa_registrar_report_para_criar_atualizacao(self):
        from unittest.mock import patch

        from . import services

        with patch(
            "chamados.views.services.registrar_report",
            wraps=services.registrar_report,
        ) as registrar_mock:
            self._post_report(status_resultante=StatusChamado.PENDENTE)

        self.assertTrue(registrar_mock.called)

    def test_view_nao_chama_atualizacao_objects_create_diretamente(self):
        import inspect

        from . import views

        codigo = inspect.getsource(views.atualizar_report_form)
        self.assertNotIn("AtualizacaoChamado.objects.create", codigo)


# =============================================================================
# Importadores flexíveis (detecção fuzzy de cabeçalho)
# =============================================================================

class ImportadorAtivosFlexivelTests(TestCase):
    """Verifica que o importador de Ativos aceita planilhas com cabeçalhos
    arbitrários (sinônimos), em qualquer ordem, com colunas extras ignoradas
    e tolerante a uma linha de título antes do header."""

    def _planilha_em_memoria(self, linhas):
        """Cria uma planilha .xlsx em memória a partir de uma lista de linhas."""
        from io import BytesIO
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        for linha in linhas:
            ws.append(list(linha))
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_planilha_do_usuario_e_aceita(self):
        """Cabeçalhos reais da planilha do parque imobiliário do usuário."""
        from .importadores import importar_ativos_excel

        cabecalhos = [
            "Ativo prisma", "CP", "Nomenclatura Prisma", "Denominação do Imóvel",
            "Endereco", "Cidade", "UF", "Região", "Lote", "Imóvel",
            "Cluster", "Tipologia", "Tipo Atendimento", "Area Construida m²",
            "Area do Terreno m²", "Area Livre m²", "Pavimentos",
            "População Fixa", "População Flutuante (diaria)",
            "Sugestao Equipe Residente (SIM / NÃO)",
            "Horario de Operação do Iovel (h)",
        ]
        linha_dados = [
            "11169", "95010", "LOJA SHOPPING IGUA", "LOJA SHOPPING IGUATEMI ALPHAVILLE",
            "ALAMEDA RIO NEGRO, 111", "Barueri", "SP", "CGSA", "1.1A", "LOJA",
            "LOJA", "A", "FIXO", "90,00", "-", "-", "2", "", "", "NÃO", "",
        ]
        arquivo = self._planilha_em_memoria([cabecalhos, linha_dados])
        resultado = importar_ativos_excel(arquivo)

        self.assertEqual(resultado.total_linhas, 1)
        self.assertEqual(resultado.criados, 1)
        self.assertEqual(resultado.ignorados, 0)
        self.assertEqual(resultado.erros, [])
        # Confere que o mapeamento foi montado
        self.assertIn("Ativo Prisma", resultado.mapeamento)
        self.assertEqual(resultado.mapeamento["Ativo Prisma"], "Ativo prisma")
        self.assertIn("Endereço", resultado.mapeamento)
        self.assertEqual(resultado.mapeamento["Endereço"], "Endereco")
        self.assertIn("UF", resultado.mapeamento)
        self.assertIn("Regional", resultado.mapeamento)
        # E que ativo foi realmente persistido com os dados certos
        ativo = Ativo.objects.get(ativo_prisma="11169")
        self.assertEqual(ativo.cidade, "Barueri")
        self.assertEqual(ativo.uf, "SP")
        self.assertEqual(ativo.regional, "CGSA")
        self.assertEqual(ativo.tipo_imovel, "LOJA")
        self.assertEqual(ativo.tipo_site_sla, "FIXO")

    def test_colunas_extras_sao_ignoradas_silenciosamente(self):
        from .importadores import importar_ativos_excel

        arquivo = self._planilha_em_memoria([
            ["Ativo prisma", "Coluna Inventada", "Nome do Site", "Endereco",
             "Cidade", "UF", "Regional", "Outra Coluna Inutil"],
            ["X-1", "lixo", "Site X", "Rua X", "Sao Paulo", "SP", "Sudeste", "lixo2"],
        ])
        resultado = importar_ativos_excel(arquivo)
        self.assertEqual(resultado.criados, 1)
        self.assertEqual(resultado.ignorados, 0)
        self.assertIn("Coluna Inventada", resultado.cabecalhos_ignorados)
        self.assertIn("Outra Coluna Inutil", resultado.cabecalhos_ignorados)

    def test_ordem_aleatoria_das_colunas_funciona(self):
        from .importadores import importar_ativos_excel

        # UF antes de Ativo Prisma; tudo embaralhado
        arquivo = self._planilha_em_memoria([
            ["UF", "Cidade", "Endereco", "Ativo Prisma", "Regional", "Nome do Site"],
            ["RJ", "Rio", "Av X", "Y-99", "Sudeste", "Site Y"],
        ])
        resultado = importar_ativos_excel(arquivo)
        self.assertEqual(resultado.criados, 1)
        self.assertEqual(resultado.ignorados, 0)
        ativo = Ativo.objects.get(ativo_prisma="Y-99")
        self.assertEqual(ativo.uf, "RJ")
        self.assertEqual(ativo.cidade, "Rio")

    def test_header_em_row_2_quando_existe_titulo_antes(self):
        from .importadores import importar_ativos_excel

        # Row 1 é um título; row 2 são os cabeçalhos reais.
        arquivo = self._planilha_em_memoria([
            ["Parque Imobiliário · Maio/2026", None, None, None, None, None],
            ["Ativo Prisma", "Nome do Site", "Endereco", "Cidade", "UF", "Regional"],
            ["Z-1", "Site Z", "Rua Z", "Curitiba", "PR", "Sul"],
        ])
        resultado = importar_ativos_excel(arquivo)
        self.assertEqual(resultado.criados, 1)
        self.assertEqual(resultado.ignorados, 0)
        self.assertTrue(Ativo.objects.filter(ativo_prisma="Z-1").exists())

    def test_planilha_sem_cabecalho_reconhecivel_aborta_limpo(self):
        from .importadores import importar_ativos_excel

        arquivo = self._planilha_em_memoria([
            ["Coluna A", "Coluna B", "Coluna C"],
            ["x", "y", "z"],
        ])
        resultado = importar_ativos_excel(arquivo)
        self.assertEqual(resultado.criados, 0)
        self.assertEqual(resultado.atualizados, 0)
        self.assertEqual(resultado.total_linhas, 0)
        self.assertEqual(len(resultado.erros), 1)
        self.assertIn("identificar", resultado.erros[0].lower())

    def test_planilha_sem_regional_ainda_e_aceita(self):
        """Regional não é mais obrigatória — linha é salva e contabilizada
        em campos_vazios."""
        from .importadores import importar_ativos_excel

        arquivo = self._planilha_em_memoria([
            ["Ativo Prisma", "Nome do Site", "Endereco", "Cidade", "UF"],
            ["W-1", "Site W", "Rua W", "Recife", "PE"],
        ])
        resultado = importar_ativos_excel(arquivo)
        self.assertEqual(resultado.criados, 1)
        self.assertEqual(resultado.ignorados, 0)
        self.assertEqual(resultado.campos_vazios.get("regional", 0), 1)
        ativo = Ativo.objects.get(ativo_prisma="W-1")
        self.assertEqual(ativo.regional, "")

    def test_acentos_e_variantes_no_cabecalho(self):
        from .importadores import importar_ativos_excel

        # Com acentos no Endereço/Região e CP que deve ser ignorado.
        arquivo = self._planilha_em_memoria([
            ["ativo prisma", "Denominação do Imóvel", "Endereço",
             "Município", "UF", "Região", "Imóvel"],
            ["A-1", "Loja A", "Av A", "Belém", "PA", "Norte", "Loja"],
        ])
        resultado = importar_ativos_excel(arquivo)
        self.assertEqual(resultado.criados, 1)
        ativo = Ativo.objects.get(ativo_prisma="A-1")
        self.assertEqual(ativo.nome_site, "Loja A")
        self.assertEqual(ativo.cidade, "Belém")
        self.assertEqual(ativo.regional, "Norte")
        self.assertEqual(ativo.tipo_imovel, "Loja")

    def test_compatibilidade_com_aliases_legados(self):
        """Os aliases antigos (importação anterior) continuam funcionando."""
        from .importadores import importar_ativos_excel

        arquivo = self._planilha_em_memoria([
            ["Prisma", "Site", "Logradouro", "Cidade", "Estado", "Regional"],
            ["LEG-1", "Site Legado", "Rua L", "Salvador", "BA", "Nordeste"],
        ])
        resultado = importar_ativos_excel(arquivo)
        self.assertEqual(resultado.criados, 1)
        ativo = Ativo.objects.get(ativo_prisma="LEG-1")
        self.assertEqual(ativo.uf, "BA")

    def test_cidade_vazia_inferida_do_endereco(self):
        """Quando a coluna Cidade está vazia mas o endereço tem o padrão
        'Rua X - Bairro - Cidade', o sistema preenche cidade automaticamente.
        Se a inferência falhar, salva mesmo assim com cidade vazia (modo
        maximamente permissivo)."""
        from .importadores import importar_ativos_excel

        arquivo = self._planilha_em_memoria([
            ["Ativo Prisma", "Nome do Site", "Endereco", "Cidade", "UF", "Regional"],
            # Cidade vazia, endereço com padrão " - ... - CIDADE"
            ["INF-1", "Site A", "Av Paulista, 100 - Bela Vista - São Paulo", "", "SP", "Sudeste"],
            # Cidade vazia, endereço sem separador (inferência falha; ainda salva)
            ["INF-2", "Site B", "Rua única sem separador", "", "PR", "Sul"],
            # Cidade preenchida (a inferência não deve sobrescrever)
            ["INF-3", "Site C", "Rua Z - bairro - Salvador", "Recife", "PE", "Nordeste"],
        ])
        resultado = importar_ativos_excel(arquivo)
        # Todas as 3 linhas são salvas — INF-2 entra com cidade vazia.
        self.assertEqual(resultado.criados, 3)
        self.assertEqual(resultado.ignorados, 0)
        self.assertEqual(resultado.cidades_inferidas_do_endereco, 1)
        # Conta a linha que ficou sem cidade
        self.assertEqual(resultado.campos_vazios.get("cidade", 0), 1)
        # INF-1 foi salvo com cidade inferida
        ativo_inf = Ativo.objects.get(ativo_prisma="INF-1")
        self.assertEqual(ativo_inf.cidade, "São Paulo")
        # INF-2 entra mas com cidade vazia
        ativo_vazio = Ativo.objects.get(ativo_prisma="INF-2")
        self.assertEqual(ativo_vazio.cidade, "")
        # INF-3 mantém a cidade da planilha (não foi sobrescrita)
        ativo_explicito = Ativo.objects.get(ativo_prisma="INF-3")
        self.assertEqual(ativo_explicito.cidade, "Recife")

    def test_extracao_de_cidade_helper_isolado(self):
        from .importadores import _extrair_cidade_do_endereco

        # Casos onde extrai
        self.assertEqual(
            _extrair_cidade_do_endereco(
                "ALAMEDA RIO NEGRO, 111 - LOJA 114 - PISO TER BARUERI"
            ),
            "PISO TER BARUERI",
        )
        self.assertEqual(
            _extrair_cidade_do_endereco("Rua A, 1 - Centro - Curitiba"),
            "Curitiba",
        )
        # Casos onde NÃO deve extrair
        self.assertEqual(_extrair_cidade_do_endereco(""), "")
        self.assertEqual(_extrair_cidade_do_endereco("Rua única sem separador"), "")
        self.assertEqual(_extrair_cidade_do_endereco("Rua X - 123"), "")  # último é numérico


class ImportadorFornecedoresFlexivelTests(TestCase):
    def _planilha_em_memoria(self, linhas):
        from io import BytesIO
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        for linha in linhas:
            ws.append(list(linha))
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_aceita_planilha_com_colunas_extras_e_sinonimos(self):
        from .importadores import importar_fornecedores_excel

        arquivo = self._planilha_em_memoria([
            ["Nome do Coordenador", "Cel", "E-mail", "Razão Social",
             "UFs atendidas", "Coluna Lixo"],
            ["João Silva", "11999998888", "j@x.com", "TechAtende", "SP, RJ", "lixo"],
        ])
        resultado = importar_fornecedores_excel(arquivo)
        self.assertEqual(resultado.criados, 1)
        self.assertEqual(resultado.ignorados, 0)
        self.assertIn("Nome", resultado.mapeamento)
        forn = Fornecedor.objects.get(nome="João Silva")
        self.assertEqual(forn.empresa, "TechAtende")
        self.assertEqual(forn.estados_atendidos, "SP, RJ")


class ImportadorObrasFlexivelTests(TestCase):
    def setUp(self):
        self.ativo = Ativo.objects.create(
            ativo_prisma="PRISMA-OBRA-FLX",
            nome_site="Site Flex",
            endereco="R F, 1",
            cidade="SP",
            uf="SP",
            regional="Sudeste",
        )

    def _planilha_em_memoria(self, linhas):
        from io import BytesIO
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        for linha in linhas:
            ws.append(list(linha))
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_aceita_planilha_obra_com_sinonimos_e_colunas_extras(self):
        from .importadores import importar_obras_excel

        arquivo = self._planilha_em_memoria([
            ["Prisma", "Nome do site (info)", "Escopo da obra",
             "Início", "Previsão de Término", "Construtora", "Lixo"],
            ["PRISMA-OBRA-FLX", "Site Flex", "Pintura externa",
             "01/05/2026", "30/06/2026", "ABC Eng.", "x"],
        ])
        resultado = importar_obras_excel(arquivo)
        self.assertEqual(resultado.criados, 1, msg=str(resultado.erros))
        self.assertEqual(resultado.ignorados, 0)
        obra = Obra.objects.get(ativo=self.ativo, descricao="Pintura externa")
        self.assertEqual(obra.responsavel, "ABC Eng.")


# =============================================================================
# Importação parcial: preservação de dados existentes
# =============================================================================
# Estes testes garantem que uma planilha parcial NÃO apaga campos preenchidos
# de registros já existentes. A regra é: coluna ausente ou valor vazio
# preserva o valor do banco; só atualiza quando a coluna existe e veio com
# conteúdo. Criação de novos registros segue permissiva.


def _planilha_em_memoria(linhas):
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for linha in linhas:
        ws.append(list(linha))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ImportacaoParcialAtivosTests(TestCase):
    """Importação parcial de Ativos NÃO pode apagar campos existentes."""

    def test_planilha_so_com_ativo_prisma_preserva_todos_os_campos(self):
        """Cenário principal: planilha com SOMENTE 'Ativo Prisma' não pode
        zerar nome_site, endereco, cidade, uf, regional, tipo_imovel,
        lider_coordenacao nem tipo_site_sla de um Ativo existente."""
        from .importadores import importar_ativos_excel

        Ativo.objects.create(
            ativo_prisma="PRISMA-001",
            nome_site="Loja Original",
            endereco="Rua Original, 100",
            cidade="São Paulo",
            uf="SP",
            regional="Sudeste",
            tipo_imovel="LOJA",
            lider_coordenacao="Líder Original",
            tipo_site_sla="FIXO",
        )

        arquivo = _planilha_em_memoria([
            ["Ativo Prisma"],
            ["PRISMA-001"],
        ])
        resultado = importar_ativos_excel(arquivo)

        self.assertEqual(resultado.atualizados, 1, msg=str(resultado.erros))
        self.assertEqual(resultado.criados, 0)
        ativo = Ativo.objects.get(ativo_prisma="PRISMA-001")
        self.assertEqual(ativo.nome_site, "Loja Original")
        self.assertEqual(ativo.endereco, "Rua Original, 100")
        self.assertEqual(ativo.cidade, "São Paulo")
        self.assertEqual(ativo.uf, "SP")
        self.assertEqual(ativo.regional, "Sudeste")
        self.assertEqual(ativo.tipo_imovel, "LOJA")
        self.assertEqual(ativo.lider_coordenacao, "Líder Original")
        self.assertEqual(ativo.tipo_site_sla, "FIXO")

    def test_planilha_parcial_atualiza_so_coluna_presente_e_preserva_outras(self):
        """Planilha com 'Ativo Prisma' + 'Cidade' atualiza cidade e preserva
        as demais colunas que NÃO vieram no arquivo."""
        from .importadores import importar_ativos_excel

        Ativo.objects.create(
            ativo_prisma="PRISMA-002",
            nome_site="Site Antigo",
            endereco="Rua Antiga, 1",
            cidade="Santos",
            uf="SP",
            regional="Sudeste",
            tipo_imovel="LOJA",
            lider_coordenacao="Líder X",
            tipo_site_sla="FIXO",
        )

        arquivo = _planilha_em_memoria([
            ["Ativo Prisma", "Cidade"],
            ["PRISMA-002", "Campinas"],
        ])
        resultado = importar_ativos_excel(arquivo)

        self.assertEqual(resultado.atualizados, 1, msg=str(resultado.erros))
        ativo = Ativo.objects.get(ativo_prisma="PRISMA-002")
        self.assertEqual(ativo.cidade, "Campinas")  # atualizou
        # Demais campos preservados
        self.assertEqual(ativo.nome_site, "Site Antigo")
        self.assertEqual(ativo.endereco, "Rua Antiga, 1")
        self.assertEqual(ativo.uf, "SP")
        self.assertEqual(ativo.regional, "Sudeste")
        self.assertEqual(ativo.tipo_imovel, "LOJA")
        self.assertEqual(ativo.lider_coordenacao, "Líder X")
        self.assertEqual(ativo.tipo_site_sla, "FIXO")

    def test_planilha_parcial_continua_criando_ativo_novo(self):
        """Planilha com APENAS 'Ativo Prisma' deve criar um Ativo novo,
        mantendo o comportamento permissivo: campos ausentes ficam como ''."""
        from .importadores import importar_ativos_excel

        arquivo = _planilha_em_memoria([
            ["Ativo Prisma"],
            ["PRISMA-NEW-001"],
        ])
        resultado = importar_ativos_excel(arquivo)

        self.assertEqual(resultado.criados, 1, msg=str(resultado.erros))
        self.assertEqual(resultado.atualizados, 0)
        ativo = Ativo.objects.get(ativo_prisma="PRISMA-NEW-001")
        self.assertEqual(ativo.nome_site, "")
        self.assertEqual(ativo.endereco, "")
        self.assertEqual(ativo.cidade, "")
        self.assertEqual(ativo.uf, "")
        self.assertEqual(ativo.regional, "")

    def test_coluna_presente_com_valor_vazio_preserva_existente(self):
        """Mesmo com a coluna presente, se o valor da célula vier vazio,
        o valor antigo no banco deve ser preservado."""
        from .importadores import importar_ativos_excel

        Ativo.objects.create(
            ativo_prisma="PRISMA-003",
            nome_site="Mantém",
            endereco="Endereço Mantido",
            cidade="Curitiba",
            uf="PR",
            regional="Sul",
        )

        arquivo = _planilha_em_memoria([
            ["Ativo Prisma", "Cidade", "UF"],
            ["PRISMA-003", "", ""],
        ])
        resultado = importar_ativos_excel(arquivo)

        self.assertEqual(resultado.atualizados, 1, msg=str(resultado.erros))
        ativo = Ativo.objects.get(ativo_prisma="PRISMA-003")
        self.assertEqual(ativo.cidade, "Curitiba")
        self.assertEqual(ativo.uf, "PR")
        self.assertEqual(ativo.nome_site, "Mantém")
        self.assertEqual(ativo.endereco, "Endereço Mantido")


class ImportacaoParcialFornecedoresTests(TestCase):
    """Importação parcial de Fornecedores NÃO pode apagar campos existentes."""

    def test_planilha_so_com_nome_preserva_todos_os_campos(self):
        """Planilha com SOMENTE 'Nome' não pode zerar telefone, email,
        empresa nem estados_atendidos de um Fornecedor existente."""
        from .importadores import importar_fornecedores_excel

        Fornecedor.objects.create(
            nome="João Silva",
            telefone="11999998888",
            email="joao@empresa.com",
            empresa="Empresa Original",
            estados_atendidos="SP, RJ",
        )

        arquivo = _planilha_em_memoria([
            ["Nome"],
            ["João Silva"],
        ])
        resultado = importar_fornecedores_excel(arquivo)

        self.assertEqual(resultado.atualizados, 1, msg=str(resultado.erros))
        self.assertEqual(resultado.criados, 0)
        forn = Fornecedor.objects.get(nome="João Silva")
        self.assertEqual(forn.telefone, "11999998888")
        self.assertEqual(forn.email, "joao@empresa.com")
        self.assertEqual(forn.empresa, "Empresa Original")
        self.assertEqual(forn.estados_atendidos, "SP, RJ")

    def test_planilha_parcial_atualiza_so_coluna_presente_e_preserva_outras(self):
        """Planilha com 'Nome' + 'Telefone' atualiza telefone e preserva
        os demais campos do fornecedor."""
        from .importadores import importar_fornecedores_excel

        Fornecedor.objects.create(
            nome="Maria Souza",
            telefone="1133334444",
            email="maria@empresa.com",
            empresa="Empresa Beta",
            estados_atendidos="MG, ES",
        )

        arquivo = _planilha_em_memoria([
            ["Nome", "Telefone"],
            ["Maria Souza", "11955554444"],
        ])
        resultado = importar_fornecedores_excel(arquivo)

        self.assertEqual(resultado.atualizados, 1, msg=str(resultado.erros))
        forn = Fornecedor.objects.get(nome="Maria Souza")
        self.assertEqual(forn.telefone, "11955554444")  # atualizou
        # Demais campos preservados
        self.assertEqual(forn.email, "maria@empresa.com")
        self.assertEqual(forn.empresa, "Empresa Beta")
        self.assertEqual(forn.estados_atendidos, "MG, ES")

    def test_planilha_parcial_continua_criando_fornecedor_novo(self):
        """Planilha com APENAS 'Nome' deve criar um Fornecedor novo,
        mantendo o comportamento permissivo: campos ausentes ficam como ''."""
        from .importadores import importar_fornecedores_excel

        arquivo = _planilha_em_memoria([
            ["Nome"],
            ["Pedro Novo"],
        ])
        resultado = importar_fornecedores_excel(arquivo)

        self.assertEqual(resultado.criados, 1, msg=str(resultado.erros))
        self.assertEqual(resultado.atualizados, 0)
        forn = Fornecedor.objects.get(nome="Pedro Novo")
        self.assertEqual(forn.telefone, "")
        self.assertEqual(forn.email, "")
        self.assertEqual(forn.empresa, "")
        self.assertEqual(forn.estados_atendidos, "")

    def test_coluna_presente_com_valor_vazio_preserva_existente(self):
        """Mesmo com a coluna presente, célula vazia preserva valor do banco."""
        from .importadores import importar_fornecedores_excel

        Fornecedor.objects.create(
            nome="Carlos Teste",
            telefone="11988887777",
            email="carlos@empresa.com",
            empresa="Empresa Carlos",
            estados_atendidos="RS",
        )

        arquivo = _planilha_em_memoria([
            ["Nome", "Telefone", "E-mail"],
            ["Carlos Teste", "", ""],
        ])
        resultado = importar_fornecedores_excel(arquivo)

        self.assertEqual(resultado.atualizados, 1, msg=str(resultado.erros))
        forn = Fornecedor.objects.get(nome="Carlos Teste")
        self.assertEqual(forn.telefone, "11988887777")
        self.assertEqual(forn.email, "carlos@empresa.com")
        self.assertEqual(forn.empresa, "Empresa Carlos")
        self.assertEqual(forn.estados_atendidos, "RS")


# =============================================================================
# Importação com arquivo .xlsx inválido/corrompido
# =============================================================================
# Garante que um arquivo nomeado como .xlsx mas com conteúdo inválido NÃO
# gere erro 500: a view captura a exceção do openpyxl/zip e devolve a
# própria tela com mensagem amigável. Nenhum registro pode ser criado.


class _ImportacaoArquivoInvalidoMixin:
    """Helpers para os testes de arquivo .xlsx inválido."""

    URL_NAME: str = ""

    def _login(self):
        from django.contrib.auth import get_user_model

        User = get_user_model()
        usuario = User.objects.create_user(
            username="operador_import", password="senhaforte123"
        )
        self.client.force_login(usuario)
        return usuario

    def _arquivo_xlsx_invalido(self):
        # Conteúdo arbitrário que NÃO é um zip válido -> dispara BadZipFile
        # no openpyxl.
        return SimpleUploadedFile(
            "bad.xlsx",
            b"isto-nao-eh-um-xlsx-valido",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _postar_arquivo_invalido(self):
        self._login()
        return self.client.post(
            reverse(self.URL_NAME),
            {"arquivo": self._arquivo_xlsx_invalido()},
            follow=False,
        )


class ImportacaoAtivosArquivoInvalidoTests(_ImportacaoArquivoInvalidoMixin, TestCase):
    URL_NAME = "chamados:ativos_import"

    def test_arquivo_xlsx_invalido_nao_retorna_500(self):
        resposta = self._postar_arquivo_invalido()
        self.assertNotEqual(resposta.status_code, 500)
        self.assertEqual(resposta.status_code, 200)

    def test_arquivo_xlsx_invalido_nao_cria_ativos(self):
        antes = Ativo.objects.count()
        self._postar_arquivo_invalido()
        self.assertEqual(Ativo.objects.count(), antes)

    def test_arquivo_xlsx_invalido_exibe_mensagem_amigavel(self):
        resposta = self._postar_arquivo_invalido()
        conteudo = resposta.content.decode("utf-8", errors="ignore")
        self.assertIn("inv", conteudo.lower())  # "inválido"
        # E não deve ter mensagem de sucesso
        self.assertNotIn("Importação concluída", conteudo)


class ImportacaoFornecedoresArquivoInvalidoTests(
    _ImportacaoArquivoInvalidoMixin, TestCase
):
    URL_NAME = "chamados:fornecedores_import"

    def test_arquivo_xlsx_invalido_nao_retorna_500(self):
        resposta = self._postar_arquivo_invalido()
        self.assertNotEqual(resposta.status_code, 500)
        self.assertEqual(resposta.status_code, 200)

    def test_arquivo_xlsx_invalido_nao_cria_fornecedores(self):
        antes = Fornecedor.objects.count()
        self._postar_arquivo_invalido()
        self.assertEqual(Fornecedor.objects.count(), antes)

    def test_arquivo_xlsx_invalido_exibe_mensagem_amigavel(self):
        resposta = self._postar_arquivo_invalido()
        conteudo = resposta.content.decode("utf-8", errors="ignore")
        self.assertIn("inv", conteudo.lower())
        self.assertNotIn("Importação concluída", conteudo)


class ImportacaoObrasArquivoInvalidoTests(_ImportacaoArquivoInvalidoMixin, TestCase):
    URL_NAME = "chamados:obras_import"

    def test_arquivo_xlsx_invalido_nao_retorna_500(self):
        resposta = self._postar_arquivo_invalido()
        self.assertNotEqual(resposta.status_code, 500)
        self.assertEqual(resposta.status_code, 200)

    def test_arquivo_xlsx_invalido_nao_cria_obras(self):
        antes = Obra.objects.count()
        self._postar_arquivo_invalido()
        self.assertEqual(Obra.objects.count(), antes)

    def test_arquivo_xlsx_invalido_exibe_mensagem_amigavel(self):
        resposta = self._postar_arquivo_invalido()
        conteudo = resposta.content.decode("utf-8", errors="ignore")
        self.assertIn("inv", conteudo.lower())
        self.assertNotIn("Importação concluída", conteudo)


# =============================================================================
# Contrato de autenticação: telas internas exigem login
# =============================================================================
# Documenta explicitamente a regra mantida pelo LoginRequiredMiddleware:
# usuário anônimo é redirecionado para /login/ (302); usuário autenticado
# acessa normalmente (200). Não testa permissões — o sistema é aberto para
# qualquer usuário logado; o login serve apenas para identificar o operador.


class AutenticacaoTelasInternasContratoTests(TestCase):
    """Garante o contrato anônimo→302 / autenticado→200 em telas internas
    representativas. Se alguma view passar a ser pública por engano (ou se
    o middleware for afrouxado), estes testes falham e sinalizam o desvio.
    """

    TELAS_INTERNAS = (
        "chamados:home",
        "chamados:ativos_list",
        "chamados:ativos_import",
        "chamados:ativo_create",
        "chamados:fornecedores_list",
        "chamados:fornecedores_import",
        "chamados:chamados_list",
        "chamados:chamado_create",
        "chamados:atualizar_report_list",
        "chamados:obras_import",
    )

    def _login(self):
        from django.contrib.auth import get_user_model

        User = get_user_model()
        usuario = User.objects.create_user(
            username="operador_contrato", password="senhaforte123"
        )
        self.client.force_login(usuario)
        return usuario

    def test_anonimo_recebe_302_em_todas_telas_internas(self):
        for nome_rota in self.TELAS_INTERNAS:
            with self.subTest(rota=nome_rota):
                response = self.client.get(reverse(nome_rota))
                self.assertEqual(
                    response.status_code,
                    302,
                    msg=f"Esperado 302 em {nome_rota} para usuário anônimo.",
                )
                self.assertIn("/login/", response["Location"])

    def test_autenticado_recebe_200_em_todas_telas_internas(self):
        self._login()
        for nome_rota in self.TELAS_INTERNAS:
            with self.subTest(rota=nome_rota):
                response = self.client.get(reverse(nome_rota))
                self.assertEqual(
                    response.status_code,
                    200,
                    msg=f"Esperado 200 em {nome_rota} para usuário autenticado.",
                )


# =============================================================================
# Métricas da Home: coerência formal
# =============================================================================
# Documenta e trava o comportamento das métricas exibidas no hero da Home
# (Prédios x Lojas / Pendentes / Concluídos). A função de produção é
# ``chamados.views._metricas_home``.
#
# Regras testadas aqui (resumo):
#
# 1. Lojas       = ``Ativo.tipo_imovel`` contém "loja" (case-insensitive).
# 2. Prédios     = negação de Lojas (inclui ``tipo_imovel`` vazio).
# 3. Ativo.ativo = False também é contado nas métricas (inventário total).
# 4. Pendentes   = ``ABERTO`` ∪ ``PENDENTE``.
# 5. Concluídos  = somente ``CONCLUIDO``.
# 6. ``CANCELADO`` e ``NAO_EMERGENCIAL`` NÃO entram em pendentes nem em
#    concluídos.
# 7. Fornecedores e Obras NÃO compõem métrica no hero hoje.
#
# Qualquer mudança nessas regras deve quebrar testes desta classe — e isso
# é proposital: a métrica é um contrato com a operação e precisa ser
# alterada explicitamente.


class HomeMetricasCoerentesTests(_LoginClienteMixin, TestCase):
    """Garante coerência formal das métricas da Home."""

    def _criar_ativo(self, prisma, tipo_imovel="", ativo=True, **kwargs):
        dados = {
            "ativo_prisma": prisma,
            "nome_site": kwargs.pop("nome_site", f"Site {prisma}"),
            "endereco": kwargs.pop("endereco", "Rua T, 1"),
            "cidade": kwargs.pop("cidade", "São Paulo"),
            "uf": kwargs.pop("uf", "SP"),
            "regional": kwargs.pop("regional", "Sudeste"),
            "tipo_imovel": tipo_imovel,
            "ativo": ativo,
        }
        dados.update(kwargs)
        return Ativo.objects.create(**dados)

    def _criar_chamado(self, ativo, status, numero_os):
        return Chamado.objects.create(
            ativo=ativo,
            numero_os=numero_os,
            data_abertura=timezone.now(),
            status=status,
        )

    def _metricas(self):
        from .views import _metricas_home

        return _metricas_home()

    # ---------------- Prédios e Lojas: classificação por tipo_imovel ---------

    def test_ativo_com_loja_no_tipo_entra_em_lojas(self):
        antes = self._metricas()
        self._criar_ativo("M-LOJA-1", tipo_imovel="LOJA")
        depois = self._metricas()
        self.assertEqual(depois["lojas_total"], antes["lojas_total"] + 1)
        self.assertEqual(depois["predios_total"], antes["predios_total"])

    def test_ativo_com_loja_shopping_no_tipo_entra_em_lojas(self):
        """Sinônimo realista: 'LOJA SHOPPING IGUATEMI' contém 'loja'."""
        antes = self._metricas()
        self._criar_ativo("M-LOJA-2", tipo_imovel="LOJA SHOPPING IGUATEMI")
        depois = self._metricas()
        self.assertEqual(depois["lojas_total"], antes["lojas_total"] + 1)
        self.assertEqual(depois["predios_total"], antes["predios_total"])

    def test_ativo_com_predio_no_tipo_entra_em_predios(self):
        antes = self._metricas()
        self._criar_ativo("M-PRED-1", tipo_imovel="PREDIO")
        depois = self._metricas()
        self.assertEqual(depois["predios_total"], antes["predios_total"] + 1)
        self.assertEqual(depois["lojas_total"], antes["lojas_total"])

    def test_ativo_com_tipos_diversos_nao_loja_entram_em_predios(self):
        """'TÉCNICO', 'SEDE', 'CD' — qualquer coisa fora de loja vira prédio."""
        antes = self._metricas()
        self._criar_ativo("M-TEC-1", tipo_imovel="TÉCNICO")
        self._criar_ativo("M-SEDE-1", tipo_imovel="SEDE ADMINISTRATIVA")
        self._criar_ativo("M-CD-1", tipo_imovel="CD")
        depois = self._metricas()
        self.assertEqual(depois["predios_total"], antes["predios_total"] + 3)
        self.assertEqual(depois["lojas_total"], antes["lojas_total"])

    # ---------------- Tipo vazio ---------------------------------------------

    def test_ativo_com_tipo_imovel_vazio_entra_em_predios(self):
        """REGRA ATUAL EXPLÍCITA: tipo_imovel='' cai em Prédios (negação de
        'loja' inclui vazio). Não há bucket 'sem classificação' hoje."""
        antes = self._metricas()
        self._criar_ativo("M-VAZIO-1", tipo_imovel="")
        depois = self._metricas()
        self.assertEqual(depois["predios_total"], antes["predios_total"] + 1)
        self.assertEqual(depois["lojas_total"], antes["lojas_total"])

    # ---------------- Ativo inativo ------------------------------------------

    def test_ativo_inativo_loja_continua_contando_em_lojas(self):
        """REGRA ATUAL EXPLÍCITA: ``Ativo.ativo=False`` também é contado.
        Métrica é inventário do parque, não 'parque operacional'."""
        antes = self._metricas()
        self._criar_ativo("M-INA-L", tipo_imovel="LOJA", ativo=False)
        depois = self._metricas()
        self.assertEqual(depois["lojas_total"], antes["lojas_total"] + 1)

    def test_ativo_inativo_predio_continua_contando_em_predios(self):
        antes = self._metricas()
        self._criar_ativo("M-INA-P", tipo_imovel="PREDIO", ativo=False)
        depois = self._metricas()
        self.assertEqual(depois["predios_total"], antes["predios_total"] + 1)

    # ---------------- Chamados: pendentes vs concluídos ----------------------

    def test_chamado_aberto_conta_em_pendentes(self):
        ativo = self._criar_ativo("M-CH-AB", tipo_imovel="LOJA")
        antes = self._metricas()
        self._criar_chamado(ativo, StatusChamado.ABERTO, "OS-AB-1")
        depois = self._metricas()
        self.assertEqual(
            depois["lojas_pendentes"], antes["lojas_pendentes"] + 1
        )
        self.assertEqual(
            depois["lojas_concluidos"], antes["lojas_concluidos"]
        )

    def test_chamado_pendente_conta_em_pendentes(self):
        ativo = self._criar_ativo("M-CH-PE", tipo_imovel="LOJA")
        antes = self._metricas()
        self._criar_chamado(ativo, StatusChamado.PENDENTE, "OS-PE-1")
        depois = self._metricas()
        self.assertEqual(
            depois["lojas_pendentes"], antes["lojas_pendentes"] + 1
        )

    def test_chamado_concluido_conta_em_concluidos(self):
        ativo = self._criar_ativo("M-CH-CO", tipo_imovel="LOJA")
        antes = self._metricas()
        self._criar_chamado(ativo, StatusChamado.CONCLUIDO, "OS-CO-1")
        depois = self._metricas()
        self.assertEqual(
            depois["lojas_concluidos"], antes["lojas_concluidos"] + 1
        )
        self.assertEqual(
            depois["lojas_pendentes"], antes["lojas_pendentes"]
        )

    def test_chamado_cancelado_nao_entra_em_pendentes_nem_concluidos(self):
        ativo = self._criar_ativo("M-CH-CA", tipo_imovel="LOJA")
        antes = self._metricas()
        self._criar_chamado(ativo, StatusChamado.CANCELADO, "OS-CA-1")
        depois = self._metricas()
        self.assertEqual(
            depois["lojas_pendentes"], antes["lojas_pendentes"]
        )
        self.assertEqual(
            depois["lojas_concluidos"], antes["lojas_concluidos"]
        )

    def test_chamado_nao_emergencial_nao_entra_em_pendentes_nem_concluidos(self):
        ativo = self._criar_ativo("M-CH-NE", tipo_imovel="LOJA")
        antes = self._metricas()
        self._criar_chamado(ativo, StatusChamado.NAO_EMERGENCIAL, "OS-NE-1")
        depois = self._metricas()
        self.assertEqual(
            depois["lojas_pendentes"], antes["lojas_pendentes"]
        )
        self.assertEqual(
            depois["lojas_concluidos"], antes["lojas_concluidos"]
        )

    def test_chamados_em_predio_caem_em_predios(self):
        """A separação prédio x loja na métrica de chamados é pelo
        ``ativo.tipo_imovel`` do chamado."""
        predio = self._criar_ativo("M-CH-PR", tipo_imovel="PREDIO")
        antes = self._metricas()
        self._criar_chamado(predio, StatusChamado.ABERTO, "OS-PRAB-1")
        self._criar_chamado(predio, StatusChamado.CONCLUIDO, "OS-PRCO-1")
        depois = self._metricas()
        self.assertEqual(
            depois["predios_pendentes"], antes["predios_pendentes"] + 1
        )
        self.assertEqual(
            depois["predios_concluidos"], antes["predios_concluidos"] + 1
        )
        self.assertEqual(
            depois["lojas_pendentes"], antes["lojas_pendentes"]
        )
        self.assertEqual(
            depois["lojas_concluidos"], antes["lojas_concluidos"]
        )

    # ---------------- Fornecedores e Obras: não exibidos no hero -------------

    def test_metricas_home_nao_expoe_contagem_de_fornecedores(self):
        """Hoje a Home não exibe contagem de Fornecedores. Se essa decisão
        mudar, esta asserção deve ser atualizada explicitamente."""
        Fornecedor.objects.create(nome="Forn Teste Métrica", ativo=True)
        Fornecedor.objects.create(nome="Forn Teste Inativo", ativo=False)
        metricas = self._metricas()
        self.assertNotIn("fornecedores_total", metricas)
        self.assertNotIn("fornecedores_ativos", metricas)

    def test_metricas_home_nao_expoe_contagem_de_obras(self):
        """Hoje a Home não exibe contagem de Obras. Se essa decisão mudar,
        esta asserção deve ser atualizada explicitamente."""
        ativo = self._criar_ativo("M-OB-1", tipo_imovel="LOJA")
        Obra.objects.create(
            ativo=ativo,
            descricao="Pintura",
            data_inicio=timezone.now().date(),
            data_fim_planejada=timezone.now().date(),
            ativa=True,
        )
        metricas = self._metricas()
        self.assertNotIn("obras_total", metricas)
        self.assertNotIn("obras_ativas", metricas)

    # ---------------- Contrato HTTP da Home ----------------------------------

    def test_home_anonimo_recebe_302(self):
        self.client.logout()
        response = self.client.get(reverse("chamados:home"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response["Location"])

    def test_home_autenticado_recebe_200(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertEqual(response.status_code, 200)

    def test_home_contexto_contem_chaves_de_metricas_esperadas(self):
        response = self.client.get(reverse("chamados:home"))
        self.assertEqual(response.status_code, 200)
        self.assertIn("metricas", response.context)
        metricas = response.context["metricas"]
        chaves_obrigatorias = {
            "predios_total",
            "lojas_total",
            "predios_pendentes",
            "lojas_pendentes",
            "predios_concluidos",
            "lojas_concluidos",
        }
        self.assertEqual(
            chaves_obrigatorias & metricas.keys(),
            chaves_obrigatorias,
            msg=f"Faltam chaves nas métricas da Home: "
                f"{chaves_obrigatorias - metricas.keys()}",
        )
