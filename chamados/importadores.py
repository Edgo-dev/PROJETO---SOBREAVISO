"""Importadores de dados operacionais do app chamados.

Cada importador segue o mesmo padrão:

1. Lê todas as linhas do .xlsx.
2. Tenta detectar a linha de cabeçalho automaticamente nas primeiras N linhas
   (planilhas com título antes do header não quebram).
3. Mapeia cada coluna reconhecida para um campo do sistema usando um
   dicionário generoso de sinônimos. Colunas não reconhecidas são ignoradas
   silenciosamente — qualquer planilha pode ser usada.
4. Devolve um `Resultado*` com contadores, erros linha-a-linha e o mapa
   "header da planilha → campo do sistema" para auditoria pelo usuário.
"""

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from unicodedata import normalize

from django.db import transaction
from openpyxl import load_workbook

from .models import Ativo, Fornecedor, Obra


# ============================================================================
# Utilitários compartilhados
# ============================================================================

VALORES_ATIVO = {"sim", "s", "ativo", "true", "1", "v", "verdadeiro"}
VALORES_INATIVO = {"nao", "n", "inativo", "false", "0", "f", "falso"}

# Quantas linhas no topo o detector tenta antes de desistir.
MAX_LINHAS_PROCURA_CABECALHO = 6


def _normalizar_cabecalho(valor) -> str:
    """Lower-case, sem acentos, com espaços e barras colapsados."""
    texto = "" if valor is None else str(valor)
    texto = texto.strip().lower()
    # Normaliza separadores comuns para espaço simples
    for sep in ("/", "\\", "-", "_", ".", "(", ")", "[", "]", ",", ";", ":"):
        texto = texto.replace(sep, " ")
    texto = " ".join(texto.split())
    texto = normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return texto


def _normalizar_texto(valor) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def _interpretar_ativo(valor) -> bool:
    """Converte 'sim/não/ativo/inativo/1/0/True/False' em boolean."""
    texto = _normalizar_cabecalho(valor)
    if not texto:
        return True
    if texto in VALORES_INATIVO:
        return False
    if texto in VALORES_ATIVO:
        return True
    return True


def _mapear_colunas(cabecalhos, mapeamento: dict) -> dict:
    """Mapeia campo_do_sistema -> índice_da_coluna na planilha.

    Para cada coluna do cabeçalho, tenta achar UM campo cujo set de aliases
    contenha o cabeçalho normalizado. Itera os campos em ordem de declaração
    (insertion order), então campos mais específicos devem vir primeiro.
    """
    colunas = {}
    for indice, cabecalho in enumerate(cabecalhos or []):
        cabecalho_normalizado = _normalizar_cabecalho(cabecalho)
        if not cabecalho_normalizado:
            continue
        for campo, aliases in mapeamento.items():
            if campo in colunas:
                continue
            if cabecalho_normalizado in aliases:
                colunas[campo] = indice
                break
    return colunas


def _detectar_cabecalho(
    linhas_iniciais: list, mapeamento: dict, campos_obrigatorios: tuple
) -> tuple:
    """Procura nas primeiras N linhas qual é o cabeçalho real.

    Útil quando a planilha tem uma linha de título antes do header. Escolhe a
    linha que cobre mais campos obrigatórios.

    Retorna ``(indice, mapa, cabecalhos_originais)`` — índice -1 se nenhuma
    linha mapeou nada.
    """
    melhor_indice = -1
    melhor_mapa = {}
    melhor_score = 0
    melhor_cabecalhos = ()

    for i, linha in enumerate(linhas_iniciais):
        mapa = _mapear_colunas(linha, mapeamento)
        # Score = quantos campos OBRIGATÓRIOS foram mapeados nessa linha.
        score = sum(1 for c in campos_obrigatorios if c in mapa)
        if score > melhor_score or (score == melhor_score and len(mapa) > len(melhor_mapa)):
            melhor_indice = i
            melhor_mapa = mapa
            melhor_score = score
            melhor_cabecalhos = linha

    return melhor_indice, melhor_mapa, melhor_cabecalhos


def _valor_linha(linha, colunas, campo):
    indice = colunas.get(campo)
    if indice is None or indice >= len(linha):
        return ""
    return _normalizar_texto(linha[indice])


_SEPARADORES_ENDERECO = re.compile(r"\s+[-–—/|]\s+")
_NAO_PARECE_CIDADE = re.compile(r"^[\d\s\-\.,]+$")


def _extrair_cidade_do_endereco(endereco: str) -> str:
    """Best-effort: tenta inferir a cidade do final do endereço.

    Estratégia: divide por separadores comuns (` - `, ` – `, ` — `, ` / `,
    ` | `) e devolve o último segmento. Se o último segmento for vazio,
    apenas números/pontuação, ou tiver mais de 60 caracteres (claramente
    não é nome de cidade), retorna string vazia.

    Exemplos:
    - "Rua X, 100 - Bairro Y - São Paulo" -> "São Paulo"
    - "ALAMEDA RIO NEGRO, 111 - LOJA 114 - PISO TER BARUERI" -> "PISO TER BARUERI"
    - "Rua única sem separador" -> ""
    - "Rua X - 123" -> ""
    """
    if not endereco:
        return ""
    partes = [p.strip() for p in _SEPARADORES_ENDERECO.split(endereco) if p.strip()]
    if len(partes) < 2:
        return ""
    ultimo = partes[-1]
    if not ultimo or len(ultimo) > 60:
        return ""
    if _NAO_PARECE_CIDADE.fullmatch(ultimo):
        return ""
    return ultimo


def _construir_mapa_amigavel(
    colunas: dict, cabecalhos: tuple, nomes_amigaveis: dict
) -> dict:
    """Constrói o mapa 'campo amigável -> header original da planilha' para
    exibir ao usuário no resultado da importação."""
    mapa = {}
    for campo, indice in colunas.items():
        if indice < len(cabecalhos):
            header_original = cabecalhos[indice]
            label_amigavel = nomes_amigaveis.get(campo, campo)
            mapa[label_amigavel] = str(header_original) if header_original is not None else ""
    return mapa


def _ler_planilha(arquivo):
    """Lê a planilha .xlsx inteira em memória (lista de tuplas)."""
    workbook = load_workbook(arquivo, read_only=True, data_only=True)
    planilha = workbook.active
    linhas = list(planilha.iter_rows(values_only=True))
    workbook.close()
    return linhas


# ============================================================================
# ATIVOS / Parque Imobiliário
# ============================================================================

MAPEAMENTO_CABECALHOS_ATIVO = {
    # IMPORTANTE: ordem importa quando aliases podem se sobrepor. Campos mais
    # específicos devem vir primeiro.
    "ativo_prisma": {
        "ativo prisma", "ativo_prisma", "prisma",
        "cod prisma", "codigo prisma", "código prisma",
        "cod ativo", "codigo do ativo", "codigo ativo",
        "id ativo", "ativo id", "n prisma", "numero prisma",
        "ativo num", "num ativo",
    },
    "nome_site": {
        "nome do site", "nome_site", "nome site", "nome operacional",
        "nome", "site", "nome do imovel",
        "denominacao do imovel", "denominacao", "denominacao imovel",
        "nomenclatura prisma", "nomenclatura",
        "nome comercial", "razao social do imovel",
        "descricao do site", "descricao site",
    },
    "endereco": {
        "endereco", "endereço", "logradouro", "rua",
        "endereco completo", "endereco do imovel", "endereco do site",
        "endereco do ativo", "logradouro completo",
    },
    "cidade": {
        "cidade", "municipio", "município", "cidade imovel", "cidade do imovel",
        "city",
    },
    "uf": {
        "uf", "estado", "sigla uf", "sigla estado", "sigla", "uf imovel",
        "uf do imovel", "uf estado",
    },
    "regional": {
        "regional", "regiao", "região",
        "regional do ativo", "regional operacional",
        "filial regional", "filial",
    },
    "tipo_imovel": {
        "tipo de imovel", "tipo imovel", "tipo_imovel",
        "imovel", "imóvel", "tipo predio",
        "categoria imovel", "tipologia imovel", "classificacao imovel",
    },
    "lider_coordenacao": {
        "lider da coordenacao", "lider_coordenacao",
        "lider regional", "lider", "líder",
        "coordenador", "coordenacao", "coordenação",
        "responsavel regional", "responsavel",
    },
    "tipo_site_sla": {
        "tipo site sla", "tipo_site_sla", "tipo site",
        "sla", "tipo sla", "tipo de sla",
        "tipo atendimento", "atendimento",
        "categoria site", "categoria do site",
    },
    "ativo": {
        "ativo?", "ativo/inativo", "status",
        "situacao", "situação", "ativacao", "ativo sn",
    },
}

# Apenas o identificador é REALMENTE obrigatório para criar/atualizar o
# registro. Demais campos viram opcionais — se vierem vazios na planilha, o
# ativo entra no banco com o campo em branco e o usuário pode filtrar/corrigir
# na Matriz de Ativos depois. A tela de resultado mostra um breakdown
# de quantas linhas tiveram cada campo vazio para auditoria.
CAMPOS_OBRIGATORIOS_ATIVO = (
    "ativo_prisma",
)

# Campos que aparecem no breakdown de qualidade dos dados ao final.
CAMPOS_RECOMENDADOS_ATIVO = (
    "nome_site",
    "endereco",
    "cidade",
    "uf",
    "regional",
)

NOMES_AMIGAVEIS_ATIVO = {
    "ativo_prisma": "Ativo Prisma",
    "nome_site": "Nome do Site",
    "endereco": "Endereço",
    "cidade": "Cidade",
    "uf": "UF",
    "regional": "Regional",
    "tipo_imovel": "Tipo de Imóvel",
    "lider_coordenacao": "Líder da Coordenação",
    "tipo_site_sla": "Tipo Site / SLA",
    "ativo": "Status (Ativo/Inativo)",
}

# Campos opcionais (texto) do Ativo que entram na atualização parcial.
# A chave (ativo_prisma), o boolean ('ativo') e o caso especial 'cidade'
# (que aceita valor inferido do endereço) têm tratamento próprio.
_CAMPOS_OPCIONAIS_ATIVO = (
    "nome_site",
    "endereco",
    "uf",
    "regional",
    "tipo_imovel",
    "lider_coordenacao",
    "tipo_site_sla",
)


def _campos_atualizar_ativo(colunas: dict, dados: dict, cidade_valor: str) -> dict:
    """Constrói o dict de campos a sobrescrever em um Ativo EXISTENTE.

    Regra de robustez (importação parcial):
    - Coluna AUSENTE no cabeçalho -> preserva o valor existente no banco.
    - Coluna PRESENTE com valor vazio -> preserva o valor existente.
    - Coluna PRESENTE com valor preenchido -> atualiza.

    Casos especiais:
    - 'cidade' aceita valor explícito ou inferido do endereço; em ambos
      sobrescreve, pois é considerada presente quando há valor não vazio.
    - 'ativo' (bool) só é tocado quando a coluna existe explicitamente;
      caso contrário, o status atual do registro é preservado.
    """
    campos = {}
    for campo in _CAMPOS_OPCIONAIS_ATIVO:
        if campo in colunas:
            valor = dados.get(campo, "")
            if valor:
                campos[campo] = valor
    if cidade_valor:
        campos["cidade"] = cidade_valor
    if "ativo" in colunas:
        campos["ativo"] = dados["ativo"]
    return campos


@dataclass
class ResultadoImportacaoAtivos:
    total_linhas: int = 0
    criados: int = 0
    atualizados: int = 0
    ignorados: int = 0
    erros: list = field(default_factory=list)
    # Mapa 'header da planilha -> campo do sistema' para exibir ao usuário.
    mapeamento: dict = field(default_factory=dict)
    # Headers lidos na linha de cabeçalho detectada (para mostrar ao usuário).
    cabecalhos_lidos: list = field(default_factory=list)
    # Cabeçalhos que existiam mas não foram mapeados (ignorados).
    cabecalhos_ignorados: list = field(default_factory=list)
    # Linhas que tinham 'Cidade' vazia mas onde o sistema conseguiu inferir
    # a cidade a partir do final do endereço (best-effort).
    cidades_inferidas_do_endereco: int = 0
    # Breakdown: quantas linhas SALVAS ficaram com cada campo recomendado
    # vazio. Permite ao usuário ver a qualidade dos dados e priorizar correção.
    campos_vazios: dict = field(default_factory=dict)


def importar_ativos_excel(arquivo, usuario=None):
    """Importa Ativos a partir de uma planilha .xlsx flexível.

    Aceita planilhas com qualquer número de colunas em qualquer ordem.
    Reconhece os cabeçalhos relevantes e ignora o resto.
    """
    del usuario
    resultado = ResultadoImportacaoAtivos()

    todas_linhas = _ler_planilha(arquivo)
    if not todas_linhas:
        return resultado

    # Detecta a linha do cabeçalho dentro das primeiras N.
    linhas_iniciais = todas_linhas[:MAX_LINHAS_PROCURA_CABECALHO]
    indice_header, colunas, cabecalhos = _detectar_cabecalho(
        linhas_iniciais, MAPEAMENTO_CABECALHOS_ATIVO, CAMPOS_OBRIGATORIOS_ATIVO
    )

    if not colunas:
        # Nenhum campo reconhecido em nenhuma das primeiras N linhas.
        primeira_linha_str = ", ".join(
            str(c) for c in (todas_linhas[0] or ()) if c
        )
        resultado.erros.append(
            "Não foi possível identificar nenhuma coluna obrigatória no cabeçalho. "
            f"Linha 1 da planilha: {primeira_linha_str or '(vazia)'}. "
            "São esperadas pelo menos as colunas: "
            "Ativo Prisma, Nome do Site, Endereço, Cidade, UF, Regional."
        )
        return resultado

    # Registra o mapeamento usado para o usuário poder auditar.
    resultado.mapeamento = _construir_mapa_amigavel(
        colunas, cabecalhos, NOMES_AMIGAVEIS_ATIVO
    )
    resultado.cabecalhos_lidos = [
        str(c) if c is not None else "" for c in cabecalhos
    ]
    indices_mapeados = set(colunas.values())
    resultado.cabecalhos_ignorados = [
        str(c) for i, c in enumerate(cabecalhos)
        if c and i not in indices_mapeados
    ]

    # Valida que o identificador único (Ativo Prisma) foi mapeado.
    if "ativo_prisma" not in colunas:
        cabecalhos_str = ", ".join(str(c) for c in cabecalhos if c)
        resultado.erros.append(
            f"Cabeçalhos detectados na planilha: {cabecalhos_str}. "
            "Coluna 'Ativo Prisma' não encontrada — sem ela é impossível "
            "identificar cada registro. Renomeie a coluna chave da sua "
            "planilha para 'Ativo Prisma' (ou sinônimos: prisma, cod ativo, etc.)."
        )
        return resultado

    # Processa as linhas de dados (tudo depois do header).
    # Envolve em transaction.atomic para evitar auto-commit por linha — em
    # planilhas com milhares de linhas a diferença é de minutos para segundos.
    primeira_linha_dados = indice_header + 1
    with transaction.atomic():
        for numero_linha, linha in enumerate(
            todas_linhas[primeira_linha_dados:], start=primeira_linha_dados + 1
        ):
            if not linha or not any(_normalizar_texto(valor) for valor in linha):
                continue

            resultado.total_linhas += 1
            endereco_valor = _valor_linha(linha, colunas, "endereco")
            cidade_valor = _valor_linha(linha, colunas, "cidade")
            # Fallback: cidade vazia + endereço presente -> tenta inferir.
            if not cidade_valor and endereco_valor:
                cidade_inferida = _extrair_cidade_do_endereco(endereco_valor)
                if cidade_inferida:
                    cidade_valor = cidade_inferida
                    resultado.cidades_inferidas_do_endereco += 1

            dados = {
                "ativo_prisma": _valor_linha(linha, colunas, "ativo_prisma"),
                "nome_site": _valor_linha(linha, colunas, "nome_site"),
                "endereco": endereco_valor,
                "cidade": cidade_valor,
                "uf": _valor_linha(linha, colunas, "uf").upper(),
                "regional": _valor_linha(linha, colunas, "regional"),
                "tipo_imovel": _valor_linha(linha, colunas, "tipo_imovel"),
                "lider_coordenacao": _valor_linha(linha, colunas, "lider_coordenacao"),
                "tipo_site_sla": _valor_linha(linha, colunas, "tipo_site_sla"),
                "ativo": _interpretar_ativo(_valor_linha(linha, colunas, "ativo")),
            }

            # Apenas o ativo_prisma é estritamente obrigatório (chave única).
            if not dados["ativo_prisma"]:
                resultado.ignorados += 1
                resultado.erros.append(
                    f"Linha {numero_linha}: 'Ativo Prisma' em branco — "
                    f"é o identificador único, não é possível salvar sem ele."
                )
                continue

            # Conta quantas linhas têm cada campo recomendado vazio.
            for campo in CAMPOS_RECOMENDADOS_ATIVO:
                if not dados.get(campo):
                    resultado.campos_vazios[campo] = resultado.campos_vazios.get(campo, 0) + 1

            # Atualização parcial preserva dados: colunas ausentes/vazias NÃO
            # sobrescrevem o que já existe. Já a criação mantém o modo
            # permissivo (campos faltantes entram como string vazia).
            registro_existente = Ativo.objects.filter(
                ativo_prisma=dados["ativo_prisma"]
            ).first()
            if registro_existente is None:
                Ativo.objects.create(
                    ativo_prisma=dados["ativo_prisma"],
                    nome_site=dados["nome_site"],
                    endereco=dados["endereco"],
                    cidade=dados["cidade"],
                    uf=dados["uf"],
                    regional=dados["regional"],
                    tipo_imovel=dados["tipo_imovel"],
                    lider_coordenacao=dados["lider_coordenacao"],
                    tipo_site_sla=dados["tipo_site_sla"],
                    ativo=dados["ativo"],
                )
                resultado.criados += 1
            else:
                campos_para_atualizar = _campos_atualizar_ativo(
                    colunas, dados, cidade_valor
                )
                if campos_para_atualizar:
                    for campo, valor in campos_para_atualizar.items():
                        setattr(registro_existente, campo, valor)
                    registro_existente.save(
                        update_fields=list(campos_para_atualizar) + ["atualizado_em"]
                    )
                resultado.atualizados += 1

    return resultado


# ============================================================================
# FORNECEDORES / Supervisores
# ============================================================================

MAPEAMENTO_CABECALHOS_FORNECEDOR = {
    "nome": {
        "nome",
        # Sinônimos atuais
        "nome do supervisor", "nome supervisor", "supervisor",
        # Sinônimos legados / planilhas antigas (mantém compatibilidade)
        "nome do coordenador", "nome coordenador", "coordenador",
        # Genéricos
        "responsavel", "responsável",
        "contato principal", "nome contato",
    },
    "telefone": {
        "telefone", "celular", "cel", "tel",
        "contato telefonico", "contato telefônico",
        "fone", "whatsapp", "numero", "telefone celular",
    },
    "email": {
        "email", "e-mail", "e mail", "endereco eletronico",
        "endereço eletrônico", "correio eletronico",
    },
    "empresa": {
        "empresa", "razao social", "razão social",
        "fornecedor", "nome fantasia", "companhia",
    },
    "estados_atendidos": {
        "estados atendidos", "estados_atendidos", "estados",
        "ufs atendidas", "ufs", "uf atendida",
        "regioes atendidas", "área de atuação", "area de atuacao",
        "abrangencia", "abrangência",
    },
}

CAMPOS_OBRIGATORIOS_FORNECEDOR = ("nome",)

NOMES_AMIGAVEIS_FORNECEDOR = {
    "nome": "Nome",
    "telefone": "Telefone",
    "email": "E-mail",
    "empresa": "Empresa",
    "estados_atendidos": "Estados atendidos",
}

COLUNAS_MODELO_FORNECEDOR = (
    "Nome",
    "Telefone",
    "E-mail",
    "Empresa",
    "Estados atendidos",
)


@dataclass
class ResultadoImportacaoFornecedores:
    total_linhas: int = 0
    criados: int = 0
    atualizados: int = 0
    ignorados: int = 0
    erros: list = field(default_factory=list)
    mapeamento: dict = field(default_factory=dict)
    cabecalhos_lidos: list = field(default_factory=list)
    cabecalhos_ignorados: list = field(default_factory=list)


def _normalizar_estados(valor: str) -> str:
    if not valor:
        return ""
    partes = [parte.strip().upper() for parte in valor.replace(";", ",").split(",")]
    return ", ".join(parte for parte in partes if parte)


# Campos opcionais do Fornecedor que entram na atualização parcial.
_CAMPOS_OPCIONAIS_FORNECEDOR = (
    "telefone",
    "email",
    "empresa",
    "estados_atendidos",
)


def _campos_atualizar_fornecedor(colunas: dict, dados: dict) -> dict:
    """Constrói o dict de campos a sobrescrever em um Fornecedor EXISTENTE.

    Mesma regra dos ativos: coluna ausente OU valor vazio preserva o valor
    atual. O campo 'ativo' (bool) NÃO consta na importação de fornecedores
    e é deliberadamente preservado no update parcial.
    """
    campos = {}
    for campo in _CAMPOS_OPCIONAIS_FORNECEDOR:
        if campo in colunas:
            valor = dados.get(campo, "")
            if valor:
                campos[campo] = valor
    return campos


def importar_fornecedores_excel(arquivo, usuario=None):
    """Importa Fornecedores/Supervisores a partir de uma planilha .xlsx flexível."""
    del usuario
    resultado = ResultadoImportacaoFornecedores()

    todas_linhas = _ler_planilha(arquivo)
    if not todas_linhas:
        return resultado

    linhas_iniciais = todas_linhas[:MAX_LINHAS_PROCURA_CABECALHO]
    indice_header, colunas, cabecalhos = _detectar_cabecalho(
        linhas_iniciais,
        MAPEAMENTO_CABECALHOS_FORNECEDOR,
        CAMPOS_OBRIGATORIOS_FORNECEDOR,
    )

    if not colunas:
        primeira_linha_str = ", ".join(
            str(c) for c in (todas_linhas[0] or ()) if c
        )
        resultado.erros.append(
            "Não foi possível identificar a coluna 'Nome' no cabeçalho. "
            f"Linha 1 da planilha: {primeira_linha_str or '(vazia)'}. "
            "É esperada pelo menos a coluna: Nome."
        )
        return resultado

    resultado.mapeamento = _construir_mapa_amigavel(
        colunas, cabecalhos, NOMES_AMIGAVEIS_FORNECEDOR
    )
    resultado.cabecalhos_lidos = [
        str(c) if c is not None else "" for c in cabecalhos
    ]
    indices_mapeados = set(colunas.values())
    resultado.cabecalhos_ignorados = [
        str(c) for i, c in enumerate(cabecalhos)
        if c and i not in indices_mapeados
    ]

    if "nome" not in colunas:
        cabecalhos_str = ", ".join(str(c) for c in cabecalhos if c)
        resultado.erros.append(
            f"Cabeçalhos detectados: {cabecalhos_str}. "
            "Coluna obrigatória 'Nome' não encontrada. "
            "Renomeie sua planilha para que a primeira coluna seja 'Nome' do supervisor."
        )
        return resultado

    primeira_linha_dados = indice_header + 1
    with transaction.atomic():
        for numero_linha, linha in enumerate(
            todas_linhas[primeira_linha_dados:], start=primeira_linha_dados + 1
        ):
            if not linha or not any(_normalizar_texto(valor) for valor in linha):
                continue

            resultado.total_linhas += 1
            dados = {
                "nome": _valor_linha(linha, colunas, "nome"),
                "telefone": _valor_linha(linha, colunas, "telefone"),
                "email": _valor_linha(linha, colunas, "email").lower(),
                "empresa": _valor_linha(linha, colunas, "empresa"),
                "estados_atendidos": _normalizar_estados(
                    _valor_linha(linha, colunas, "estados_atendidos")
                ),
            }

            if not dados["nome"]:
                resultado.ignorados += 1
                resultado.erros.append(
                    f"Linha {numero_linha}: coluna 'Nome' em branco."
                )
                continue

            # Atualização parcial preserva dados existentes; criação mantém o
            # modo permissivo (campos faltantes entram como string vazia).
            registro_existente = Fornecedor.objects.filter(
                nome=dados["nome"]
            ).first()
            if registro_existente is None:
                Fornecedor.objects.create(
                    nome=dados["nome"],
                    telefone=dados["telefone"],
                    email=dados["email"],
                    empresa=dados["empresa"],
                    estados_atendidos=dados["estados_atendidos"],
                    ativo=True,
                )
                resultado.criados += 1
            else:
                campos_para_atualizar = _campos_atualizar_fornecedor(colunas, dados)
                if campos_para_atualizar:
                    for campo, valor in campos_para_atualizar.items():
                        setattr(registro_existente, campo, valor)
                    registro_existente.save(
                        update_fields=list(campos_para_atualizar) + ["atualizado_em"]
                    )
                resultado.atualizados += 1

    return resultado


# ============================================================================
# OBRAS
# ============================================================================

MAPEAMENTO_CABECALHOS_OBRA = {
    "ativo_prisma": {
        "ativo prisma", "ativo_prisma", "prisma",
        "cod prisma", "codigo prisma", "cod ativo",
        "id ativo", "ativo id",
    },
    "descricao": {
        "descricao da obra", "descricao_obra", "descricao",
        "obra", "do que se trata", "objeto da obra",
        "escopo", "escopo da obra", "atividade",
    },
    "data_inicio": {
        "data inicio", "data_inicio", "data de inicio",
        "inicio", "data inicial", "início",
        "data de inicio da obra", "start date",
    },
    "data_fim_planejada": {
        "data fim planejada", "data fim", "data_fim_planejada", "data_fim",
        "data planejada para termino", "previsao de termino",
        "termino", "data termino", "data prevista",
        "data de termino", "previsao termino", "end date",
        "previsao de entrega", "data de entrega prevista",
    },
    "responsavel": {
        "responsavel", "responsavel pela obra",
        "empresa responsavel", "empresa", "fornecedor",
        "construtora", "executor", "executor da obra",
    },
    "observacoes": {
        "observacoes", "observações", "obs", "observacao",
        "comentarios", "comentários", "notas",
    },
}

CAMPOS_OBRIGATORIOS_OBRA = (
    "ativo_prisma",
    "descricao",
    "data_inicio",
    "data_fim_planejada",
)

NOMES_AMIGAVEIS_OBRA = {
    "ativo_prisma": "Ativo Prisma",
    "descricao": "Descrição da Obra",
    "data_inicio": "Data Início",
    "data_fim_planejada": "Data Fim Planejada",
    "responsavel": "Responsável",
    "observacoes": "Observações",
}

COLUNAS_MODELO_OBRA = (
    "Ativo Prisma",
    "Nome do site",
    "Endereço",
    "Cidade",
    "UF",
    "Regional",
    "Tipo de imóvel",
    "Descrição da obra",
    "Data início",
    "Data fim planejada",
    "Responsável",
    "Observações",
)


@dataclass
class ResultadoImportacaoObras:
    total_linhas: int = 0
    criados: int = 0
    atualizados: int = 0
    ignorados: int = 0
    erros: list = field(default_factory=list)
    mapeamento: dict = field(default_factory=dict)
    cabecalhos_lidos: list = field(default_factory=list)
    cabecalhos_ignorados: list = field(default_factory=list)


def _parse_data(valor):
    """Aceita datetime, date ou string em dd/mm/aaaa, aaaa-mm-dd ou variantes."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        texto = valor.strip()
        if not texto:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d.%m.%Y"):
            try:
                return datetime.strptime(texto, fmt).date()
            except ValueError:
                continue
    return None


def importar_obras_excel(arquivo, usuario=None):
    """Importa Obras a partir de uma planilha .xlsx flexível."""
    del usuario
    resultado = ResultadoImportacaoObras()

    todas_linhas = _ler_planilha(arquivo)
    if not todas_linhas:
        return resultado

    linhas_iniciais = todas_linhas[:MAX_LINHAS_PROCURA_CABECALHO]
    indice_header, colunas, cabecalhos = _detectar_cabecalho(
        linhas_iniciais, MAPEAMENTO_CABECALHOS_OBRA, CAMPOS_OBRIGATORIOS_OBRA
    )

    if not colunas:
        primeira_linha_str = ", ".join(
            str(c) for c in (todas_linhas[0] or ()) if c
        )
        resultado.erros.append(
            "Não foi possível identificar as colunas obrigatórias da obra. "
            f"Linha 1 da planilha: {primeira_linha_str or '(vazia)'}. "
            "São esperadas as colunas: Ativo Prisma, Descrição da Obra, "
            "Data Início, Data Fim Planejada."
        )
        return resultado

    resultado.mapeamento = _construir_mapa_amigavel(
        colunas, cabecalhos, NOMES_AMIGAVEIS_OBRA
    )
    resultado.cabecalhos_lidos = [
        str(c) if c is not None else "" for c in cabecalhos
    ]
    indices_mapeados = set(colunas.values())
    resultado.cabecalhos_ignorados = [
        str(c) for i, c in enumerate(cabecalhos)
        if c and i not in indices_mapeados
    ]

    faltantes_mapeamento = [
        c for c in CAMPOS_OBRIGATORIOS_OBRA if c not in colunas
    ]
    if faltantes_mapeamento:
        faltantes_str = ", ".join(
            NOMES_AMIGAVEIS_OBRA[c] for c in faltantes_mapeamento
        )
        cabecalhos_str = ", ".join(str(c) for c in cabecalhos if c)
        resultado.erros.append(
            f"Cabeçalhos detectados: {cabecalhos_str}. "
            f"Coluna(s) obrigatória(s) NÃO encontrada(s): {faltantes_str}."
        )
        return resultado

    primeira_linha_dados = indice_header + 1
    with transaction.atomic():
        for numero_linha, linha in enumerate(
            todas_linhas[primeira_linha_dados:], start=primeira_linha_dados + 1
        ):
            if not linha or not any(_normalizar_texto(valor) for valor in linha):
                continue

            resultado.total_linhas += 1
            ativo_prisma = _valor_linha(linha, colunas, "ativo_prisma")
            descricao = _valor_linha(linha, colunas, "descricao")
            data_inicio_raw = (
                linha[colunas["data_inicio"]]
                if "data_inicio" in colunas and colunas["data_inicio"] < len(linha)
                else None
            )
            data_fim_raw = (
                linha[colunas["data_fim_planejada"]]
                if "data_fim_planejada" in colunas
                and colunas["data_fim_planejada"] < len(linha)
                else None
            )
            data_inicio = _parse_data(data_inicio_raw)
            data_fim_planejada = _parse_data(data_fim_raw)
            responsavel = _valor_linha(linha, colunas, "responsavel")
            observacoes = _valor_linha(linha, colunas, "observacoes")

            faltantes = []
            if not ativo_prisma:
                faltantes.append("Ativo Prisma")
            if not descricao:
                faltantes.append("Descrição da Obra")
            if data_inicio is None:
                faltantes.append("Data Início")
            if data_fim_planejada is None:
                faltantes.append("Data Fim Planejada")

            if faltantes:
                resultado.ignorados += 1
                resultado.erros.append(
                    f"Linha {numero_linha}: campos obrigatórios em branco/inválidos: "
                    f"{', '.join(faltantes)}."
                )
                continue

            if data_fim_planejada < data_inicio:
                resultado.ignorados += 1
                resultado.erros.append(
                    f"Linha {numero_linha}: data fim planejada anterior à data de início."
                )
                continue

            try:
                ativo = Ativo.objects.get(ativo_prisma=ativo_prisma)
            except Ativo.DoesNotExist:
                resultado.ignorados += 1
                resultado.erros.append(
                    f"Linha {numero_linha}: Ativo Prisma '{ativo_prisma}' não cadastrado. "
                    "Importe o ativo pela Matriz de Ativos antes."
                )
                continue

            _obra, criado = Obra.objects.update_or_create(
                ativo=ativo,
                data_inicio=data_inicio,
                descricao=descricao,
                defaults={
                    "data_fim_planejada": data_fim_planejada,
                    "responsavel": responsavel,
                    "observacoes": observacoes,
                    "ativa": True,
                },
            )
            if criado:
                resultado.criados += 1
            else:
                resultado.atualizados += 1

    return resultado


# ============================================================================
# Aliases retro-compatíveis (testes antigos referenciam estes nomes)
# ============================================================================

# Mantém a API pública anterior funcionando.
MAPEAMENTO_CABECALHOS = MAPEAMENTO_CABECALHOS_ATIVO
CAMPOS_OBRIGATORIOS = CAMPOS_OBRIGATORIOS_ATIVO
