"""Views do app chamados."""

import logging
import mimetypes
from datetime import date
from io import BytesIO
from zipfile import BadZipFile

from django.contrib import messages
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.core.cache import cache
from django.core.paginator import Paginator
from django.db.models import Case, Count, IntegerField, Q, Value, When
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_http_methods, require_POST
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException

from . import services
from .forms import (
    AtivoForm,
    AtivosImportForm,
    ChamadoForm,
    FornecedorForm,
    FornecedoresImportForm,
    LoginForm,
    ObraForm,
    ObrasImportForm,
    RegisterForm,
    RegistrarReportForm,
)
from .importadores import (
    COLUNAS_MODELO_FORNECEDOR,
    COLUNAS_MODELO_OBRA,
    importar_ativos_excel,
    importar_fornecedores_excel,
    importar_obras_excel,
)
from .models import Ativo, Chamado, Fornecedor, Obra, StatusChamado

ATIVOS_POR_PAGINA = 20
CHAMADOS_POR_PAGINA = 20
FORNECEDORES_POR_PAGINA = 20
OBRAS_POR_PAGINA = 20
ATUALIZAR_REPORT_POR_PAGINA = 20

logger = logging.getLogger(__name__)

# Exceções típicas de planilha Excel inválida/corrompida quando o arquivo
# chega ao openpyxl. Tratadas pelas views de importação para devolver a
# própria tela com mensagem amigável (em vez de 500).
EXCECOES_PLANILHA_INVALIDA = (BadZipFile, InvalidFileException, ValueError)
MENSAGEM_PLANILHA_INVALIDA = (
    "Arquivo Excel inválido ou corrompido. Envie uma planilha .xlsx válida."
)


def _montar_querystring(request, excluir=("page",)):
    """Retorna querystring sem os parâmetros listados em excluir."""
    params = request.GET.copy()
    for chave in excluir:
        params.pop(chave, None)
    return params.urlencode()


def serve_evidencia(request, caminho):
    """Serve arquivos de evidência apenas para usuários autenticados."""
    if not request.user.is_authenticated:
        from django.contrib.auth.views import redirect_to_login
        return redirect_to_login(request.get_full_path())
    import os as _os
    from django.conf import settings as _settings
    base = _os.path.realpath(
        _os.path.join(_settings.MEDIA_ROOT, "evidencias")
    )
    real = _os.path.realpath(
        _os.path.join(_settings.MEDIA_ROOT, "evidencias", caminho)
    )
    if not real.startswith(base + _os.sep) and real != base:
        raise Http404
    if not _os.path.isfile(real):
        raise Http404
    content_type, _ = mimetypes.guess_type(real)
    content_type = content_type or "application/octet-stream"
    return FileResponse(open(real, "rb"), content_type=content_type)


def _obras_ativas_qs(ativo):
    hoje = date.today()
    return Obra.objects.filter(
        ativo=ativo,
        ativa=True,
        data_fim_real__isnull=True,
        data_inicio__lte=hoje,
        data_fim_planejada__gte=hoje,
    ).order_by("-data_inicio")


COLUNAS_EXPORTAR_ATUALIZAR_REPORT = [
    "Número da OS",
    "Status atual",
    "Data de abertura",
    "Ativo Prisma",
    "Nome do site",
    "Endereço",
    "Cidade",
    "UF",
    "Regional",
    "Tipo de imóvel",
    "Tipo site/SLA",
    "Líder coordenação",
    "Fornecedor",
    "Solicitante",
    "Contato solicitante",
    "Command Center",
    "Denominação",
    "Ação tomada",
    "Detalhamento da situação",
    "Primeiro report pendente",
    "Quantidade de atualizações",
    "Última atualização",
    "Data da última atualização",
    "Status da última atualização",
    "Criado em",
    "Atualizado em",
]


def _formatar_data_excel(valor):
    if not valor:
        return "-"
    return timezone.localtime(valor).strftime("%d/%m/%Y %H:%M")


def _normalizar_filtros_atualizar_report(params):
    return {
        "q": params.get("q", "").strip(),
        "data_inicio": params.get("data_inicio", "").strip(),
        "data_fim": params.get("data_fim", "").strip(),
        "status": params.get("status", "").strip(),
        "regional": params.get("regional", "").strip(),
        "fornecedor": params.get("fornecedor", "").strip(),
        "cidade": params.get("cidade", "").strip(),
        "primeiro_report": params.get("primeiro_report", "").strip(),
    }


def _filtrar_chamados_atualizar_report(params):
    filtros = _normalizar_filtros_atualizar_report(params)
    queryset = Chamado.objects.select_related("ativo", "fornecedor").prefetch_related(
        "atualizacoes"
    )

    q = filtros["q"]
    if q:
        queryset = queryset.filter(
            Q(numero_os__icontains=q)
            | Q(ativo__ativo_prisma__icontains=q)
            | Q(ativo__nome_site__icontains=q)
            | Q(ativo__cidade__icontains=q)
            | Q(ativo__regional__icontains=q)
        )

    data_inicio = parse_date(filtros["data_inicio"]) if filtros["data_inicio"] else None
    data_fim = parse_date(filtros["data_fim"]) if filtros["data_fim"] else None
    if data_inicio is not None:
        queryset = queryset.filter(data_abertura__date__gte=data_inicio)
    if data_fim is not None:
        queryset = queryset.filter(data_abertura__date__lte=data_fim)

    status_validos = {valor for valor, _ in StatusChamado.choices}
    if filtros["status"] in status_validos:
        queryset = queryset.filter(status=filtros["status"])

    if filtros["regional"]:
        queryset = queryset.filter(ativo__regional__icontains=filtros["regional"])
    if filtros["cidade"]:
        queryset = queryset.filter(ativo__cidade__icontains=filtros["cidade"])
    if filtros["fornecedor"]:
        try:
            queryset = queryset.filter(fornecedor_id=int(filtros["fornecedor"]))
        except (TypeError, ValueError):
            pass

    if filtros["primeiro_report"] == "sim":
        queryset = queryset.filter(atualizacoes__isnull=True)

    return queryset, filtros


def _q_lojas(prefixo: str = "") -> Q:
    """Filtro Q: ativos cujo tipo_imovel contém 'loja' (case-insensitive).

    Cobre 'Loja', 'LOJA', 'Loja Comercial', 'LOJA SHOPPING ...', etc.
    """
    campo = f"{prefixo}tipo_imovel__icontains" if prefixo else "tipo_imovel__icontains"
    return Q(**{campo: "loja"})


def _q_predios(prefixo: str = "") -> Q:
    """Filtro Q: tudo que NÃO for Loja é considerado Prédio.

    Cobre 'Prédio', 'PRÉDIO', 'Predio', 'PREDIO', 'TÉCNICO', 'SEDE', valor
    vazio, etc. — qualquer tipo_imovel que não contenha 'loja'.
    """
    return ~_q_lojas(prefixo)


def _metricas_home() -> dict:
    """Calcula os totais Prédios x Lojas exibidos nos cards do hero da Home.

    Regras formais (cobertas por ``HomeMetricasCoerentesTests``):

    - **Lojas**: ``Ativo.tipo_imovel`` contém ``"loja"`` (case-insensitive,
      acentos não importam pois o filtro é ``__icontains`` literal).
    - **Prédios**: NEGAÇÃO de Lojas — qualquer ``tipo_imovel`` que NÃO contenha
      ``"loja"``. Isso inclui ``tipo_imovel`` vazio (``""``), valores como
      ``"PRÉDIO"``, ``"TÉCNICO"``, ``"SEDE"`` etc.
    - **Status do Ativo**: a contagem NÃO filtra por ``ativo=True``;
      ``Ativo.ativo=False`` também é contabilizado. Trata-se de inventário
      total do parque, não de "ativos operacionais".
    - **Pendentes**: chamados com status ``ABERTO`` OU ``PENDENTE``
      (união). ``CANCELADO`` e ``NAO_EMERGENCIAL`` NÃO entram.
    - **Concluídos**: chamados com status estritamente ``CONCLUIDO``.
      ``CANCELADO`` e ``NAO_EMERGENCIAL`` NÃO entram.
    - **Fornecedores** e **Obras** NÃO compõem métrica no hero hoje.
    """
    metricas_ativos = Ativo.objects.aggregate(
        predios=Count("id", filter=_q_predios()),
        lojas=Count("id", filter=_q_lojas()),
    )

    pendentes_qs = Chamado.objects.filter(
        Q(status=StatusChamado.ABERTO) | Q(status=StatusChamado.PENDENTE)
    )
    metricas_chamados = pendentes_qs.aggregate(
        predios_pendentes=Count("id", filter=_q_predios("ativo__")),
        lojas_pendentes=Count("id", filter=_q_lojas("ativo__")),
    )

    concluidos_qs = Chamado.objects.filter(status=StatusChamado.CONCLUIDO)
    metricas_concluidos = concluidos_qs.aggregate(
        predios_concluidos=Count("id", filter=_q_predios("ativo__")),
        lojas_concluidos=Count("id", filter=_q_lojas("ativo__")),
    )

    return {
        "predios_total":      metricas_ativos["predios"],
        "lojas_total":        metricas_ativos["lojas"],
        "predios_pendentes":  metricas_chamados["predios_pendentes"],
        "lojas_pendentes":    metricas_chamados["lojas_pendentes"],
        "predios_concluidos": metricas_concluidos["predios_concluidos"],
        "lojas_concluidos":   metricas_concluidos["lojas_concluidos"],
    }


def home(request):
    """Pagina inicial do sistema com cards placeholder."""
    cards = [
        {
            "titulo": "Matriz de Ativos",
            "descricao": "Consultar ativos, filtrar endereços e abrir chamados a partir do parque imobiliário.",
            "url": reverse("chamados:ativos_list"),
        },
        {
            "titulo": "Importar Parque Imobiliário",
            "descricao": "Carregar planilha .xlsx para popular ou atualizar a Matriz de Ativos.",
            "url": reverse("chamados:ativos_import"),
        },
        {
            "titulo": "Novo Chamado",
            "descricao": "Registrar nova OS emergencial vinculada a um ativo.",
            "url": reverse("chamados:chamado_create"),
        },
        {
            "titulo": "Atualizar Report",
            "descricao": "Registrar reports, atualizar status e copiar o texto de acompanhamento.",
            "url": reverse("chamados:atualizar_report_list"),
        },
        {
            "titulo": "Contato Fornecedor",
            "descricao": "Consultar, cadastrar e atualizar contatos de fornecedores.",
            "url": reverse("chamados:fornecedores_list"),
        },
        {
            "titulo": "Consolidado Obras",
            "descricao": "Cadastrar, importar e acompanhar as obras do parque imobiliário.",
            "url": reverse("chamados:obras_list"),
        },
    ]
    return render(
        request,
        "chamados/home.html",
        {"cards": cards, "metricas": _metricas_home()},
    )


def ativos_list(request):
    """Matriz de Ativos: listagem com filtros, busca e paginacao."""
    q = request.GET.get("q", "").strip()
    cidade = request.GET.get("cidade", "").strip()
    uf = request.GET.get("uf", "").strip()
    regional = request.GET.get("regional", "").strip()
    ativo_param = request.GET.get("ativo", "todos").strip() or "todos"
    sem = request.GET.get("sem", "").strip()

    queryset = Ativo.objects.all()

    if q:
        queryset = queryset.filter(
            Q(ativo_prisma__icontains=q)
            | Q(nome_site__icontains=q)
            | Q(endereco__icontains=q)
            | Q(cidade__icontains=q)
            | Q(regional__icontains=q)
        )
    if cidade:
        queryset = queryset.filter(cidade__icontains=cidade)
    if uf:
        queryset = queryset.filter(uf__iexact=uf)
    if regional:
        queryset = queryset.filter(regional__icontains=regional)
    if ativo_param == "sim":
        queryset = queryset.filter(ativo=True)
    elif ativo_param == "nao":
        queryset = queryset.filter(ativo=False)

    # Filtro de qualidade: ativos com determinado campo vazio.
    CAMPOS_QUALIDADE = {
        "nome_site", "endereco", "cidade", "uf", "regional", "tipo_imovel"
    }
    if sem in CAMPOS_QUALIDADE:
        queryset = queryset.filter(**{sem: ""})

    total = queryset.count()

    # Contadores de qualidade (totais sem filtro de busca aplicado, mas
    # respeitando o status ativo/inativo para não confundir).
    base_qualidade = Ativo.objects.all()
    if ativo_param == "sim":
        base_qualidade = base_qualidade.filter(ativo=True)
    elif ativo_param == "nao":
        base_qualidade = base_qualidade.filter(ativo=False)
    contadores_qualidade = {
        "nome_site": base_qualidade.filter(nome_site="").count(),
        "endereco": base_qualidade.filter(endereco="").count(),
        "cidade": base_qualidade.filter(cidade="").count(),
        "uf": base_qualidade.filter(uf="").count(),
        "regional": base_qualidade.filter(regional="").count(),
        "tipo_imovel": base_qualidade.filter(tipo_imovel="").count(),
    }
    total_com_problema = sum(1 for v in contadores_qualidade.values() if v > 0)

    paginator = Paginator(queryset, ATIVOS_POR_PAGINA)
    page_obj = paginator.get_page(request.GET.get("page"))

    querystring_filtros = _montar_querystring(request)

    context = {
        "page_obj": page_obj,
        "ativos": page_obj.object_list,
        "total": total,
        "filtros": {
            "q": q,
            "cidade": cidade,
            "uf": uf,
            "regional": regional,
            "ativo": ativo_param,
            "sem": sem,
        },
        "contadores_qualidade": contadores_qualidade,
        "tem_problemas_de_qualidade": total_com_problema > 0,
        "querystring_filtros": querystring_filtros,
    }
    return render(request, "chamados/ativos_list.html", context)


def ativo_create(request):
    """Cadastro de novo ativo."""
    if request.method == "POST":
        form = AtivoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Ativo cadastrado com sucesso.")
            return redirect("chamados:ativos_list")
    else:
        form = AtivoForm()
    return render(
        request,
        "chamados/ativo_form.html",
        {"form": form, "modo": "novo", "titulo": "Novo ativo"},
    )


def ativos_import(request):
    """Importa a planilha do parque imobiliario para a Matriz de Ativos."""
    resultado = None
    if request.method == "POST":
        form = AtivosImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resultado = importar_ativos_excel(
                    form.cleaned_data["arquivo"],
                    usuario=request.user,
                )
            except EXCECOES_PLANILHA_INVALIDA:
                # Arquivo .xlsx ilegível pelo openpyxl (zip quebrado, formato
                # diferente etc.). Nenhuma gravação ocorreu porque a leitura
                # da planilha acontece antes do bloco transacional.
                logger.warning(
                    "Importação de Ativos abortada por arquivo inválido.",
                    exc_info=True,
                )
                messages.error(request, MENSAGEM_PLANILHA_INVALIDA)
            else:
                messages.success(request, "Importação concluída.")
    else:
        form = AtivosImportForm()

    return render(
        request,
        "chamados/ativos_import.html",
        {"form": form, "resultado": resultado},
    )


def ativo_update(request, pk):
    """Edicao de ativo existente."""
    ativo = get_object_or_404(Ativo, pk=pk)
    if request.method == "POST":
        form = AtivoForm(request.POST, instance=ativo)
        if form.is_valid():
            form.save()
            messages.success(request, "Ativo atualizado com sucesso.")
            return redirect("chamados:ativos_list")
    else:
        form = AtivoForm(instance=ativo)
    return render(
        request,
        "chamados/ativo_form.html",
        {
            "form": form,
            "modo": "editar",
            "titulo": f"Editar ativo {ativo.ativo_prisma}",
            "ativo": ativo,
        },
    )


def ativo_detail(request, pk):
    """Detalhe do ativo."""
    ativo = get_object_or_404(Ativo, pk=pk)
    return render(request, "chamados/ativo_detail.html", {"ativo": ativo})


def fornecedores_list(request):
    """Listagem operacional de supervisores/fornecedores."""
    q = request.GET.get("q", "").strip()
    empresa = request.GET.get("empresa", "").strip()
    estado = request.GET.get("estado", "").strip().upper()
    ativo_param = request.GET.get("ativo", "todos").strip() or "todos"

    queryset = Fornecedor.objects.all()
    if q:
        queryset = queryset.filter(
            Q(nome__icontains=q)
            | Q(telefone__icontains=q)
            | Q(email__icontains=q)
            | Q(empresa__icontains=q)
            | Q(estados_atendidos__icontains=q)
        )
    if empresa:
        queryset = queryset.filter(empresa__icontains=empresa)
    if estado:
        queryset = queryset.filter(estados_atendidos__icontains=estado)
    if ativo_param == "sim":
        queryset = queryset.filter(ativo=True)
    elif ativo_param == "nao":
        queryset = queryset.filter(ativo=False)

    total = queryset.count()
    ativos_count = queryset.filter(ativo=True).count()
    inativos_count = total - ativos_count
    empresas_count = (
        queryset.exclude(empresa="").values("empresa").distinct().count()
    )

    paginator = Paginator(queryset, FORNECEDORES_POR_PAGINA)
    page_obj = paginator.get_page(request.GET.get("page"))

    querystring_filtros = _montar_querystring(request)

    return render(
        request,
        "chamados/fornecedores_list.html",
        {
            "page_obj": page_obj,
            "fornecedores": page_obj.object_list,
            "total": total,
            "stats": {
                "ativos": ativos_count,
                "inativos": inativos_count,
                "empresas": empresas_count,
            },
            "filtros": {
                "q": q,
                "empresa": empresa,
                "estado": estado,
                "ativo": ativo_param,
            },
            "querystring_filtros": querystring_filtros,
        },
    )


def fornecedores_import(request):
    """Importa supervisores/fornecedores via planilha .xlsx."""
    resultado = None
    if request.method == "POST":
        form = FornecedoresImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resultado = importar_fornecedores_excel(
                    form.cleaned_data["arquivo"],
                    usuario=request.user,
                )
            except EXCECOES_PLANILHA_INVALIDA:
                logger.warning(
                    "Importação de Fornecedores abortada por arquivo inválido.",
                    exc_info=True,
                )
                messages.error(request, MENSAGEM_PLANILHA_INVALIDA)
            else:
                messages.success(request, "Importação concluída.")
    else:
        form = FornecedoresImportForm()

    return render(
        request,
        "chamados/fornecedores_import.html",
        {"form": form, "resultado": resultado},
    )


def fornecedores_template_download(request):
    """Gera planilha .xlsx em branco com as colunas esperadas."""
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Supervisores"
    planilha.append(list(COLUNAS_MODELO_FORNECEDOR))

    cabecalho_font = Font(bold=True)
    for celula in planilha[1]:
        celula.font = cabecalho_font

    larguras = [28, 18, 32, 28, 28]
    for indice, largura in enumerate(larguras, start=1):
        planilha.column_dimensions[planilha.cell(row=1, column=indice).column_letter].width = largura

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        'attachment; filename="modelo_supervisores.xlsx"'
    )
    return response


def consolidado_obras(request):
    """Compatibilidade: a URL antiga agora redireciona para a tela de Obras."""
    return redirect("chamados:obras_list")


def fornecedor_create(request):
    """Cadastro operacional de supervisor."""
    if request.method == "POST":
        form = FornecedorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Supervisor cadastrado com sucesso.")
            return redirect("chamados:fornecedores_list")
    else:
        form = FornecedorForm(initial={"ativo": True})
    return render(
        request,
        "chamados/fornecedor_form.html",
        {"form": form, "titulo": "Novo Supervisor"},
    )


def fornecedor_update(request, pk):
    """Edicao operacional de supervisor."""
    fornecedor = get_object_or_404(Fornecedor, pk=pk)
    if request.method == "POST":
        form = FornecedorForm(request.POST, instance=fornecedor)
        if form.is_valid():
            form.save()
            messages.success(request, "Supervisor atualizado com sucesso.")
            return redirect("chamados:fornecedores_list")
    else:
        form = FornecedorForm(instance=fornecedor)
    return render(
        request,
        "chamados/fornecedor_form.html",
        {"form": form, "titulo": f"Editar supervisor {fornecedor.nome}"},
    )


def fornecedor_detail(request, pk):
    """Detalhe operacional do fornecedor."""
    fornecedor = get_object_or_404(Fornecedor, pk=pk)
    chamados = fornecedor.chamados.select_related("ativo").order_by(
        "-data_abertura", "-criado_em"
    )
    return render(
        request,
        "chamados/fornecedor_detail.html",
        {
            "fornecedor": fornecedor,
            "quantidade_chamados": chamados.count(),
            "ultimos_chamados": chamados[:10],
        },
    )


def fornecedor_delete(request, pk):
    fornecedor = get_object_or_404(Fornecedor, pk=pk)
    if request.method == "POST":
        nome = fornecedor.nome
        if fornecedor.chamados.exists():
            messages.error(
                request,
                f"Não é possível excluir '{nome}' pois possui chamados vinculados. "
                "Considere inativá-lo em vez de excluir."
            )
            return redirect("chamados:fornecedores_list")
        fornecedor.delete()
        messages.success(request, f"Supervisor '{nome}' excluído com sucesso.")
        return redirect("chamados:fornecedores_list")
    return render(
        request,
        "chamados/fornecedor_confirm_delete.html",
        {"fornecedor": fornecedor},
    )


@require_http_methods(["GET", "POST"])
def chamado_create(request):
    """Cadastro de novo chamado emergencial.

    Aceita querystring ?ativo=<pk> para pre-selecionar o ativo.
    """
    ativo_param = request.GET.get("ativo") or request.POST.get("ativo")
    ativo_preselecionado = None
    if ativo_param:
        try:
            ativo_preselecionado = Ativo.objects.filter(
                pk=int(ativo_param), ativo=True
            ).first()
        except (TypeError, ValueError):
            ativo_preselecionado = None

    if request.method == "POST":
        form = ChamadoForm(request.POST)
        if form.is_valid():
            chamado = services.criar_chamado(form, usuario=request.user)
            messages.success(request, "Chamado criado com sucesso.")
            return redirect("chamados:chamado_detail", pk=chamado.pk)
    else:
        initial = {}
        if ativo_preselecionado is not None:
            initial["ativo"] = ativo_preselecionado
        if request.user.is_authenticated:
            nome_logado = request.user.get_full_name() or request.user.username
            if nome_logado:
                initial["command_center"] = nome_logado
        form = ChamadoForm(initial=initial)

    obras_ativas = []
    if ativo_preselecionado is not None:
        obras_ativas = list(_obras_ativas_qs(ativo_preselecionado))

    return render(
        request,
        "chamados/chamado_form.html",
        {
            "form": form,
            "ativo_preselecionado": ativo_preselecionado,
            "obras_ativas": obras_ativas,
            "titulo": "Novo chamado",
        },
    )


def chamados_list(request):
    """Listagem simples de chamados com busca opcional."""
    q = request.GET.get("q", "").strip()
    queryset = Chamado.objects.select_related("ativo", "fornecedor")

    if q:
        queryset = queryset.filter(
            Q(numero_os__icontains=q)
            | Q(ativo__ativo_prisma__icontains=q)
            | Q(ativo__nome_site__icontains=q)
            | Q(ativo__cidade__icontains=q)
            | Q(ativo__regional__icontains=q)
        )

    total = queryset.count()

    paginator = Paginator(queryset, CHAMADOS_POR_PAGINA)
    page_obj = paginator.get_page(request.GET.get("page"))

    querystring_filtros = _montar_querystring(request)

    return render(
        request,
        "chamados/chamados_list.html",
        {
            "page_obj": page_obj,
            "chamados": page_obj.object_list,
            "total": total,
            "filtros": {"q": q},
            "querystring_filtros": querystring_filtros,
        },
    )


def chamado_detail(request, pk):
    """Detalhe completo de um chamado."""
    chamado = get_object_or_404(
        Chamado.objects.select_related(
            "ativo", "fornecedor", "criado_por", "atualizado_por"
        ).prefetch_related(
            "atualizacoes__evidencias",
            "atualizacoes__criado_por",
        ),
        pk=pk,
    )
    atualizacoes = chamado.atualizacoes.all()
    return render(
        request,
        "chamados/chamado_detail.html",
        {
            "chamado": chamado,
            "atualizacoes": atualizacoes,
            "tempo_aberto": _formatar_tempo_aberto(chamado.data_abertura),
            "sla_class": (
                _classificar_sla(chamado.data_abertura)
                if chamado.status in (
                    StatusChamado.ABERTO.value,
                    StatusChamado.PENDENTE.value,
                )
                else None
            ),
            "mostrar_sla": chamado.status in (
                StatusChamado.ABERTO.value,
                StatusChamado.PENDENTE.value,
            ),
        },
    )


def atualizar_report_list(request):
    """Listagem de chamados para selecao do report a atualizar.

    A ordenacao empurra chamados ja finalizados (Concluido, Cancelado e
    Nao emergencial) para o fim da lista — nenhum deles exige acao do
    operador, entao nao precisam concorrer pela atencao na esteira.
    Dentro de cada grupo mantemos a ordem padrao (-data_abertura).
    """
    queryset, filtros = _filtrar_chamados_atualizar_report(request.GET)
    status_finalizados = [
        StatusChamado.CONCLUIDO,
        StatusChamado.CANCELADO,
        StatusChamado.NAO_EMERGENCIAL,
    ]
    queryset = queryset.annotate(
        _ordem_status=Case(
            When(status__in=status_finalizados, then=Value(1)),
            default=Value(0),
            output_field=IntegerField(),
        )
    ).order_by("_ordem_status", "-data_abertura", "-criado_em")

    total = queryset.count()
    contadores_status = queryset.aggregate(
        abertos=Count("id", filter=Q(status=StatusChamado.ABERTO)),
        pendentes=Count("id", filter=Q(status=StatusChamado.PENDENTE)),
        concluidos=Count("id", filter=Q(status=StatusChamado.CONCLUIDO)),
        cancelados=Count("id", filter=Q(status=StatusChamado.CANCELADO)),
        nao_emergenciais=Count("id", filter=Q(status=StatusChamado.NAO_EMERGENCIAL)),
        primeiro_report_pendente=Count("id", filter=Q(atualizacoes__isnull=True)),
    )
    contadores = {
        "total": total,
        "abertos": contadores_status["abertos"],
        "pendentes": contadores_status["pendentes"],
        "concluidos": contadores_status["concluidos"],
        "cancelados": contadores_status["cancelados"],
        "nao_emergenciais": contadores_status["nao_emergenciais"],
        "primeiro_report_pendente": contadores_status["primeiro_report_pendente"],
    }

    paginator = Paginator(queryset, ATUALIZAR_REPORT_POR_PAGINA)
    page_obj = paginator.get_page(request.GET.get("page"))

    chamados = list(page_obj.object_list)
    for chamado in chamados:
        # Usa o cache do prefetch_related em vez de disparar exists() por chamado
        chamado.primeiro_report = len(chamado.atualizacoes.all()) == 0
        if chamado.status in (
            StatusChamado.ABERTO.value,
            StatusChamado.PENDENTE.value,
        ):
            chamado.sla_class = _classificar_sla(chamado.data_abertura)
            chamado.tempo_aberto_label = _formatar_tempo_aberto(chamado.data_abertura)
        else:
            chamado.sla_class = None
            chamado.tempo_aberto_label = None

    querystring_filtros = _montar_querystring(request)

    base_qs = _montar_querystring(request, excluir=("page",))
    sep = "&" if base_qs else ""
    querystring_com_primeiro_report = (
        base_qs + (f"{sep}primeiro_report=sim"
                   if "primeiro_report" not in request.GET else "")
    )
    querystring_sem_primeiro_report = _montar_querystring(
        request, excluir=("page", "primeiro_report")
    )

    # Dropdowns refletem apenas regionais/cidades que existem na esteira
    # de chamados (nao o parque imobiliario inteiro).
    _cache_key_regionais = "dropdown_regionais_chamados"
    _cache_key_cidades   = "dropdown_cidades_chamados"
    TTL = 120  # segundos

    regionais_disponiveis = cache.get(_cache_key_regionais)
    if regionais_disponiveis is None:
        regionais_disponiveis = list(
            Chamado.objects.exclude(ativo__regional="")
            .exclude(ativo__regional__isnull=True)
            .values_list("ativo__regional", flat=True)
            .distinct()
            .order_by("ativo__regional")
        )
        cache.set(_cache_key_regionais, regionais_disponiveis, TTL)

    cidades_disponiveis = cache.get(_cache_key_cidades)
    if cidades_disponiveis is None:
        cidades_disponiveis = list(
            Chamado.objects.exclude(ativo__cidade="")
            .exclude(ativo__cidade__isnull=True)
            .values_list("ativo__cidade", flat=True)
            .distinct()
            .order_by("ativo__cidade")
        )
        cache.set(_cache_key_cidades, cidades_disponiveis, TTL)

    return render(
        request,
        "chamados/atualizar_report_list.html",
        {
            "page_obj": page_obj,
            "chamados": chamados,
            "total": total,
            "contadores": contadores,
            "fornecedores": Fornecedor.objects.filter(ativo=True),
            "status_choices": StatusChamado.choices,
            "regionais_disponiveis": regionais_disponiveis,
            "cidades_disponiveis": cidades_disponiveis,
            "filtros": filtros,
            "querystring_filtros": querystring_filtros,
            "querystring_com_primeiro_report": querystring_com_primeiro_report,
            "querystring_sem_primeiro_report": querystring_sem_primeiro_report,
        },
    )


def atualizar_report_exportar_excel(request):
    """Exporta a listagem de Atualizar Report respeitando os filtros da tela."""
    queryset, _filtros = _filtrar_chamados_atualizar_report(request.GET)
    queryset = queryset.select_related(
        "ativo", "fornecedor"
    ).prefetch_related("atualizacoes")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Atualizar Report"
    sheet.append(COLUNAS_EXPORTAR_ATUALIZAR_REPORT)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:{sheet.cell(row=1, column=len(COLUNAS_EXPORTAR_ATUALIZAR_REPORT)).coordinate}"
    )

    for chamado in queryset:
        atualizacoes = list(chamado.atualizacoes.all())
        ultima_atualizacao = atualizacoes[-1] if atualizacoes else None
        sheet.append(
            [
                chamado.numero_os,
                chamado.get_status_display(),
                _formatar_data_excel(chamado.data_abertura),
                chamado.ativo.ativo_prisma,
                chamado.ativo.nome_site,
                chamado.ativo.endereco,
                chamado.ativo.cidade,
                chamado.ativo.uf,
                chamado.ativo.regional,
                chamado.ativo.tipo_imovel or "-",
                chamado.ativo.tipo_site_sla or "-",
                chamado.ativo.lider_coordenacao or "-",
                chamado.fornecedor.nome if chamado.fornecedor_id else "-",
                chamado.solicitante or "-",
                chamado.contato_solicitante or "-",
                chamado.command_center or "-",
                chamado.denominacao or "-",
                chamado.acao_tomada or "-",
                chamado.detalhamento_situacao or "-",
                "Não" if atualizacoes else "Sim",
                len(atualizacoes),
                ultima_atualizacao.texto_atualizacao if ultima_atualizacao else "-",
                _formatar_data_excel(ultima_atualizacao.criado_em)
                if ultima_atualizacao
                else "-",
                ultima_atualizacao.get_status_resultante_display()
                if ultima_atualizacao
                else "-",
                _formatar_data_excel(chamado.criado_em),
                _formatar_data_excel(chamado.atualizado_em),
            ]
        )

    for column_cells in sheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        sheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_length + 2, 12), 45
        )

    arquivo = BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)

    nome_arquivo = timezone.localtime().strftime("atualizar_report_%Y%m%d_%H%M%S.xlsx")
    response = HttpResponse(
        arquivo.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response


def _formatar_tempo_aberto(data_abertura) -> str:
    """Retorna um rotulo compacto tipo '3min', '5h', '1d 4h' para o cabecalho."""
    if data_abertura is None:
        return ""
    delta = timezone.now() - data_abertura
    total_minutos = int(delta.total_seconds() // 60)
    if total_minutos < 0:
        total_minutos = 0
    if total_minutos < 60:
        return f"{total_minutos}min"
    horas_totais = total_minutos // 60
    if horas_totais < 24:
        return f"{horas_totais}h"
    dias = horas_totais // 24
    horas = horas_totais % 24
    if horas:
        return f"{dias}d {horas}h"
    return f"{dias}d"


def _classificar_sla(data_abertura) -> str:
    """Retorna 'ok', 'alerta' ou 'critico' baseado no tempo aberto."""
    if data_abertura is None:
        return "ok"
    delta = timezone.now() - data_abertura
    horas = delta.total_seconds() / 3600
    if horas <= 4:
        return "ok"
    if horas <= 12:
        return "alerta"
    return "critico"


def atualizar_report_form(request, pk):
    """Formulario de report com persistencia delegada ao service."""
    chamado = get_object_or_404(
        Chamado.objects.select_related("ativo", "fornecedor").prefetch_related(
            "atualizacoes"
        ),
        pk=pk,
    )
    primeiro_report = services.is_primeiro_report(chamado)

    if request.method == "POST":
        form = RegistrarReportForm(request.POST)
        if form.is_valid():
            status_resultante = form.cleaned_data.get("status_resultante") or None
            if not primeiro_report and status_resultante is None:
                form.add_error(
                    "status_resultante",
                    "Status resultante é obrigatório para atualizações posteriores.",
                )
            else:
                evidencias = request.FILES.getlist("evidencias")
                try:
                    services.registrar_report(
                        chamado=chamado,
                        texto_atualizacao=form.cleaned_data["texto_atualizacao"],
                        status_resultante=status_resultante,
                        usuario=request.user,
                        evidencias=evidencias,
                    )
                except ValueError as exc:
                    form.add_error(None, str(exc))
                else:
                    messages.success(request, "Report registrado com sucesso.")
                    return redirect("chamados:atualizar_report_form", pk=chamado.pk)
    else:
        form = RegistrarReportForm()

    texto_whatsapp = services.gerar_texto_whatsapp(chamado)

    return render(
        request,
        "chamados/atualizar_report_form.html",
        {
            "chamado": chamado,
            "form": form,
            "primeiro_report": primeiro_report,
            "texto_whatsapp": texto_whatsapp,
            "tempo_aberto": _formatar_tempo_aberto(chamado.data_abertura),
        },
    )


# === Obras =================================================================

def obras_list(request):
    """Listagem de obras com filtros simples."""
    q = request.GET.get("q", "").strip()
    situacao = request.GET.get("situacao", "todas").strip() or "todas"

    queryset = Obra.objects.select_related("ativo")
    if q:
        queryset = queryset.filter(
            Q(ativo__ativo_prisma__icontains=q)
            | Q(ativo__nome_site__icontains=q)
            | Q(ativo__cidade__icontains=q)
            | Q(ativo__regional__icontains=q)
            | Q(descricao__icontains=q)
            | Q(responsavel__icontains=q)
        )

    hoje = date.today()
    if situacao == "em_andamento":
        queryset = queryset.filter(
            ativa=True,
            data_fim_real__isnull=True,
            data_inicio__lte=hoje,
            data_fim_planejada__gte=hoje,
        )
    elif situacao == "atrasada":
        queryset = queryset.filter(
            ativa=True,
            data_fim_real__isnull=True,
            data_fim_planejada__lt=hoje,
        )
    elif situacao == "planejada":
        queryset = queryset.filter(
            ativa=True,
            data_fim_real__isnull=True,
            data_inicio__gt=hoje,
        )
    elif situacao == "concluida":
        queryset = queryset.filter(data_fim_real__isnull=False)

    total = queryset.count()
    hoje_obras = date.today()
    stats_obras = Obra.objects.aggregate(
        em_andamento=Count("id", filter=Q(
            ativa=True,
            data_fim_real__isnull=True,
            data_inicio__lte=hoje_obras,
            data_fim_planejada__gte=hoje_obras,
        )),
        atrasadas=Count("id", filter=Q(
            ativa=True,
            data_fim_real__isnull=True,
            data_fim_planejada__lt=hoje_obras,
        )),
        planejadas=Count("id", filter=Q(
            ativa=True,
            data_fim_real__isnull=True,
            data_inicio__gt=hoje_obras,
        )),
        concluidas=Count("id", filter=Q(data_fim_real__isnull=False)),
    )
    em_andamento_count = stats_obras["em_andamento"]
    atrasadas_count    = stats_obras["atrasadas"]
    planejadas_count   = stats_obras["planejadas"]
    concluidas_count   = stats_obras["concluidas"]

    paginator = Paginator(queryset, OBRAS_POR_PAGINA)
    page_obj = paginator.get_page(request.GET.get("page"))

    querystring_filtros = _montar_querystring(request)

    return render(
        request,
        "chamados/obras_list.html",
        {
            "page_obj": page_obj,
            "obras": page_obj.object_list,
            "total": total,
            "stats": {
                "em_andamento": em_andamento_count,
                "atrasadas": atrasadas_count,
                "planejadas": planejadas_count,
                "concluidas": concluidas_count,
            },
            "filtros": {"q": q, "situacao": situacao},
            "querystring_filtros": querystring_filtros,
        },
    )


@require_http_methods(["GET", "POST"])
def obra_create(request):
    """Cadastro individual de obra."""
    ativo_param = request.GET.get("ativo")
    initial = {}
    if ativo_param:
        try:
            ativo = Ativo.objects.filter(pk=int(ativo_param), ativo=True).first()
            if ativo:
                initial["ativo"] = ativo
        except (TypeError, ValueError):
            pass

    if request.method == "POST":
        form = ObraForm(request.POST)
        if form.is_valid():
            obra = form.save()
            messages.success(request, "Obra cadastrada com sucesso.")
            return redirect("chamados:obra_detail", pk=obra.pk)
    else:
        form = ObraForm(initial=initial)

    return render(
        request,
        "chamados/obra_form.html",
        {"form": form, "modo": "novo"},
    )


@require_http_methods(["GET", "POST"])
def obra_update(request, pk):
    """Edição de obra existente."""
    obra = get_object_or_404(Obra, pk=pk)
    if request.method == "POST":
        form = ObraForm(request.POST, instance=obra)
        if form.is_valid():
            form.save()
            messages.success(request, "Obra atualizada com sucesso.")
            return redirect("chamados:obra_detail", pk=obra.pk)
    else:
        form = ObraForm(instance=obra)
    return render(
        request,
        "chamados/obra_form.html",
        {
            "form": form,
            "modo": "editar",
            "obra": obra,
        },
    )


def obra_detail(request, pk):
    """Detalhe da obra."""
    obra = get_object_or_404(Obra.objects.select_related("ativo"), pk=pk)

    hoje = date.today()
    total_dias = (obra.data_fim_planejada - obra.data_inicio).days or 1
    dias_passados = (min(hoje, obra.data_fim_planejada) - obra.data_inicio).days
    progresso = max(0, min(100, int(dias_passados / total_dias * 100)))

    if obra.data_fim_real:
        dias_restantes = None
        dias_atraso = None
    elif hoje > obra.data_fim_planejada:
        dias_restantes = None
        dias_atraso = (hoje - obra.data_fim_planejada).days
    else:
        dias_restantes = (obra.data_fim_planejada - hoje).days
        dias_atraso = None

    alerta_proximidade = dias_restantes is not None and dias_restantes <= 7

    return render(
        request,
        "chamados/obra_detail.html",
        {
            "obra": obra,
            "progresso": progresso,
            "dias_restantes": dias_restantes,
            "dias_atraso": dias_atraso,
            "alerta_proximidade": alerta_proximidade,
        },
    )


@require_http_methods(["GET", "POST"])
def obra_concluir(request, pk):
    obra = get_object_or_404(Obra, pk=pk)
    if request.method == "POST":
        concluida = services.concluir_obra(obra)
        if concluida:
            messages.success(
                request,
                f"Obra em '{obra.ativo.nome_site}' marcada como concluída.",
            )
        else:
            messages.warning(request, "Esta obra já foi concluída.")
        return redirect("chamados:obra_detail", pk=obra.pk)
    return render(
        request,
        "chamados/obra_concluir_confirm.html",
        {"obra": obra},
    )


def obras_import(request):
    """Importa obras via planilha .xlsx."""
    resultado = None
    if request.method == "POST":
        form = ObrasImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resultado = importar_obras_excel(
                    form.cleaned_data["arquivo"],
                    usuario=request.user,
                )
            except EXCECOES_PLANILHA_INVALIDA:
                logger.warning(
                    "Importação de Obras abortada por arquivo inválido.",
                    exc_info=True,
                )
                messages.error(request, MENSAGEM_PLANILHA_INVALIDA)
            else:
                messages.success(request, "Importação concluída.")
    else:
        form = ObrasImportForm()

    return render(
        request,
        "chamados/obras_import.html",
        {"form": form, "resultado": resultado},
    )


def obras_template_download(request):
    """Gera planilha .xlsx modelo em branco para o cadastro de obras."""
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Obras"
    planilha.append(list(COLUNAS_MODELO_OBRA))

    cabecalho_font = Font(bold=True)
    for celula in planilha[1]:
        celula.font = cabecalho_font

    larguras = [16, 30, 36, 22, 6, 18, 22, 50, 14, 18, 26, 30]
    for indice, largura in enumerate(larguras, start=1):
        planilha.column_dimensions[
            planilha.cell(row=1, column=indice).column_letter
        ].width = largura

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        'attachment; filename="modelo_obras.xlsx"'
    )
    return response


@require_http_methods(["GET"])
def ativos_autocomplete(request):
    """Endpoint JSON usado pelo combobox de Ativo Prisma no Novo Chamado.

    Retorna uma janela paginada de ativos ativos que casem com o termo digitado
    em qualquer um dos campos mais relevantes (prisma, nome do site, cidade, UF).
    Aceita ?offset=N para scroll infinito; tamanho da página = 30.
    """
    PAGINA = 30
    termo = (request.GET.get("q") or "").strip()
    try:
        offset = max(0, int(request.GET.get("offset") or 0))
    except (TypeError, ValueError):
        offset = 0

    queryset = Ativo.objects.filter(ativo=True)

    if termo:
        queryset = queryset.filter(
            Q(ativo_prisma__icontains=termo)
            | Q(nome_site__icontains=termo)
            | Q(cidade__icontains=termo)
            | Q(uf__icontains=termo)
        )

    queryset = queryset.order_by("ativo_prisma")
    total = queryset.count()
    pagina_qs = queryset[offset : offset + PAGINA]

    resultados = []
    for ativo in pagina_qs:
        partes_sub = []
        if ativo.nome_site:
            partes_sub.append(ativo.nome_site)
        if ativo.cidade and ativo.uf:
            partes_sub.append(f"{ativo.cidade}/{ativo.uf}")
        elif ativo.cidade:
            partes_sub.append(ativo.cidade)
        elif ativo.uf:
            partes_sub.append(ativo.uf)
        if ativo.regional:
            partes_sub.append(ativo.regional)
        resultados.append({
            "id": ativo.id,
            "label": ativo.ativo_prisma or f"Ativo {ativo.id}",
            "sub": " · ".join(partes_sub),
        })

    return JsonResponse({
        "results": resultados,
        "total": total,
        "offset": offset,
        "has_more": offset + len(resultados) < total,
    })


@require_http_methods(["GET"])
def ativo_obras_ativas(request, pk):
    """Endpoint JSON consumido pelo formulário de Novo Chamado.

    Retorna as obras em andamento (ou atrasadas) do ativo informado, para que
    a UI alerte o usuário sobre intervenções já em curso no endereço.
    """
    ativo = get_object_or_404(Ativo, pk=pk)
    hoje = date.today()
    obras_qs = Obra.objects.filter(
        ativo=ativo,
        ativa=True,
        data_fim_real__isnull=True,
        data_inicio__lte=hoje,
    ).order_by("-data_inicio")

    obras = []
    for obra in obras_qs:
        if obra.data_fim_planejada < hoje:
            situacao = "atrasada"
        else:
            situacao = "em_andamento"
        obras.append({
            "id": obra.id,
            "descricao": obra.descricao,
            "data_inicio": obra.data_inicio.isoformat(),
            "data_fim_planejada": obra.data_fim_planejada.isoformat(),
            "data_inicio_br": obra.data_inicio.strftime("%d/%m/%Y"),
            "data_fim_planejada_br": obra.data_fim_planejada.strftime("%d/%m/%Y"),
            "responsavel": obra.responsavel,
            "situacao": situacao,
            "situacao_label": "Atrasada" if situacao == "atrasada" else "Em andamento",
            "detalhe_url": reverse("chamados:obra_detail", args=[obra.id]),
        })

    return JsonResponse({
        "ativo_id": ativo.id,
        "ativo_prisma": ativo.ativo_prisma,
        "nome_site": ativo.nome_site,
        "endereco": ativo.endereco,
        "cidade": ativo.cidade,
        "uf": ativo.uf,
        "regional": ativo.regional,
        "tipo_imovel": ativo.tipo_imovel,
        "lider_coordenacao": ativo.lider_coordenacao,
        "tipo_site_sla": ativo.tipo_site_sla,
        "hoje": hoje.isoformat(),
        "obras": obras,
    })


# ============================================================
#  Autenticacao (login / logout)
# ============================================================

@require_http_methods(["GET", "POST"])
def login_view(request):
    """Tela de login por nome + sobrenome + senha."""
    if request.user.is_authenticated:
        return redirect("chamados:home")

    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            user = authenticate(
                request,
                first_name=form.cleaned_data["first_name"],
                last_name=form.cleaned_data["last_name"],
                password=form.cleaned_data["password"],
            )
            if user is not None:
                auth_login(request, user)
                next_url = (
                    request.POST.get("next") or request.GET.get("next") or ""
                ).strip()
                if next_url and next_url.startswith("/") and not next_url.startswith("//"):
                    return redirect(next_url)
                return redirect("chamados:home")
            form.add_error(None, "Nome, sobrenome ou senha incorretos.")
    else:
        form = LoginForm()

    return render(
        request,
        "chamados/login.html",
        {"form": form, "next": request.GET.get("next", "")},
    )


@require_POST
def logout_view(request):
    auth_logout(request)
    return redirect("chamados:login")


@require_http_methods(["GET", "POST"])
def register_view(request):
    """Auto-cadastro: nome + sobrenome + e-mail + senha (com confirmacao)."""
    if request.user.is_authenticated:
        return redirect("chamados:home")

    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = services.registrar_usuario(
                first_name=form.cleaned_data["first_name"],
                last_name=form.cleaned_data["last_name"],
                email=form.cleaned_data["email"],
                password=form.cleaned_data["password"],
            )
            # Como ha mais de um backend configurado, especificamos qual
            # backend "logou" o usuario para o auth_login.
            auth_login(
                request,
                user,
                backend="chamados.backends.NomeSobrenomeBackend",
            )
            messages.success(
                request, f"Conta criada com sucesso. Bem-vindo, {user.first_name}!"
            )
            return redirect("chamados:home")
    else:
        form = RegisterForm()

    return render(request, "chamados/register.html", {"form": form})
